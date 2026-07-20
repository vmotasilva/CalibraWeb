"""Views para gerenciamento de listas de presença."""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.urls import reverse
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import logging
import pandas as pd
import os

logger = logging.getLogger(__name__)

from django.http import FileResponse
from django.conf import settings

from procedures.models import (
    ListaPresenca, RegistroTreinamento, Procedimento, ParticipanteExterno, 
    PlanejamentoTreinamento, TemplateListaPresenca, MapeamentoCampoListaPresenca
)
from procedures.forms.lista_presenca_forms import (
    ListaPresencaForm, RegistroTreinamentoFormSet, ImportacaoTreinamentoForm, ParticipanteExternoForm
)
from procedures.utils.name_matching import tentar_linkar_colaborador
from rh.models import Colaborador


@login_required
def lista_presenca_list_view(request):
    """Lista todas as listas de presença com filtros e remoção em massa."""
    # Processar remoção em massa
    if request.method == 'POST' and 'delete_selected' in request.POST:
        ids_selecionados = request.POST.getlist('lista_ids')
        if ids_selecionados:
            count = ListaPresenca.objects.filter(id__in=ids_selecionados).delete()[0]
            messages.success(request, f'✅ {count} lista(s) removida(s) com sucesso!')
        else:
            messages.warning(request, '⚠️ Nenhuma lista selecionada.')
        return redirect('procedures:lista_presenca_list')
    
    # Buscar listas
    listas = ListaPresenca.objects.all()
    
    # Aplicar filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    instrutor_id = request.GET.get('instrutor')
    busca = request.GET.get('busca')
    
    if data_inicio:
        listas = listas.filter(data_sessao__gte=data_inicio)
    if data_fim:
        listas = listas.filter(data_sessao__lte=data_fim)
    if instrutor_id:
        listas = listas.filter(instrutor_id=instrutor_id)
    if busca:
        from django.db.models import Q
        listas = listas.filter(
            Q(codigo__icontains=busca) |
            Q(titulo__icontains=busca) |
            Q(local__icontains=busca) |
            Q(instrutor_nome__icontains=busca)  # Buscar também por instrutor nome livre
        )
    
    listas = listas.order_by('-data_sessao', '-codigo')
    
    # Adicionar contadores para cada lista
    for lista in listas:
        registros = lista.registros.all()
        # Contar participantes distintos (colaboradores + externos)
        colaboradores = registros.filter(colaborador__isnull=False).values('colaborador').distinct().count()
        externos = registros.filter(participante_externo__isnull=False).values('participante_externo').distinct().count()
        lista.total_participantes = colaboradores + externos
        # Contar procedimentos (pode haver registros sem procedimento)
        lista.total_procedimentos = registros.filter(procedimento__isnull=False).values('procedimento').distinct().count()
        lista.total_registros = registros.count()
        
        # Obter procedimentos únicos
        procs_set = set()
        procs = []
        for reg in registros.select_related('procedimento'):
            if reg.procedimento and reg.procedimento not in procs_set:
                procs_set.add(reg.procedimento)
                procs.append(reg.procedimento)
        lista.procedimentos_lista = procs
    
    # Buscar instrutores para o filtro
    instrutores = Colaborador.objects.filter(
        listas_presenca_como_instrutor__isnull=False
    ).distinct().order_by('nome_completo')
    
    # Calcular totais
    total_listas = listas.count()
    total_registros = RegistroTreinamento.objects.filter(lista_presenca__in=listas).count()
    
    context = {
        'listas': listas,
        'instrutores': instrutores,
        'total_listas': total_listas,
        'total_registros': total_registros,
        'filtros': {
            'data_inicio': data_inicio or '',
            'data_fim': data_fim or '',
            'instrutor': instrutor_id or '',
            'busca': busca or '',
        }
    }
    return render(request, 'procedures/lista_presenca_list.html', context)


@login_required
def lista_presenca_create_view(request):
    """Cria nova lista de presença com múltiplos registros."""
    if request.method == 'POST':
        form = ListaPresencaForm(request.POST)
        formset = RegistroTreinamentoFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                # Salvar a lista de presença
                lista = form.save(commit=False)
                lista.criado_por = request.user
                lista.save()
                
                # Salvar os registros de treinamento
                registros = formset.save(commit=False)
                for registro in registros:
                    registro.lista_presenca = lista
                    # Se data_treinamento não foi preenchida, usar data da sessão
                    if not registro.data_treinamento:
                        registro.data_treinamento = lista.data_sessao
                    
                    # Tentar linkar colaborador usando nome ou FK
                    if registro.colaborador_nome or registro.colaborador:
                        colab_linkado = tentar_linkar_colaborador(
                            nome_texto=registro.colaborador_nome,
                            colaborador_fk=registro.colaborador,
                            threshold=0.85
                        )
                        # Salvar FK se conseguiu match, mesmo que não tenha FK manual
                        if colab_linkado:
                            registro.colaborador = colab_linkado
                    
                    registro.save()
                
                # Deletar registros marcados para exclusão
                for obj in formset.deleted_objects:
                    obj.delete()
                
                messages.success(request, f'Lista de presença {lista.codigo} criada com sucesso!')
                return redirect('procedures:lista_presenca_detail', pk=lista.pk)
    else:
        form = ListaPresencaForm()
        formset = RegistroTreinamentoFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'action': 'create',
    }
    return render(request, 'procedures/lista_presenca_form.html', context)


@login_required
def lista_presenca_detail_view(request, pk):
    """Exibe detalhes da lista de presença."""
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    # Buscar registros com relacionamentos
    registros = lista.registros.select_related(
        'colaborador', 'participante_externo', 'procedimento'
    ).order_by('colaborador__nome_completo', 'participante_externo__nome_completo')
    
    # Estatísticas
    from django.db.models import Count
    colaboradores_count = registros.filter(colaborador__isnull=False).values('colaborador').distinct().count()
    externos_count = registros.filter(participante_externo__isnull=False).values('participante_externo').distinct().count()
    procedimentos_count = registros.filter(procedimento__isnull=False).values('procedimento').distinct().count()
    tipos_registro = registros.values('tipo').annotate(count=Count('tipo')).order_by('-count')
    
    stats = {
        'total_participantes': colaboradores_count + externos_count,
        'total_colaboradores': colaboradores_count,
        'total_externos': externos_count,
        'total_procedimentos': procedimentos_count,
        'total_registros': registros.count(),
        'tipos_registro': tipos_registro,
    }
    
    # Obter template e mapeamentos se existirem
    template_mapeamentos = {}
    if lista.template:
        mapeamentos = lista.template.mapeamentos.all()
        for m in mapeamentos:
            if m.campo_dados:  # Apenas mapeamentos com campo_dados definido
                template_mapeamentos[m.placeholder] = {
                    'label': m.get_campo_dados_display(),
                    'placeholder': m.placeholder,
                    'campo_dados': m.campo_dados,
                    'formato': m.formato,
                }
    
    context = {
        'lista': lista,
        'registros': registros,
        'stats': stats,
        'template_mapeamentos': template_mapeamentos,
    }
    return render(request, 'procedures/lista_presenca_detail.html', context)


@login_required
def lista_presenca_edit_view(request, pk):
    """Edita lista de presença existente."""
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    if request.method == 'POST':
        form = ListaPresencaForm(request.POST, instance=lista)
        formset = RegistroTreinamentoFormSet(request.POST, instance=lista)
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                
                registros = formset.save(commit=False)
                for registro in registros:
                    if not registro.data_treinamento:
                        registro.data_treinamento = lista.data_sessao
                    
                    # Tentar linkar colaborador usando nome ou FK
                    if registro.colaborador_nome or registro.colaborador:
                        colab_linkado = tentar_linkar_colaborador(
                            nome_texto=registro.colaborador_nome,
                            colaborador_fk=registro.colaborador,
                            threshold=0.85
                        )
                        # Salvar FK se conseguiu match, mesmo que não tenha FK manual
                        if colab_linkado:
                            registro.colaborador = colab_linkado
                    
                    registro.save()
                
                for obj in formset.deleted_objects:
                    obj.delete()
                
                messages.success(request, f'Lista de presença {lista.codigo} atualizada!')
                return redirect('procedures:lista_presenca_detail', pk=lista.pk)
    else:
        form = ListaPresencaForm(instance=lista)
        formset = RegistroTreinamentoFormSet(instance=lista)
    
    # Buscar colaboradores e procedimentos únicos registrados
    registros = lista.registros.select_related(
        'colaborador', 'participante_externo', 'procedimento'
    ).all()
    
    # Listar colaboradores únicos com contagem
    colaboradores_registrados = []
    for registro in registros:
        if registro.colaborador or registro.colaborador_nome:
            if registro.colaborador:
                nome = registro.colaborador.nome_completo
                matricula = registro.colaborador.matricula
            else:
                nome = registro.colaborador_nome
                matricula = '—'
            
            # Encontrar ou criar entrada
            colab_entry = next((c for c in colaboradores_registrados if c['nome'] == nome), None)
            if not colab_entry:
                colab_entry = {
                    'nome': nome,
                    'matricula': matricula,
                    'tipo': 'Interno' if registro.colaborador else 'Externo',
                    'count': 0
                }
                colaboradores_registrados.append(colab_entry)
            colab_entry['count'] += 1
        elif registro.participante_externo:
            nome = registro.participante_externo.nome_completo
            colab_entry = next((c for c in colaboradores_registrados if c['nome'] == nome), None)
            if not colab_entry:
                colab_entry = {
                    'nome': nome,
                    'matricula': '—',
                    'tipo': 'Externo',
                    'count': 0
                }
                colaboradores_registrados.append(colab_entry)
            colab_entry['count'] += 1
    
    # Listar procedimentos únicos com contagem
    procedimentos_registrados = []
    for registro in registros:
        if registro.procedimento:
            proc = {
                'codigo': registro.procedimento.codigo,
                'nome': registro.procedimento.nome,
                'revisao': registro.procedimento.numero_revisao or 'N/A'
            }
            # Encontrar ou criar entrada
            proc_entry = next((p for p in procedimentos_registrados if p['codigo'] == proc['codigo']), None)
            if not proc_entry:
                proc_entry = {
                    'codigo': proc['codigo'],
                    'nome': proc['nome'],
                    'revisao': proc['revisao'],
                    'count': 0
                }
                procedimentos_registrados.append(proc_entry)
            proc_entry['count'] += 1
    
    context = {
        'form': form,
        'formset': formset,
        'lista': lista,
        'action': 'edit',
        'colaboradores_registrados': colaboradores_registrados,
        'procedimentos_registrados': procedimentos_registrados,
    }
    return render(request, 'procedures/lista_presenca_form.html', context)


@login_required
def lista_presenca_delete_view(request, pk):
    """Deleta lista de presença (e seus registros associados)."""
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    if request.method == 'POST':
        codigo = lista.codigo
        lista.delete()
        messages.success(request, f'Lista de presença {codigo} removida com sucesso!')
        return redirect('procedures:lista_presenca_list')
    
    context = {'lista': lista}
    return render(request, 'procedures/lista_presenca_confirm_delete.html', context)


@login_required
def lista_presenca_export_pdf_view(request, pk):
    """Exporta lista de presença como PDF usando o template mapeado ou layout genérico."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    import PyPDF2
    
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    # Verificar se existe template mapeado com arquivo PDF
    if lista.template and lista.template.arquivo_pdf_template:
        # Tentar gerar PDF usando o template mapeado
        try:
            return _gerar_pdf_com_template_mapeado(lista)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception(f"Erro ao gerar PDF com template: {e}")
            # Fallback para layout genérico
            pass
    
    # Fallback: Layout genérico (código anterior)
    registros = lista.registros.select_related('colaborador', 'procedimento').order_by('colaborador__nome_completo')
    
    # Criar buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"LISTA DE PRESENÇA - {lista.codigo}", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações da sessão
    info_data = [
        ['Título:', lista.titulo],
        ['Instrutor:', str(lista.instrutor) if lista.instrutor else '-'],
        ['Data:', lista.data_sessao.strftime('%d/%m/%Y')],
        ['Horário:', f"{lista.hora_inicio.strftime('%H:%M') if lista.hora_inicio else '-'} às {lista.hora_fim.strftime('%H:%M') if lista.hora_fim else '-'}"],
        ['Carga Horária:', f"{lista.carga_horaria}h" if lista.carga_horaria else '-'],
        ['Local:', lista.local or '-'],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 14*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 1*cm))
    
    # Tabela de participantes
    participantes_data = [['Nº', 'Nome', 'Matrícula', 'Procedimento(s)', 'Assinatura']]
    
    # Agrupar por colaborador
    from itertools import groupby
    for i, (colaborador_id, registros_grupo) in enumerate(groupby(registros, key=lambda r: r.colaborador.id), 1):
        registros_list = list(registros_grupo)
        primeiro = registros_list[0]
        procedimentos = ', '.join([r.procedimento.codigo for r in registros_list])
        
        participantes_data.append([
            str(i),
            primeiro.colaborador.nome_completo,
            primeiro.colaborador.matricula,
            procedimentos,
            ''  # Espaço para assinatura
        ])
    
    participantes_table = Table(participantes_data, colWidths=[1*cm, 6*cm, 3*cm, 5*cm, 3*cm])
    participantes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(participantes_table)
    
    # Construir PDF
    doc.build(elements)
    
    # Retornar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lista_presenca_{lista.codigo}.pdf"'
    return response


def _gerar_pdf_com_template_mapeado(lista):
    """
    Gera PDF preenchendo o template PDF com dados da lista de presença.
    Usa pdfrw para preencher campos de formulário e reportlab para overlay de texto.
    """
    from pdfrw import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from io import BytesIO
    import os
    
    # Obter template
    template = lista.template
    if not template or not template.arquivo_pdf_template:
        raise ValueError("Template sem arquivo PDF")
    
    # Obter mapeamentos
    mapeamentos = MapeamentoCampoListaPresenca.objects.filter(template=template)
    if not mapeamentos.exists():
        raise ValueError("Template sem mapeamentos")
    
    # Preparar dados para preencher
    dados = {}
    for mapeamento in mapeamentos:
        placeholder = mapeamento.placeholder
        campo_dados = mapeamento.campo_dados
        
        # Extrair valor do campo
        valor = ''
        if campo_dados == 'titulo':
            valor = lista.titulo or ''
        elif campo_dados == 'facilitador':
            valor = lista.instrutor_nome or str(lista.instrutor) if lista.instrutor else ''
        elif campo_dados == 'data':
            valor = lista.data_sessao.strftime('%d/%m/%Y') if lista.data_sessao else ''
        elif campo_dados == 'hora_inicio':
            valor = lista.hora_inicio.strftime('%H:%M') if lista.hora_inicio else ''
        elif campo_dados == 'hora_fim':
            valor = lista.hora_fim.strftime('%H:%M') if lista.hora_fim else ''
        elif campo_dados == 'carga_horaria':
            valor = f"{lista.carga_horaria}h" if lista.carga_horaria else ''
        elif campo_dados == 'local':
            valor = lista.local or ''
        elif campo_dados == 'procedimentos':
            procedimentos = lista.registros.values_list('procedimento__codigo', flat=True).distinct()
            valor = ', '.join(procedimentos) if procedimentos else ''
        elif campo_dados == 'empresa':
            valor = ''  # Adicionar lógica se necessário
        elif campo_dados == 'departamento':
            valor = ''  # Adicionar lógica se necessário
        
        if valor:
            dados[placeholder] = str(valor)
    
    # Ler PDF template
    pdf_path = template.arquivo_pdf_template.path
    
    try:
        # Tentar com pdfrw primeiro (para campos de formulário)
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Verificar se tem campos de formulário
        if reader.Root.AcroForm:
            # Preencher campos de formulário
            for field in reader.Root.AcroForm.Fields:
                if field.T:
                    field_name = field.T[1:-1]  # Remove parênteses
                    
                    # Procurar nos dados
                    for placeholder, valor in dados.items():
                        if placeholder.lower() in field_name.lower() or field_name.lower() in placeholder.lower():
                            field.V = f'({valor})'
                            break
        
        # Copiar páginas
        for page in reader.pages:
            writer.addpage(page)
        
        # Salvar em buffer
        output = BytesIO()
        writer.write(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="lista_presenca_{lista.codigo}.pdf"'
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception(f"Erro ao processar PDF com pdfrw: {e}")
        
        # Fallback: Usar pdfplumber para overlay de texto
        try:
            import pdfplumber
            from PyPDF2 import PdfReader as PyPdfReader, PdfWriter as PyPdfWriter
            
            # Criar um PDF com os textos sobrepostos
            packet = BytesIO()
            can = canvas.Canvas(packet, pagesize=A4)
            
            # Escrever dados no canvas (posições podem precisar ajuste)
            y_position = 750
            for placeholder, valor in dados.items():
                can.drawString(100, y_position, f"{placeholder}: {valor}")
                y_position -= 20
            
            can.save()
            
            # Fazer merge do template com o overlay
            packet.seek(0)
            overlay_pdf = PyPdfReader(packet)
            
            base_pdf = PyPdfReader(pdf_path)
            output = PyPdfWriter()
            
            for page_num in range(len(base_pdf.pages)):
                base_page = base_pdf.pages[page_num]
                overlay_page = overlay_pdf.pages[0]
                base_page.merge_page(overlay_page)
                output.add_page(base_page)
            
            # Salvar resultado
            result_buffer = BytesIO()
            output.write(result_buffer)
            result_buffer.seek(0)
            
            response = HttpResponse(result_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="lista_presenca_{lista.codigo}.pdf"'
            return response
            
        except Exception as e2:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception(f"Erro ao fazer overlay: {e2}")
            raise ValueError(f"Não foi possível preencher o template PDF: {e2}")


@login_required
def lista_presenca_importar_view(request):
    """Importa treinamentos em massa via Excel com estrutura completa de 28 colunas."""
    if request.method == 'POST':
        form = ImportacaoTreinamentoForm(request.POST, request.FILES)
        
        if form.is_valid():
            # Validar se arquivo foi enviado
            if 'arquivo' not in request.FILES:
                # Se for AJAX, retornar JSON de erro
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': 'Nenhum arquivo foi selecionado.'}, status=400)

                messages.error(request, 'Nenhum arquivo foi selecionado.')
                context = {'form': form}
                return render(request, 'procedures/lista_presenca_importar.html', context)
            
            arquivo = request.FILES['arquivo']
            criar_listas = form.cleaned_data['criar_listas_automaticamente']
            sobrescrever = form.cleaned_data['sobrescrever_existentes']
            criar_participante_externo = form.cleaned_data.get('criar_participante_externo', False)
            
            try:
                # Ler arquivo Excel - sem fazer parsing de datas (deixar como string)
                # Vamos fazer o parsing no loop para ter melhor controle
                df = pd.read_excel(arquivo)
                
                # Validar se o arquivo tem dados
                if df.empty:
                    messages.error(request, 'O arquivo Excel está vazio. Por favor, adicione dados ao arquivo.')
                    context = {'form': form}
                    return render(request, 'procedures/lista_presenca_importar.html', context)
                
                # Validar colunas obrigatórias
                colunas_obrigatorias = ['matricula', 'data_inicio_treinamento']
                colunas_faltantes = [col for col in colunas_obrigatorias if col not in df.columns]
                
                if colunas_faltantes:
                    messages.error(request, f'Colunas obrigatórias ausentes: {", ".join(colunas_faltantes)}')
                    messages.info(request, 'Baixe o template Excel e certifique-se de usar os nomes de colunas corretos.')
                    context = {'form': form}
                    return render(request, 'procedures/lista_presenca_importar.html', context)
                
                # Processar importação
                resultados = processar_importacao(df, criar_listas, sobrescrever, request.user, criar_participante_externo)
                
                # Mostrar resultados
                if resultados['criados'] > 0:
                    messages.success(request, f"✅ {resultados['criados']} registros criados com sucesso!")
                if resultados['atualizados'] > 0:
                    messages.info(request, f"ℹ️ {resultados['atualizados']} registros atualizados.")
                if resultados['listas_criadas'] > 0:
                    messages.info(request, f"📋 {resultados['listas_criadas']} listas de presença criadas automaticamente.")
                if resultados.get('participantes_externos_criados', 0) > 0:
                    messages.info(request, f"📌 {resultados['participantes_externos_criados']} participantes externos criados automaticamente.")
                if resultados.get('skipped', 0) > 0:
                    messages.info(request, f"ℹ️ {resultados['skipped']} registros já existiam e foram sincronizados (idempotente).")

                # Salvar relatório de erros em cache (por usuário) com fallback para sessão e adicionar link para download
                download_link = None
                if resultados['erros'] > 0:
                    messages.warning(request, f"⚠️ {resultados['erros']} erros encontrados.")
                    # Mostrar primeiros 5 erros
                    for erro in resultados['mensagens_erro'][:5]:
                        messages.error(request, erro)
                    if len(resultados['mensagens_erro']) > 5:
                        messages.error(request, f"... e mais {len(resultados['mensagens_erro']) - 5} erros.")
                    try:
                        from django.core.cache import cache
                        cache_key = f"import_erros_user_{request.user.id}" if request.user and request.user.is_authenticated else None
                        if cache_key:
                            cache.set(cache_key, resultados['mensagens_erro'], timeout=3600)
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.info(f"[IMPORT] Erros salvos em cache: {cache_key}")
                        # Fallback para sessão (compatibilidade)
                        request.session['ultimo_relatorio_erros'] = resultados['mensagens_erro']
                        request.session.modified = True
                    except Exception:
                        logger.exception('Não foi possível salvar relatório de erros (cache/session).')

                    # Preparar link para download do relatório completo
                    from django.urls import reverse
                    download_link = reverse('procedures:lista_presenca_erros_download')
                    from django.utils.html import format_html
                    messages.info(request, format_html('Baixe o relatório completo de erros: <a href="{}" class="alert-link">Download CSV</a>', download_link))

                # Se a requisição for AJAX, retornar JSON com os resultados
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    data = {
                        'criados': resultados['criados'],
                        'atualizados': resultados['atualizados'],
                        'erros': resultados['erros'],
                        'listas_criadas': resultados['listas_criadas'],
                        'mensagens_erro': resultados['mensagens_erro'],
                    }
                    if download_link:
                        data['download_url'] = download_link
                    return JsonResponse(data)
    
                # Redirecionar apenas se houve sucesso — validar se 'next' é seguro
                from django.urls import reverse
                from django.utils.http import url_has_allowed_host_and_scheme

                candidate_next = request.POST.get('next') or ''
                if candidate_next and url_has_allowed_host_and_scheme(candidate_next, allowed_hosts={request.get_host()}):
                    next_url = candidate_next
                else:
                    next_url = reverse('procedures:lista_presenca_list')

                if resultados['criados'] > 0 or resultados['atualizados'] > 0:
                    return redirect(next_url)
                else:
                    # Se nenhum registro foi criado/atualizado, mostrar mensagem e permanecer na página de origem
                    messages.error(request, 'Nenhum registro foi importado. Verifique os erros acima.')
                    return redirect(next_url)
            except pd.errors.EmptyDataError:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': 'O arquivo Excel está vazio ou corrompido.'}, status=400)

                messages.error(request, 'O arquivo Excel está vazio ou corrompido.')
                context = {'form': form}
                return render(request, 'procedures/lista_presenca_importar.html', context)
            except Exception as e:
                # Em modo AJAX, retornar JSON com detalhes técnicos (limitado)
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)

                messages.error(request, f'❌ Erro ao processar arquivo: {str(e)}')
                import traceback
                messages.error(request, f'Detalhes técnicos: {traceback.format_exc()[:200]}')
                context = {'form': form}
                return render(request, 'procedures/lista_presenca_importar.html', context)
        else:
            # Form inválido - retornar erros; se for AJAX, responder em JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                from django.http import JsonResponse
                # converter errors para dict simples
                errors = {k: [str(e) for e in v] for k, v in form.errors.items()}
                return JsonResponse({'success': False, 'errors': errors}, status=400)

            messages.error(request, 'Por favor, corrija os erros no formulário.')
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            context = {'form': form}
            return render(request, 'procedures/lista_presenca_importar.html', context)
    else:
        form = ImportacaoTreinamentoForm()
    
    context = {
        'form': form,
    }
    return render(request, 'procedures/lista_presenca_importar.html', context)


@login_required
def lista_presenca_erros_download_view(request):
    """Baixa um CSV com o relatório de erros da última importação (armazenado em sessão)."""
    from django.http import HttpResponse
    import csv
    from io import StringIO

    # Preferir cache por usuário, depois fallback para sessão
    from django.core.cache import cache
    cache_key = f"import_erros_user_{request.user.id}" if request.user and request.user.is_authenticated else None
    erros = None
    if cache_key:
        erros = cache.get(cache_key)
    if not erros:
        erros = request.session.get('ultimo_relatorio_erros', [])
    if not erros:
        messages.info(request, 'Nenhum relatório de erros disponível para download.')
        return redirect('procedures:lista_presenca_importar')

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['linha', 'mensagem'])

    for e in erros:
        if isinstance(e, str) and e.startswith('Linha') and ':' in e:
            parts = e.split(':', 1)
            linha = parts[0].replace('Linha', '').strip()
            mensagem = parts[1].strip()
        else:
            linha = ''
            mensagem = str(e)
        writer.writerow([linha, mensagem])

    csv_content = si.getvalue()
    response = HttpResponse(csv_content, content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="erros_importacao_treinamentos.csv"'
    return response

@login_required
def lista_presenca_download_template_view(request):
    """Gera template Excel para importação de treinamentos com estrutura completa."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Treinamentos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos - 28 colunas conforme especificação
    headers = [
        'cpf_colaborador',                      # 1
        'nome_colaborador',                     # 2
        'empresa',                              # 3
        'genero',                               # 4
        'matricula',                            # 5 - vincula ao colaborador
        'vinculo_emprego',                      # 6
        'cargo',                                # 7
        'centro_custo',                         # 8
        'status_ocupacao',                      # 9
        'estado_unidade',                       # 10
        'categoria_comunicacao',                # 11
        'metodologia_treinamento',              # 12
        'tipo',                                 # 13 - tipo do colaborador/treinamento
        'area_conhecimento',                    # 14 - será automatizado pelo procedimento
        'titulo_treinamento',                   # 15
        'nome_procedimento',                    # 16
        'codigo_documento',                     # 17 - código do procedimento
        'numero_revisao',                       # 18
        'data_inicio_treinamento',              # 19
        'data_final_treinamento',               # 20
        'mes',                                  # 21
        'facilitador_fornecedor',               # 22 - instrutor
        'carga_horaria',                        # 23 - formato hh:mm
        'custo_treinamento',                    # 24 - R$ por pessoa
        'carga_horaria_horas',                  # 25 - em horas decimais
        'necessita_avaliacao_eficacia',         # 26
        'data_limite_avaliacao_eficacia',       # 27
        'observacao'                            # 28
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Exemplos de dados
    exemplos = [
        [
            '123.456.789-00',           # cpf_colaborador
            'João Silva',               # nome_colaborador
            'Empresa ABC',              # empresa
            'M',                        # genero
            '123456',                   # matricula - vincula ao colaborador
            'CLT',                      # vinculo_emprego
            'Técnico de Laboratório',   # cargo
            'LAB001',                   # centro_custo
            'Ativo',                    # status_ocupacao
            'SP',                       # estado_unidade
            'Presencial',               # categoria_comunicacao
            'Teórico-Prático',          # metodologia_treinamento
            'PROCEDIMENTO',             # tipo
            'Qualidade',                # area_conhecimento
            'Treinamento PO-001',       # titulo_treinamento
            'Procedimento Operacional 001',  # nome_procedimento
            'PO-001',                   # codigo_documento
            '03',                       # numero_revisao
            '2025-01-15',               # data_inicio_treinamento
            '2025-01-15',               # data_final_treinamento
            'Janeiro',                  # mes
            'Maria Santos',             # facilitador_fornecedor
            '04:00',                    # carga_horaria (hh:mm)
            '0.00',                     # custo_treinamento
            '4',                        # carga_horaria_horas
            'SIM',                      # necessita_avaliacao_eficacia
            '2025-02-15',               # data_limite_avaliacao_eficacia
            'Primeiro treinamento'      # observacao
        ],
        [
            '987.654.321-00',
            'Maria Oliveira',
            'Empresa ABC',
            'F',
            '123457',
            'CLT',
            'Analista de Qualidade',
            'LAB002',
            'Ativo',
            'SP',
            'Online',
            'Teórico',
            'CAPACITACAO',
            'Metrologia',
            'Curso de Calibração',
            '',                         # sem procedimento
            '',                         # sem código
            '',                         # sem revisão
            '2025-01-20',
            '2025-01-22',
            'Janeiro',
            'Instituto XYZ',
            '16:00',
            '500.00',
            '16',
            'NAO',
            '',
            'Curso externo'
        ],
    ]
    
    for row_num, row_data in enumerate(exemplos, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 18, 'B': 25, 'C': 20, 'D': 8, 'E': 12,
        'F': 15, 'G': 25, 'H': 15, 'I': 15, 'J': 12,
        'K': 20, 'L': 22, 'M': 15, 'N': 20, 'O': 30,
        'P': 35, 'Q': 18, 'R': 12, 'S': 20, 'T': 20,
        'U': 12, 'V': 30, 'W': 12, 'X': 15, 'Y': 12,
        'Z': 20, 'AA': 25, 'AB': 30
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Adicionar aba de instruções
    ws_instrucoes = wb.create_sheet("Instruções")
    instrucoes = [
        ["INSTRUÇÕES PARA IMPORTAÇÃO DE TREINAMENTOS"],
        [""],
        ["Estrutura do Template:"],
        [""],
        ["COLUNAS OBRIGATÓRIAS:"],
        ["5. matricula - Vincula o treinamento ao colaborador"],
        ["19. data_inicio_treinamento - Data de início do treinamento"],
        [""],
        ["COLUNAS CONDICIONAIS:"],
        ["17. codigo_documento - Obrigatório se tipo='PROCEDIMENTO'"],
        ["15. titulo_treinamento - Obrigatório se não houver procedimento"],
        [""],
        ["INFORMAÇÕES DO COLABORADOR (preenchidas automaticamente pelo sistema):"],
        ["1. cpf_colaborador"],
        ["2. nome_colaborador"],
        ["3. empresa"],
        ["4. genero"],
        ["6. vinculo_emprego"],
        ["7. cargo"],
        ["8. centro_custo"],
        ["9. status_ocupacao"],
        ["10. estado_unidade"],
        [""],
        ["INFORMAÇÕES DO TREINAMENTO:"],
        ["11. categoria_comunicacao - Ex: Presencial, Online, EAD"],
        ["12. metodologia_treinamento - Ex: Teórico, Prático, Teórico-Prático"],
        ["13. tipo - PROCEDIMENTO, ALINHAMENTO, REUNIAO, CAPACITACAO, OUTRO"],
        ["14. area_conhecimento - Área de conhecimento (preenchido automaticamente pelo procedimento)"],
        ["15. titulo_treinamento - Título do treinamento"],
        ["16. nome_procedimento - Nome do procedimento (se aplicável)"],
        ["17. codigo_documento - Código do procedimento (ex: PO-001)"],
        ["18. numero_revisao - Número da revisão do procedimento"],
        [""],
        ["DATAS E HORÁRIOS:"],
        ["19. data_inicio_treinamento - Formato: AAAA-MM-DD (ex: 2025-01-15)"],
        ["20. data_final_treinamento - Data final do treinamento"],
        ["21. mes - Mês de referência (ex: Janeiro)"],
        ["22. facilitador_fornecedor - Nome do responsável pela aplicação"],
        ["23. carga_horaria - Formato: hh:mm (ex: 04:00)"],
        ["24. custo_treinamento - Valor em R$ por pessoa (ex: 500.00)"],
        ["25. carga_horaria_horas - Carga horária em horas decimais (ex: 4)"],
        [""],
        ["AVALIAÇÃO DE EFICÁCIA:"],
        ["26. necessita_avaliacao_eficacia - SIM ou NAO"],
        ["27. data_limite_avaliacao_eficacia - Data limite para avaliação"],
        [""],
        ["OBSERVAÇÕES:"],
        ["28. observacao - Observações adicionais"],
        [""],
        ["IMPORTANTE:"],
        ["1. A matrícula (coluna 5) é usada para vincular o treinamento ao colaborador"],
        ["2. Se o tipo for PROCEDIMENTO, o código_documento (coluna 17) é obrigatório"],
        ["3. A área_conhecimento (coluna 14) será preenchida automaticamente pelo procedimento cadastrado"],
        ["4. Formato de data: AAAA-MM-DD (ex: 2025-01-15)"],
        ["5. Carga horária: formato hh:mm (ex: 04:00 para 4 horas)"],
        ["6. Custo: apenas números e decimais com ponto (ex: 500.00)"],
    ]
    
    for row_num, row_data in enumerate(instrucoes, 1):
        cell = ws_instrucoes.cell(row=row_num, column=1)
        cell.value = row_data[0]
        if "INSTRUÇÕES" in row_data[0] or "COLUNAS" in row_data[0] or "INFORMAÇÕES" in row_data[0] or "DATAS" in row_data[0] or "AVALIAÇÃO" in row_data[0] or "OBSERVAÇÕES" in row_data[0]:
            cell.font = Font(bold=True, size=12)
    
    ws_instrucoes.column_dimensions['A'].width = 100
    
    # Salvar em buffer
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_importacao_treinamentos.xlsx'
    return response


def atualizar_demandas_apos_importacao():
    """
    Sincroniza as demandas (gaps) de todos os colaboradores após importação de treinamentos.
    
    Esta função é chamada após a importação bem-sucedida de registros de treinamento.
    Ela recalcula os gaps de demanda para cada colaborador comparando:
    - Procedimentos requeridos pelo perfil atribuído
    - Procedimentos já treinados (registrados em RegistroTreinamento)
    
    O sistema utiliza esses dados na análise de gaps para atualizar a matriz de demandas
    e reduzir o número de gaps pendentes para cada colaborador.
    """
    from procedures.models import ColaboradorPerfil, Procedimento
    
    try:
        # Buscar todos os colaboradores com perfis atribuídos e ativos
        colaboradores_perfis = ColaboradorPerfil.objects.filter(
            ativo=True
        ).select_related('colaborador', 'perfil')
        
        colaboradores_processados = 0
        total_demandas_reduzidas = 0
        
        for cp in colaboradores_perfis:
            colaborador = cp.colaborador
            perfil = cp.perfil
            
            # Buscar procedimentos requeridos pelo perfil (via subgrupos_treinamento)
            procedimentos_requeridos = Procedimento.objects.filter(
                subgrupos_treinamento__grupo__perfil=perfil
            ).distinct()
            
            # Buscar procedimentos já treinados por este colaborador
            procedimentos_treinados_ids = set(
                RegistroTreinamento.objects.filter(
                    colaborador=colaborador,
                    procedimento__isnull=False
                ).values_list('procedimento_id', flat=True)
            )
            
            # Calcular demanda pendente (gaps) 
            demanda_pendente = procedimentos_requeridos.exclude(
                id__in=procedimentos_treinados_ids
            ).count()
            
            total_requerido = procedimentos_requeridos.count()
            
            # Log para auditoria
            import logging
            logger = logging.getLogger(__name__)
            logger.info(
                f"Colaborador: {colaborador.nome_completo} | "
                f"Perfil: {perfil.nome} | "
                f"Requerido: {total_requerido} | "
                f"Treinado: {total_requerido - demanda_pendente} | "
                f"Demanda Pendente: {demanda_pendente}"
            )
            
            colaboradores_processados += 1
            total_demandas_reduzidas += (total_requerido - demanda_pendente)
        
        # Log final
        import logging
        logger = logging.getLogger(__name__)
        logger.info(
            f"✅ Demandas atualizadas para {colaboradores_processados} colaboradores | "
            f"Total de {total_demandas_reduzidas} demandas satisfeitas por novos treinamentos"
        )
        
        return True
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"❌ Erro ao atualizar demandas após importação: {str(e)}", exc_info=True)
        return False


def processar_importacao(df, criar_listas, sobrescrever, usuario, criar_participante_externo=False):
    """Processa DataFrame com dados de importação - estrutura completa com 28 colunas com OTIMIZAÇÕES."""
    resultados = {
        'criados': 0,
        'atualizados': 0,
        'erros': 0,
        'listas_criadas': 0,
        'mensagens_erro': [],
        'skipped': 0,
        'participantes_externos_criados': 0,
    }
    
    # OTIMIZAÇÃO: Cache em memória para evitar queries repetidas
    cache_colaboradores = {}
    cache_procedimentos = {}
    cache_listas = {}
    
    # Pré-carregar TODOS os colaboradores e procedimentos que serão usados
    import logging
    logger = logging.getLogger(__name__)
    logger.info("[IMPORT] Pré-carregando dados do banco...")

    # Normalizadores
    def norm_matricula(s):
        if s is None:
            return ''
        return str(s).strip()

    def norm_matricula_key(s):
        return norm_matricula(s).upper()

    def norm_cpf(s):
        if s is None:
            return ''
        return ''.join(ch for ch in str(s) if ch.isdigit())

    def norm_codigo_proc(s):
        if s is None:
            return ''
        return ''.join(ch for ch in str(s).upper() if ch.isalnum())

    # Colaboradores - criar múltiplos índices para matching robusto
    todos_colaboradores = {}
    colaboradores_por_cpf = {}
    colaboradores_por_nome = {}
    for c in Colaborador.objects.all():
        mkey = norm_matricula_key(c.matricula)
        todos_colaboradores[mkey] = c
        mkey_nozeros = mkey.lstrip('0')
        if mkey_nozeros and mkey_nozeros not in todos_colaboradores:
            todos_colaboradores[mkey_nozeros] = c
        cpf_key = norm_cpf(getattr(c, 'cpf', None))
        if cpf_key:
            colaboradores_por_cpf[cpf_key] = c
        nome_key = str(getattr(c, 'nome_completo', '')).strip().upper()
        if nome_key:
            colaboradores_por_nome[nome_key] = c

    # Procedimentos - índice por código normalizado
    todos_procedimentos = {}
    for p in Procedimento.objects.all():
        pkey = norm_codigo_proc(p.codigo)
        if pkey:
            todos_procedimentos[pkey] = p

    # Ajustar cache para usar a mesma chave de agrupamento
    todas_listas = {}
    for l in ListaPresenca.objects.all():
        codigo_proc = ''
        facilitador_key = ''
        if l.titulo:
            # Tentar extrair o código do procedimento do título (assumindo padrão novo)
            partes = l.titulo.split('-')
            if len(partes) >= 4:
                codigo_proc = partes[3]
            if len(partes) >= 5:
                facilitador_key = partes[4]
        chave = (codigo_proc, l.data_sessao, facilitador_key)
        todas_listas[chave] = l
    
    # Pré-carregar registros existentes para verificação duplicada mais rápida
    # Separar em dois dicionários: um para registros com procedimento, outro para sem
    # Suporte para participantes externos (colaborador NULL) usando prefixo 'c' ou 'e' na chave
    registros_com_procedimento = {}  # ('c'/'e', participante_id, procedimento_id) -> True
    registros_sem_procedimento = {}  # ('c'/'e', participante_id, titulo_treinamento, data_treinamento) -> True
    
    for reg in RegistroTreinamento.objects.all().values('colaborador_id', 'participante_externo_id', 'procedimento_id', 'titulo_treinamento', 'data_treinamento'):
        participante_prefix = 'c' if reg['colaborador_id'] else 'e'
        participante_id = reg['colaborador_id'] if reg['colaborador_id'] else reg['participante_externo_id']
        if reg['procedimento_id']:
            # Registros com procedimento - incluir também a data para permitir múltiplos registros do mesmo procedimento em datas diferentes
            chave = (participante_prefix, participante_id, reg['procedimento_id'], reg['data_treinamento'])
            registros_com_procedimento[chave] = True
        else:
            # Registros sem procedimento - constraint é (participante, titulo_treinamento, data_treinamento)
            chave = (participante_prefix, participante_id, reg['titulo_treinamento'], reg['data_treinamento'])
            registros_sem_procedimento[chave] = True
    
    # Dicionário para agrupar por sessão
    sessoes = {}
    
    # OTIMIZAÇÃO: Batch para criar múltiplos registros de uma vez
    registros_para_criar = []
    registros_para_atualizar = []
    
    for index, row in df.iterrows():
        try:
            erros_linha = []  # coletar vários erros dessa linha
            participante_externo = None
            # Buscar colaborador pela matrícula (campo obrigatório)
            if 'matricula' not in row or pd.isna(row['matricula']):
                resultados['erros'] += 1
                msg = f"Linha {index + 2}: Matrícula ausente"
                resultados['mensagens_erro'].append(msg)

                continue
                
            matricula_raw = str(row['matricula']).strip()
            matricula_str = matricula_raw
            colaborador = None
            # Tentativas de match: matricula (normalizada), matricula sem zeros, cpf, nome similar
            mkey = norm_matricula_key(matricula_str)
            if mkey in todos_colaboradores:
                colaborador = todos_colaboradores[mkey]
                match_method = 'matricula'
            else:
                mkey_nozeros = mkey.lstrip('0')
                if mkey_nozeros and mkey_nozeros in todos_colaboradores:
                    colaborador = todos_colaboradores[mkey_nozeros]
                    match_method = 'matricula_nozeros'
                else:
                    # Tentar CPF
                    cpf_val = str(row.get('cpf_colaborador', '') or '')
                    cpf_key = norm_cpf(cpf_val)
                    if cpf_key and cpf_key in colaboradores_por_cpf:
                        colaborador = colaboradores_por_cpf[cpf_key]
                        match_method = 'cpf'
                    else:
                        # Tentar nome similar
                        nome_val = str(row.get('nome_colaborador', '') or '')
                        if nome_val:
                            possivel = tentar_linkar_colaborador(nome_val)
                            if possivel:
                                colaborador = possivel
                                match_method = f'nome_similar ({possivel.id})'
            if not colaborador:
                # Se modo tolerante, criar participante externo
                if criar_participante_externo:
                    nome_ext = str(row.get('nome_colaborador') or row.get('colaborador_nome') or '').strip()
                    cpf_ext = str(row.get('cpf_colaborador') or '').strip() or None
                    empresa_ext = str(row.get('empresa') or '').strip() or None
                    participante_externo = ParticipanteExterno.objects.create(
                        nome_completo=nome_ext or f"Participante Externo Linha {index + 2}",
                        cpf=cpf_ext,
                        empresa=empresa_ext
                    )
                    resultados['participantes_externos_criados'] = resultados.get('participantes_externos_criados', 0) + 1
                    match_method = 'created_externo'
                else:
                    resultados['erros'] += 1
                    msg = f"Linha {index + 2}: Colaborador com matrícula '{matricula_str}' não encontrado no sistema"
                    resultados['mensagens_erro'].append(msg)
                    continue

            # Determinar tipo de treinamento
            tipo = row.get('tipo', 'PROCEDIMENTO')
            if pd.isna(tipo) or not tipo:
                tipo = 'PROCEDIMENTO'
            tipo = tipo.upper()
            
            # Validar tipo
            tipos_validos = ['PROCEDIMENTO', 'ALINHAMENTO', 'REUNIAO', 'CAPACITACAO', 'OUTRO']
            if tipo not in tipos_validos:
                tipo = 'PROCEDIMENTO'
            
            # Buscar procedimento (obrigatório apenas se tipo=PROCEDIMENTO)
            procedimento = None
            if tipo == 'PROCEDIMENTO':
                if 'codigo_documento' in row and pd.notna(row['codigo_documento']):
                    codigo_proc = str(row['codigo_documento']).strip()
                    codigo_key = norm_codigo_proc(codigo_proc)
                    procedimento = None
                    match_proc_method = None
                    if codigo_key in todos_procedimentos:
                        procedimento = todos_procedimentos[codigo_key]
                        match_proc_method = 'codigo_norm'
                    else:
                        # Tentar sem pontos/formatos alternativos
                        alt = codigo_key.replace('.', '').replace(' ', '')
                        if alt in todos_procedimentos:
                            procedimento = todos_procedimentos[alt]
                            match_proc_method = 'codigo_alt'
                    if not procedimento:
                        resultados['erros'] += 1
                        msg = f"Linha {index + 2}: Procedimento {row['codigo_documento']} não encontrado"
                        resultados['mensagens_erro'].append(msg)

                        continue
                    else:
                        pass

                else:
                    resultados['erros'] += 1
                    msg = f"Linha {index + 2}: Tipo PROCEDIMENTO requer codigo_documento"
                    resultados['mensagens_erro'].append(msg)


                    continue
            
            # Obter título do treinamento
            titulo_treinamento = row.get('titulo_treinamento', '')
            if pd.isna(titulo_treinamento):
                titulo_treinamento = ''
            # VALIDAÇÃO: Limitar a 200 caracteres (conforme modelo)
            titulo_treinamento = str(titulo_treinamento)[:200]
            
            # Se não tem procedimento, título é obrigatório
            if not procedimento and not titulo_treinamento:
                resultados['erros'] += 1
                msg = f"Linha {index + 2}: Treinamento sem procedimento requer titulo_treinamento"
                resultados['mensagens_erro'].append(msg)

                continue
            
            # Converter data de início (obrigatória)
            if 'data_inicio_treinamento' not in row or pd.isna(row['data_inicio_treinamento']):
                resultados['erros'] += 1
                msg = f"Linha {index + 2}: data_inicio_treinamento ausente"
                resultados['mensagens_erro'].append(msg)
                continue
            
            try:
                # Converter para data, tratando diferentes formatos
                valor_data = row['data_inicio_treinamento']
                if isinstance(valor_data, str):
                    # Se for string, tentar parsing com múltiplos formatos
                    try:
                        data_treinamento = pd.to_datetime(valor_data, format='%d/%m/%Y').date()
                    except (ValueError, TypeError):
                        try:
                            data_treinamento = pd.to_datetime(valor_data, format='%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            data_treinamento = pd.to_datetime(valor_data).date()
                else:
                    # Se for timestamp/datetime, converter diretamente
                    data_treinamento = pd.to_datetime(valor_data).date()
                    
                # Validar se data é válida
                if data_treinamento is None or (isinstance(data_treinamento, float) and pd.isna(data_treinamento)):
                    raise ValueError("Data nula após conversão")
            except Exception as e:
                resultados['erros'] += 1
                msg = f"Linha {index + 2}: Erro ao converter data_inicio_treinamento ({str(e)}). Valor: {row['data_inicio_treinamento']}"
                resultados['mensagens_erro'].append(msg)
                continue
            
            # Converter data final (opcional)
            data_final_treinamento = None
            if 'data_final_treinamento' in row and pd.notna(row['data_final_treinamento']):
                try:
                    valor_data = row['data_final_treinamento']
                    if isinstance(valor_data, str):
                        try:
                            data_final_treinamento = pd.to_datetime(valor_data, format='%d/%m/%Y').date()
                        except (ValueError, TypeError):
                            try:
                                data_final_treinamento = pd.to_datetime(valor_data, format='%Y-%m-%d').date()
                            except (ValueError, TypeError):
                                data_final_treinamento = pd.to_datetime(valor_data).date()
                    else:
                        data_final_treinamento = pd.to_datetime(valor_data).date()
                except (ValueError, TypeError):
                    pass  # Ignorar erros em data final (opcional)
            
            # Buscar facilitador/fornecedor (opcional)
            facilitador_fornecedor = row.get('facilitador_fornecedor', '')
            if pd.isna(facilitador_fornecedor):
                facilitador_fornecedor = ''
            # VALIDAÇÃO: Limitar a 200 caracteres (conforme modelo)
            facilitador_fornecedor = str(facilitador_fornecedor)[:200]
            
            # Obter revisão treinada
            revisao_treinada = ''
            if procedimento:
                revisao_treinada = str(row.get('numero_revisao', procedimento.numero_revisao or ''))
            if pd.isna(revisao_treinada) or not revisao_treinada:
                revisao_treinada = '01'
            # VALIDAÇÃO: Limitar revisao_treinada a 10 caracteres (conforme modelo)
            revisao_treinada = str(revisao_treinada)[:10]
            
            # Processar campos adicionais
            categoria_comunicacao = row.get('categoria_comunicacao', '')
            if pd.isna(categoria_comunicacao):
                categoria_comunicacao = ''
            # VALIDAÇÃO: Limitar a 100 caracteres
            categoria_comunicacao = str(categoria_comunicacao)[:100]
                
            metodologia_treinamento = row.get('metodologia_treinamento', '')
            if pd.isna(metodologia_treinamento):
                metodologia_treinamento = ''
            # VALIDAÇÃO: Limitar a 100 caracteres
            metodologia_treinamento = str(metodologia_treinamento)[:100]
            
            # Área de conhecimento - priorizar do procedimento
            area_conhecimento = ''
            if procedimento and procedimento.area_conhecimento:
                area_conhecimento = procedimento.area_conhecimento
            elif 'area_conhecimento' in row and pd.notna(row['area_conhecimento']):
                area_conhecimento = str(row['area_conhecimento'])
            # VALIDAÇÃO: Limitar a 200 caracteres
            area_conhecimento = str(area_conhecimento)[:200]
            
            # Carga horária - VALIDAÇÃO: Limitar a 10 caracteres (hh:mm format)
            carga_horaria = row.get('carga_horaria', '')
            if pd.isna(carga_horaria):
                carga_horaria = ''
            else:
                carga_horaria = str(carga_horaria).strip()
                # Se tiver valor, validar que é formato hh:mm
                if carga_horaria and ':' not in carga_horaria:
                    # Se for número puro (4 horas), converter para hh:00
                    try:
                        horas = int(float(carga_horaria))
                        carga_horaria = f"{horas:02d}:00"
                    except (ValueError, TypeError):
                        carga_horaria = ''
            # Limitar a 10 caracteres
            carga_horaria = carga_horaria[:10]
            
            # NOVIDADE: Campos da Lista de Presença (hora e carga da sessão)
            hora_inicio_sessao = None
            hora_fim_sessao = None
            carga_horaria_sessao = None
            
            # Extrair hora de início da sessão (formato hh:mm)
            if 'hora_inicio' in row and pd.notna(row['hora_inicio']):
                try:
                    valor = str(row['hora_inicio']).strip()
                    # Tentar converter para TimeField
                    if ':' in valor:
                        hora_inicio_sessao = pd.to_datetime(valor, format='%H:%M').time()
                except (ValueError, TypeError):
                    pass
            
            # Extrair hora de fim da sessão (formato hh:mm)
            if 'hora_fim' in row and pd.notna(row['hora_fim']):
                try:
                    valor = str(row['hora_fim']).strip()
                    if ':' in valor:
                        hora_fim_sessao = pd.to_datetime(valor, format='%H:%M').time()
                except (ValueError, TypeError):
                    pass
            
            # Extrair carga horária da sessão (em horas decimais)
            if 'carga_horaria_lista' in row and pd.notna(row['carga_horaria_lista']):
                try:
                    carga_horaria_sessao = Decimal(str(row['carga_horaria_lista']).replace(',', '.'))
                except (InvalidOperation, ValueError, TypeError):
                    pass
            
            # Custo do treinamento
            custo_treinamento = None
            if 'custo_treinamento' in row and pd.notna(row['custo_treinamento']):
                try:
                    custo_treinamento = Decimal(str(row['custo_treinamento']).replace(',', '.'))
                except (InvalidOperation, ValueError, TypeError):
                    pass
            
            # Mês de referência
            mes_referencia = row.get('mes', '')
            if pd.isna(mes_referencia):
                mes_referencia = ''
            # VALIDAÇÃO: Limitar a 100 caracteres
            mes_referencia = str(mes_referencia)[:100]
            
            # Necessita avaliação de eficácia
            necessita_avaliacao = False
            if 'necessita_avaliacao_eficacia' in row and pd.notna(row['necessita_avaliacao_eficacia']):
                valor = str(row['necessita_avaliacao_eficacia']).upper()
                necessita_avaliacao = valor in ['SIM', 'S', 'TRUE', '1', 'YES']
            
            # Data limite para avaliação de eficácia
            data_limite_avaliacao = None
            if 'data_limite_avaliacao_eficacia' in row and pd.notna(row['data_limite_avaliacao_eficacia']):
                try:
                    valor_data = row['data_limite_avaliacao_eficacia']
                    if isinstance(valor_data, str):
                        try:
                            data_limite_avaliacao = pd.to_datetime(valor_data, format='%d/%m/%Y').date()
                        except (ValueError, TypeError):
                            try:
                                data_limite_avaliacao = pd.to_datetime(valor_data, format='%Y-%m-%d').date()
                            except (ValueError, TypeError):
                                data_limite_avaliacao = pd.to_datetime(valor_data).date()
                    else:
                        data_limite_avaliacao = pd.to_datetime(valor_data).date()
                except (ValueError, TypeError):
                    pass  # Ignorar erros em data limite
            
            # Descrição/observações
            observacoes = row.get('observacao', '')
            if pd.isna(observacoes):
                observacoes = ''
            
            # Criar chave de sessão para agrupamento (codigo_documento, data_inicio_treinamento, facilitador_fornecedor)
            lista = None
            codigo_proc = str(row['codigo_documento']).strip() if 'codigo_documento' in row and pd.notna(row['codigo_documento']) else ''
            facilitador_key = str(facilitador_fornecedor).strip() if facilitador_fornecedor else ''
            chave_sessao = (codigo_proc, data_treinamento, facilitador_key)

            # Determinar se vamos criar/usar ListaPresenca ou não
            lista = None
            if criar_listas:
                if chave_sessao not in sessoes:
                    # Gerar título no formato ANO+MÊS+DIA+CODIGO PROCEDIMENTO+MATRICULA DO FORNECEDOR
                    ano = data_treinamento.year if data_treinamento else ''
                    mes = f"{data_treinamento.month:02d}" if data_treinamento else ''
                    dia = f"{data_treinamento.day:02d}" if data_treinamento else ''
                    codigo = codigo_proc if codigo_proc else ''
                    matricula_fornecedor = ''
                    if facilitador_fornecedor:
                        import re
                        match = re.search(r'\d+', facilitador_fornecedor)
                        if match:
                            matricula_fornecedor = match.group(0)
                        else:
                            matricula_fornecedor = facilitador_fornecedor.replace(' ', '').upper()
                    partes = [str(ano), str(mes), str(dia), codigo, matricula_fornecedor]
                    titulo_lista = '-'.join([p for p in partes if p])

                    # Usar cache com a chave correta
                    if chave_sessao in todas_listas:
                        lista = todas_listas[chave_sessao]
                    else:
                        lista = ListaPresenca.objects.create(
                            titulo=titulo_lista[:200],
                            instrutor_nome=facilitador_fornecedor,
                            instrutor=None,
                            data_sessao=data_treinamento,
                            hora_inicio=hora_inicio_sessao,
                            hora_fim=hora_fim_sessao,
                            carga_horaria=carga_horaria_sessao,
                            local='',
                            criado_por=usuario
                        )
                        todas_listas[chave_sessao] = lista
                        resultados['listas_criadas'] += 1

                    sessoes[chave_sessao] = lista
                else:
                    lista = sessoes[chave_sessao]
            else:
                # Não agrupar em listas: manter lista = None e usar data_treinamento direto
                lista = None
            # OTIMIZAÇÃO: Acumular registros para criar em batch
            # Verificar se já existe registro usando a chave correta (suporta colaborador e participante_externo)
            # Usar data da lista (se houver) como data_registro
            data_registro = lista.data_sessao if lista else data_treinamento
            if procedimento:
                participante_prefix = 'c' if colaborador else 'e'
                participante_id = colaborador.id if colaborador else (participante_externo.id if participante_externo else None)
                chave_existencia = (participante_prefix, participante_id, procedimento.id, data_registro)
                existe = chave_existencia in registros_com_procedimento
            else:
                participante_prefix = 'c' if colaborador else 'e'
                participante_id = colaborador.id if colaborador else (participante_externo.id if participante_externo else None)
                chave_existencia = (participante_prefix, participante_id, titulo_treinamento, data_treinamento)
                existe = chave_existencia in registros_sem_procedimento

            if existe:
                # Em vez de tratar como erro, atualizar para evitar duplicação (comportamento idempotente)
                data_registro = lista.data_sessao if lista else data_treinamento
                registros_para_atualizar.append({
                    'colaborador': colaborador,
                    'participante_externo': participante_externo,
                    'procedimento': procedimento,
                    'titulo_treinamento': titulo_treinamento,
                    'data_treinamento': data_treinamento,
                    'dados': {
                        'tipo': tipo,
                        'lista_presenca': lista,
                        'data_treinamento': data_registro,  # Usar data da lista
                        'revisao_treinada': revisao_treinada,
                        'observacoes': observacoes,
                        'categoria_comunicacao': categoria_comunicacao,
                        'metodologia_treinamento': metodologia_treinamento,
                        'area_conhecimento': area_conhecimento,
                        'facilitador_fornecedor': facilitador_fornecedor,
                        'carga_horaria': carga_horaria,
                        'custo_treinamento': custo_treinamento,
                        'data_final_treinamento': data_final_treinamento,
                        'mes_referencia': mes_referencia,
                        'necessita_avaliacao_eficacia': necessita_avaliacao,
                        'data_limite_avaliacao_eficacia': data_limite_avaliacao
                    }
                })
                resultados['skipped'] += 1
            else:
                # Adicionar à lista de criação
                # Usar data da lista de presença como data_treinamento
                data_registro = lista.data_sessao if lista else data_treinamento
                
                registros_para_criar.append(RegistroTreinamento(
                    colaborador=colaborador,
                    participante_externo=participante_externo if 'participante_externo' in locals() else None,
                    colaborador_nome=(colaborador.nome_completo if colaborador else (str(row.get('nome_colaborador') or '')[:200])),
                    procedimento=procedimento,
                    tipo=tipo,
                    titulo_treinamento=titulo_treinamento,
                    lista_presenca=lista,
                    data_treinamento=data_registro,  # Usar data da lista
                    revisao_treinada=revisao_treinada,
                    observacoes=observacoes,
                    categoria_comunicacao=categoria_comunicacao,
                    metodologia_treinamento=metodologia_treinamento,
                    area_conhecimento=area_conhecimento,
                    facilitador_fornecedor=facilitador_fornecedor,
                    carga_horaria=carga_horaria,
                    custo_treinamento=custo_treinamento,
                    data_final_treinamento=data_final_treinamento,
                    mes_referencia=mes_referencia,
                    necessita_avaliacao_eficacia=necessita_avaliacao,
                    data_limite_avaliacao_eficacia=data_limite_avaliacao
                ))
                
        except Exception as e:
            resultados['erros'] += 1
            msg = f"Linha {index + 2}: {str(e)}"
            resultados['mensagens_erro'].append(msg)
        # Registrar erro da linha
            resultados['mensagens_erro'].append(msg)
            continue


    
    # OTIMIZAÇÃO: Usar update_or_create para evitar duplicatas
    # Isso é mais lento que bulk_create mas não falha com duplicatas
    logger.info(f"[IMPORT] Processando {len(registros_para_criar)} registros...")
    # Helper to coerce pandas timestamps/NaT to native date or None
    from datetime import datetime, date
    def _safe_date(val):
        if val is None:
            return None
        # pandas NaT
        try:
            import pandas as _pd
            if isinstance(val, _pd._libs.tslibs.nattype.NaTType) or (_pd.isna(val) if hasattr(_pd, 'isna') else False):
                return None
        except Exception:
            pass
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        try:
            # Try pandas conversion
            import pandas as _pd
            return _pd.to_datetime(val).date()
        except Exception:
            return None

    if registros_para_criar:
        for reg in registros_para_criar:
            try:
                # Prepare filter kwargs depending on whether registro refers to colaborador or participante_externo
                if reg.procedimento:
                    # Incluir data na chave do filtro para permitir múltiplos registros do mesmo procedimento em datas diferentes
                    if reg.colaborador:
                        filter_kwargs = {'colaborador': reg.colaborador, 'procedimento': reg.procedimento, 'data_treinamento': _safe_date(reg.data_treinamento)}
                    else:
                        filter_kwargs = {'participante_externo': getattr(reg, 'participante_externo', None), 'procedimento': reg.procedimento, 'data_treinamento': _safe_date(reg.data_treinamento)}

                    defaults = {
                        'tipo': reg.tipo,
                        'titulo_treinamento': reg.titulo_treinamento,
                        'lista_presenca': reg.lista_presenca,
                        'data_treinamento': _safe_date(reg.data_treinamento),
                        'revisao_treinada': reg.revisao_treinada,
                        'observacoes': reg.observacoes,
                        'categoria_comunicacao': reg.categoria_comunicacao,
                        'metodologia_treinamento': reg.metodologia_treinamento,
                        'area_conhecimento': reg.area_conhecimento,
                        'facilitador_fornecedor': reg.facilitador_fornecedor,
                        'carga_horaria': reg.carga_horaria,
                        'custo_treinamento': reg.custo_treinamento,
                        'data_final_treinamento': _safe_date(reg.data_final_treinamento),
                        'mes_referencia': reg.mes_referencia,
                        'necessita_avaliacao_eficacia': reg.necessita_avaliacao_eficacia,
                        'data_limite_avaliacao_eficacia': _safe_date(reg.data_limite_avaliacao_eficacia),
                        'colaborador_nome': reg.colaborador_nome,
                        'participante_externo': getattr(reg, 'participante_externo', None),
                    }

                    obj, created = RegistroTreinamento.objects.update_or_create(
                        **filter_kwargs,
                        defaults=defaults
                    )
                    if created:
                        resultados['criados'] += 1
                    else:
                        resultados['atualizados'] += 1
                else:
                    # Sem procedimento - usar update_or_create com chave adequada
                    if reg.colaborador:
                        filter_kwargs = {'colaborador': reg.colaborador, 'titulo_treinamento': reg.titulo_treinamento, 'data_treinamento': _safe_date(reg.data_treinamento)}
                    else:
                        filter_kwargs = {'participante_externo': getattr(reg, 'participante_externo', None), 'titulo_treinamento': reg.titulo_treinamento, 'data_treinamento': _safe_date(reg.data_treinamento)}

                    defaults = {
                        'colaborador_nome': reg.colaborador_nome,
                        'procedimento': reg.procedimento,
                        'tipo': reg.tipo,
                        'lista_presenca': reg.lista_presenca,
                        'revisao_treinada': reg.revisao_treinada,
                        'observacoes': reg.observacoes,
                        'categoria_comunicacao': reg.categoria_comunicacao,
                        'metodologia_treinamento': reg.metodologia_treinamento,
                        'area_conhecimento': reg.area_conhecimento,
                        'facilitador_fornecedor': reg.facilitador_fornecedor,
                        'carga_horaria': reg.carga_horaria,
                        'custo_treinamento': reg.custo_treinamento,
                        'data_final_treinamento': _safe_date(reg.data_final_treinamento),
                        'mes_referencia': reg.mes_referencia,
                        'necessita_avaliacao_eficacia': reg.necessita_avaliacao_eficacia,
                        'data_limite_avaliacao_eficacia': _safe_date(reg.data_limite_avaliacao_eficacia),
                        'participante_externo': getattr(reg, 'participante_externo', None)
                    }

                    obj, created = RegistroTreinamento.objects.update_or_create(
                        **filter_kwargs,
                        defaults=defaults
                    )
                    if created:
                        resultados['criados'] += 1
                    else:
                        resultados['atualizados'] += 1
            except Exception as e:
                resultados['erros'] += 1
                resultados['mensagens_erro'].append(f"Erro ao criar/atualizar registro: {str(e)}")
                logger.exception(f"Erro ao criar/atualizar registro: {str(e)}")
    
    # OTIMIZAÇÃO: Atualizar registros em batch
    logger.info(f"[IMPORT] Atualizando {len(registros_para_atualizar)} registros...")
    if registros_para_atualizar:
        for item in registros_para_atualizar:
            try:
                # Montar filtro adequado (colaborador ou participante_externo)
                if item.get('procedimento'):
                    # Incluir data_treinamento para selecionar o registro correto (mesmo procedimento em datas diferentes)
                    data_sel = item['dados'].get('data_treinamento')
                    if item.get('colaborador'):
                        filter_kwargs = {'colaborador': item['colaborador'], 'procedimento': item['procedimento'], 'data_treinamento': data_sel}
                    else:
                        filter_kwargs = {'participante_externo': item.get('participante_externo'), 'procedimento': item['procedimento'], 'data_treinamento': data_sel}
                else:
                    if item.get('colaborador'):
                        filter_kwargs = {'colaborador': item['colaborador'], 'titulo_treinamento': item['titulo_treinamento'], 'data_treinamento': item['data_treinamento']}
                    else:
                        filter_kwargs = {'participante_externo': item.get('participante_externo'), 'titulo_treinamento': item['titulo_treinamento'], 'data_treinamento': item['data_treinamento']}

                reg = RegistroTreinamento.objects.get(**filter_kwargs)
                for chave, valor in item['dados'].items():
                    setattr(reg, chave, valor)
                reg.save()
                resultados['atualizados'] += 1
            except Exception as e:
                resultados['erros'] += 1
                resultados['mensagens_erro'].append(f"Erro ao atualizar registro: {str(e)}")
                logger.exception(f"Erro ao atualizar registro: {str(e)}")
    
    # Atualizar demandas após importação bem-sucedida
    if resultados['criados'] > 0 or resultados['atualizados'] > 0:
        logger.info(f"[IMPORT] Atualizando demandas...")
        atualizar_demandas_apos_importacao()
    
    logger.info(f"[IMPORT] Concluído! Criados: {resultados['criados']}, Atualizados: {resultados['atualizados']}, Erros: {resultados['erros']}")
    return resultados

# ============================================================================
# GERAÇÃO DE LISTAS DE PRESENÇA COM TEMPLATE
# ============================================================================

@login_required
def upload_template_lista_presenca(request):
    """Upload de PDF template para lista de presença"""
    
    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao')
        arquivo_pdf = request.FILES.get('arquivo_pdf_template')
        tem_pagina_assinatura = request.POST.get('tem_pagina_assinatura') == 'on'
        num_linhas_assinatura = request.POST.get('num_linhas_assinatura', '20')
        
        if not all([nome, arquivo_pdf]):
            messages.error(request, 'Preencha todos os campos obrigatórios.')
            return render(request, 'procedures/upload_template.html')
        
        try:
            template = TemplateListaPresenca.objects.create(
                nome=nome,
                descricao=descricao,
                arquivo_pdf_template=arquivo_pdf,
                tem_pagina_assinatura=tem_pagina_assinatura,
                num_linhas_assinatura=int(num_linhas_assinatura),
                tipo_arquivo='pdf'
            )
            messages.success(request, f'Template "{template.nome}" enviado com sucesso! Agora mapeie os placeholders.')
            return redirect('procedures:mapear_template_fields', template_id=template.id)
        except Exception as e:
            messages.error(request, f'Erro ao enviar template: {str(e)}')
    
    context = {
        'tipo_arquivo_choices': TemplateListaPresenca.TIPO_ARQUIVO,
    }
    return render(request, 'procedures/upload_template.html', context)


@login_required
def upload_pdf_template(request, template_id):
    """API para fazer upload de novo PDF do template"""
    from django.http import JsonResponse
    
    template = get_object_or_404(TemplateListaPresenca, id=template_id)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
    
    arquivo_pdf = request.FILES.get('arquivo_pdf_template')
    
    if not arquivo_pdf:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo selecionado'})
    
    if not arquivo_pdf.name.lower().endswith('.pdf'):
        return JsonResponse({'success': False, 'error': 'Por favor, envie um arquivo PDF válido'})
    
    try:
        # Deletar arquivo anterior se existir
        if template.arquivo_pdf_template:
            template.arquivo_pdf_template.delete()
        
        # Salvar novo arquivo
        template.arquivo_pdf_template = arquivo_pdf
        template.save()
        
        return JsonResponse({'success': True, 'message': 'PDF enviado com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao carregar PDF: {str(e)}'})


@login_required
def remove_pdf_template(request, template_id):
    """API para remover PDF do template"""
    from django.http import JsonResponse
    
    template = get_object_or_404(TemplateListaPresenca, id=template_id)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)
    
    try:
        if template.arquivo_pdf_template:
            template.arquivo_pdf_template.delete()
            template.arquivo_pdf_template = None
            template.save()
            return JsonResponse({'success': True, 'message': 'PDF removido com sucesso!'})
        else:
            return JsonResponse({'success': False, 'error': 'Nenhum PDF para remover'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao remover PDF: {str(e)}'})


def extract_placeholders_from_pdf(pdf_path):
    """Extrai placeholders do PDF procurando por padrões {{placeholder}} ou retorna uma lista vazia se não encontrar"""
    import PyPDF2
    import re
    
    placeholders = set()
    
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                try:
                    text = page.extract_text()
                    # Procurar padrões {{palavra}} ou {{palavra_com_underscore}}
                    matches = re.findall(r'\{\{(\w+(?:_\w+)*)\}\}', text)
                    placeholders.update(matches)
                except Exception:
                    logger.warning(
                        "Falha ao extrair texto da pagina %s do PDF %s", page_num, pdf_path,
                        exc_info=True,
                    )
                    continue
    except Exception as e:
        logger.error("Erro ao extrair placeholders do PDF %s: %s", pdf_path, e, exc_info=True)
    
    return sorted(list(placeholders)) if placeholders else []


@login_required
def mapear_template_fields(request, template_id):
    """Interface para mapear placeholders do PDF para campos de dados - agora com suporte a click manual"""
    
    template = get_object_or_404(TemplateListaPresenca, id=template_id)
    mapeamentos_existentes = MapeamentoCampoListaPresenca.objects.filter(template=template)
    
    if request.method == 'POST':
        # Processar mapeamentos de placeholders (novo formato com campo_dados[placeholder])
        campo_dados_dict = {}
        for key, value in request.POST.items():
            if key.startswith('campo_dados[') and key.endswith(']'):
                placeholder = key[12:-1]  # Extrai placeholder de "campo_dados[placeholder]"
                campo_dados_dict[placeholder] = value
        
        # Validar que todos os placeholders foram mapeados
        if not campo_dados_dict:
            messages.error(request, 'Nenhum placeholder foi mapeado. Por favor, mapeie pelo menos um campo.')
            return redirect('procedures:mapear_template_fields', template_id=template_id)
        
        # Limpar mapeamentos anteriores
        MapeamentoCampoListaPresenca.objects.filter(template=template).delete()
        
        # Criar novos mapeamentos
        mapeamentos_criados = 0
        for placeholder, campo_dados in campo_dados_dict.items():
            if placeholder and campo_dados:
                try:
                    MapeamentoCampoListaPresenca.objects.create(
                        template=template,
                        placeholder=placeholder,
                        campo_dados=campo_dados,
                    )
                    mapeamentos_criados += 1
                except Exception as e:
                    messages.warning(request, f'Erro ao mapear placeholder {placeholder}: {str(e)}')
        
        # Atualizar status de mapeamento completo
        template.mapeamento_completo = mapeamentos_criados > 0
        template.save()
        
        messages.success(request, f'Mapeamento de {mapeamentos_criados} campo(s) salvo para "{template.nome}"!')
        return redirect('procedures:gerenciar_templates_presenca')
    
    # Extrair placeholders dinamicamente do PDF se existir
    placeholders = []
    mapeamento_dict = {}
    has_pdf = False
    
    if template.arquivo_pdf_template:
        has_pdf = True
        try:
            pdf_path = template.arquivo_pdf_template.path
            placeholders = extract_placeholders_from_pdf(pdf_path)
        except Exception as e:
            messages.warning(request, f'Erro ao extrair placeholders: {str(e)}')
            placeholders = []
        
        # Se não encontrou placeholders {{}} automáticos, permitir mapeamento manual
        if not placeholders:
            # Usar placeholders pré-definidos como sugestão para mapeamento manual
            placeholders = [
                'titulo', 'facilitador', 'data', 'hora_inicio',
                'hora_fim', 'carga_horaria', 'local', 'procedimentos'
            ]
    
    # Criar dicionário para lookup rápido de mapeamentos
    for mapeamento in mapeamentos_existentes:
        mapeamento_dict[mapeamento.placeholder] = mapeamento
    
    context = {
        'template': template,
        'placeholders': placeholders,
        'mapeamentos_existentes': mapeamentos_existentes,
        'mapeamento_dict': mapeamento_dict,
        'campos_disponiveis': MapeamentoCampoListaPresenca.CAMPOS_DISPONIVEIS,
        'has_pdf': has_pdf,
    }
    return render(request, 'procedures/mapear_template_fields.html', context)


@login_required
def serve_pdf_template(request, template_id):
    """Serve o arquivo PDF do template para visualização"""
    from django.http import FileResponse, Http404
    
    template = get_object_or_404(TemplateListaPresenca, id=template_id)
    
    if not template.arquivo_pdf_template:
        raise Http404("PDF não encontrado")
    
    try:
        response = FileResponse(template.arquivo_pdf_template.open('rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{template.arquivo_pdf_template.name}"'
        return response
    except Exception as e:
        raise Http404(f"Erro ao servir PDF: {str(e)}")


@login_required
def selecionar_template_lista_presenca(request):
    """Seleciona planejamento e template para gerar lista de presença"""
    
    if request.method == 'POST':
        planejamento_id = request.POST.get('planejamento_id')
        template_id = request.POST.get('template_id')
        
        if not planejamento_id or not template_id:
            messages.error(request, 'Selecione um planejamento e um template.')
            return redirect('procedures:selecionar_template_lista_presenca')
        
        try:
            planejamento = PlanejamentoTreinamento.objects.get(id=planejamento_id)
            template = TemplateListaPresenca.objects.get(id=template_id, ativo=True)
        except (PlanejamentoTreinamento.DoesNotExist, TemplateListaPresenca.DoesNotExist):
            messages.error(request, 'Planejamento ou template não encontrado.')
            return redirect('procedures:selecionar_template_lista_presenca')
        
        return redirect('procedures:gerar_lista_presenca_pdf',
                       planejamento_id=planejamento_id,
                       template_id=template_id)
    
    context = {
        'planejamentos': PlanejamentoTreinamento.objects.filter(
            status__in=['pendente', 'agendado']
        ).order_by('-data_prevista'),
        'templates': TemplateListaPresenca.objects.filter(ativo=True),
    }
    return render(request, 'procedures/lista_presenca_selecionar_template.html', context)


@login_required
def gerar_lista_presenca_desde_planejamento(request, planejamento_id):
    """Formulário simplificado para gerar lista de presença a partir de um planejamento"""
    
    planejamento = get_object_or_404(PlanejamentoTreinamento, id=planejamento_id)
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        
        if not template_id:
            messages.error(request, 'Selecione um template.')
            return redirect('procedures:gerar_lista_presenca_desde_planejamento', planejamento_id=planejamento_id)
        
        try:
            template = TemplateListaPresenca.objects.get(id=template_id, ativo=True)
        except TemplateListaPresenca.DoesNotExist:
            messages.error(request, 'Template não encontrado.')
            return redirect('procedures:gerar_lista_presenca_desde_planejamento', planejamento_id=planejamento_id)
        
        return redirect('procedures:gerar_lista_presenca_pdf',
                       planejamento_id=planejamento_id,
                       template_id=template_id)
    
    context = {
        'planejamento': planejamento,
        'templates': TemplateListaPresenca.objects.filter(ativo=True),
    }
    return render(request, 'procedures/gerar_lista_presenca_desde_planejamento.html', context)


@login_required
def upload_lista_presenca_assinada(request, pk):
    """Upload da lista de presença assinada (evidência documental)"""
    
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo_assinado')
        
        if not arquivo:
            messages.error(request, 'Selecione um arquivo.')
            return redirect('procedures:upload_lista_presenca_assinada', pk=pk)
        
        # Validar tipo de arquivo
        extensao = os.path.splitext(arquivo.name)[1].lower()
        tipos_permitidos = ['.pdf', '.jpg', '.jpeg', '.png', '.tiff', '.tif']
        
        if extensao not in tipos_permitidos:
            messages.error(request, f'Tipo de arquivo não permitido. Use: {", ".join(tipos_permitidos)}')
            return redirect('procedures:upload_lista_presenca_assinada', pk=pk)
        
        # Validar tamanho (máximo 50MB)
        if arquivo.size > 50 * 1024 * 1024:
            messages.error(request, 'Arquivo muito grande. Máximo 50 MB.')
            return redirect('procedures:upload_lista_presenca_assinada', pk=pk)
        
        try:
            # Remover arquivo anterior se existir
            if lista.arquivo_assinado:
                try:
                    if os.path.exists(lista.arquivo_assinado.path):
                        os.remove(lista.arquivo_assinado.path)
                except OSError:
                    logger.warning(
                        "Falha ao remover arquivo assinado anterior da lista %s", lista.pk,
                        exc_info=True,
                    )
            
            # Salvar novo arquivo
            lista.arquivo_assinado = arquivo
            lista.data_upload_assinado = timezone.now()
            lista.save()
            
            messages.success(request, f'✅ Arquivo "{arquivo.name}" enviado com sucesso como evidência documental!')
            return redirect('procedures:lista_presenca_detail', pk=pk)
        
        except Exception as e:
            messages.error(request, f'Erro ao enviar arquivo: {str(e)}')
            return redirect('procedures:upload_lista_presenca_assinada', pk=pk)
    
    context = {
        'lista': lista,
    }
    return render(request, 'procedures/upload_lista_presenca_assinada.html', context)


@login_required
def remover_lista_presenca_assinada(request, pk):
    """Remove o arquivo da lista de presença assinada"""
    
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    if request.method == 'POST':
        try:
            if lista.arquivo_assinado:
                try:
                    if os.path.exists(lista.arquivo_assinado.path):
                        os.remove(lista.arquivo_assinado.path)
                except OSError:
                    logger.warning(
                        "Falha ao remover arquivo assinado da lista %s", lista.pk,
                        exc_info=True,
                    )
                
                lista.arquivo_assinado = None
                lista.data_upload_assinado = None
                lista.save()
                messages.success(request, '✅ Arquivo removido com sucesso!')
            else:
                messages.info(request, 'Esta lista não possui arquivo assinado.')
        
        except Exception as e:
            messages.error(request, f'Erro ao remover arquivo: {str(e)}')
    
    return redirect('procedures:lista_presenca_detail', pk=pk)


@login_required
def visualizar_lista_presenca_assinada(request, pk):
    """Visualiza o arquivo da lista de presença assinada"""
    
    lista = get_object_or_404(ListaPresenca, pk=pk)
    
    if not lista.arquivo_assinado:
        messages.error(request, 'Esta lista não possui arquivo assinado.')
        return redirect('procedures:lista_presenca_detail', pk=pk)
    
    try:
        arquivo_path = lista.arquivo_assinado.path
        with open(arquivo_path, 'rb') as arquivo:
            response = HttpResponse(arquivo.read())
            response['Content-Type'] = 'application/octet-stream'
            response['Content-Disposition'] = f'inline; filename="{lista.arquivo_assinado.name}"'
            return response
    
    except Exception as e:
        messages.error(request, f'Erro ao abrir arquivo: {str(e)}')
        return redirect('procedures:lista_presenca_detail', pk=pk)



@login_required
def gerar_lista_presenca_pdf(request, planejamento_id, template_id):
    """Gera lista de presença a partir do planejamento e template"""
    
    planejamento = get_object_or_404(PlanejamentoTreinamento, id=planejamento_id)
    template = get_object_or_404(TemplateListaPresenca, id=template_id, ativo=True)
    mapeamentos = MapeamentoCampoListaPresenca.objects.filter(template=template)
    
    if request.method == 'POST':
        # Criar lista de presença com informações do formulário
        data_sessao = request.POST.get('data_sessao', planejamento.data_prevista)
        hora_inicio = request.POST.get('hora_inicio', '')
        hora_fim = request.POST.get('hora_fim', '')
        local = request.POST.get('local', planejamento.local)
        
        try:
            lista = ListaPresenca.objects.create(
                titulo=planejamento.titulo,
                instrutor=planejamento.instrutor,
                instrutor_nome=planejamento.instrutor.nome_completo if planejamento.instrutor else '',
                data_sessao=data_sessao,
                hora_inicio=hora_inicio or None,
                hora_fim=hora_fim or None,
                carga_horaria=planejamento.carga_horaria,
                local=local,
                template=template,  # ADICIONAR O TEMPLATE AQUI
                criado_por=request.user,
            )
            
            # Adicionar procedimentos do planejamento à lista
            for proc in planejamento.procedimentos.all():
                # Assumindo que existe um modelo RegistroTreinamento que vincula procedimentos
                pass
            
            messages.success(request, 'Lista de presença criada com sucesso!')
            return redirect('procedures:lista_presenca_detail', pk=lista.id)
        except Exception as e:
            messages.error(request, f'Erro ao criar lista de presença: {str(e)}')
    
    context = {
        'planejamento': planejamento,
        'template': template,
        'mapeamentos': mapeamentos,
    }
    return render(request, 'procedures/gerar_lista_presenca.html', context)


@login_required
def gerenciar_templates_presenca_view(request):
    """
    Gerenciamento central de templates de listas de presença.
    Lista templates, permite criar novo, editar e deletar.
    """
    # Buscar todos os templates
    templates = TemplateListaPresenca.objects.all().order_by('-criado_em')
    
    # Adicionar informações sobre mapeamento
    for template in templates:
        # Contar placeholders mapeados (com campo_dados definido)
        template.total_campos_mapeados = template.mapeamentos.filter(
            campo_dados__isnull=False
        ).exclude(campo_dados='').count()
        template.total_campos_obrigatorios = len(template.get_placeholders_list())
        # Calcular porcentagem de progresso para a barra visual
        if template.total_campos_obrigatorios > 0:
            template.progresso_porcentagem = int((template.total_campos_mapeados / template.total_campos_obrigatorios) * 100)
        else:
            template.progresso_porcentagem = 0
    
    # Buscar ações recentes
    acao = request.GET.get('acao')
    mensagem_sucesso = None
    
    if request.method == 'POST':
        acao_post = request.POST.get('acao')
        
        if acao_post == 'novo':
            # Criar novo template
            nome = request.POST.get('nome', '').strip()
            descricao = request.POST.get('descricao', '').strip()
            
            if not nome:
                messages.error(request, '⚠️ Nome do template é obrigatório.')
            else:
                template = TemplateListaPresenca.objects.create(
                    nome=nome,
                    descricao=descricao,
                )
                messages.success(request, f'✅ Template "{nome}" criado com sucesso!')
                return redirect(f"{reverse('procedures:gerenciar_templates_presenca')}?acao=novo")
        
        elif acao_post == 'deletar':
            # Deletar template
            template_id = request.POST.get('template_id')
            try:
                template = TemplateListaPresenca.objects.get(id=template_id)
                nome = template.nome
                template.delete()
                messages.success(request, f'✅ Template "{nome}" deletado com sucesso!')
            except TemplateListaPresenca.DoesNotExist:
                messages.error(request, '⚠️ Template não encontrado.')
            return redirect('procedures:gerenciar_templates_presenca')
    
    context = {
        'templates': templates,
        'acao': acao,
        'total_templates': templates.count(),
    }
    return render(request, 'procedures/gerenciar_templates_presenca.html', context)


@login_required
def api_procedimentos_json_view(request):
    """API para carregar lista de procedimentos em JSON."""
    from django.http import JsonResponse
    
    procedimentos = Procedimento.objects.values('id', 'codigo', 'nome').order_by('codigo')
    return JsonResponse(list(procedimentos), safe=False)


@login_required
def api_colaboradores_json_view(request):
    """API para carregar lista de colaboradores em JSON."""
    from django.http import JsonResponse
    
    colaboradores = Colaborador.objects.values('id', 'nome_completo').order_by('nome_completo')
    return JsonResponse([
        {'id': c['id'], 'nome': c['nome_completo']}
        for c in colaboradores
    ], safe=False)


@login_required
def api_colaboradores_busca_view(request):
    """API para buscar colaboradores com filtro."""
    from django.http import JsonResponse
    from django.db.models import Q
    
    termo = request.GET.get('q', '').strip()
    
    if len(termo) < 2:
        return JsonResponse([], safe=False)
    
    colaboradores = Colaborador.objects.filter(
        Q(nome_completo__icontains=termo) | 
        Q(matricula__icontains=termo)
    ).values('id', 'nome_completo').order_by('nome_completo')[:20]  # Limitar a 20 resultados
    
    return JsonResponse(list(colaboradores), safe=False)


@login_required
def api_procedimentos_busca_view(request):
    """API para buscar procedimentos com filtro."""
    from django.http import JsonResponse
    from django.db.models import Q
    
    termo = request.GET.get('q', '').strip()
    
    if len(termo) < 1:
        return JsonResponse([], safe=False)
    
    procedimentos = Procedimento.objects.filter(
        Q(codigo__icontains=termo) | 
        Q(nome__icontains=termo)
    ).values('id', 'codigo', 'nome').order_by('codigo')[:20]  # Limitar a 20 resultados
    
    return JsonResponse(list(procedimentos), safe=False)


@login_required
def lista_presenca_export_view(request):
    """Exporta todas as listas de presença (com filtros aplicados) para Excel."""
    from datetime import datetime
    from django.db.models import Q
    
    # Aplicar os mesmos filtros da listagem
    listas = ListaPresenca.objects.all()
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    instrutor_id = request.GET.get('instrutor')
    busca = request.GET.get('busca')
    
    if data_inicio:
        listas = listas.filter(data_sessao__gte=data_inicio)
    if data_fim:
        listas = listas.filter(data_sessao__lte=data_fim)
    if instrutor_id:
        listas = listas.filter(instrutor_id=instrutor_id)
    if busca:
        listas = listas.filter(
            Q(codigo__icontains=busca) |
            Q(titulo__icontains=busca) |
            Q(local__icontains=busca) |
            Q(instrutor_nome__icontains=busca)
        )
    
    listas = listas.order_by('-data_sessao', '-codigo')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listas de Presença"
    
    # Definir estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Adicionar cabeçalho
    headers = ['Código', 'Título', 'Data Sessão', 'Instrutor', 'Local', 'Carga Horária', 'Participantes', 'Registros', 'Observações']
    ws.append(headers)
    
    # Formatar cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Definir largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30
    
    # Adicionar dados
    for lista in listas:
        # Contar participantes únicos e registros
        participantes_unicos = lista.registrotreinamento_set.values('colaborador_id').distinct().count()
        registros_count = lista.registrotreinamento_set.count()
        
        # Instrutor (preferir nome do BD, depois texto livre)
        instrutor_nome = lista.instrutor.nome_completo if lista.instrutor else lista.instrutor_nome or '—'
        
        row = [
            lista.codigo,
            lista.titulo,
            lista.data_sessao.strftime('%d/%m/%Y') if lista.data_sessao else '—',
            instrutor_nome,
            lista.local or '—',
            str(lista.carga_horaria) if lista.carga_horaria else '—',
            participantes_unicos,
            registros_count,
            lista.observacoes or '—'
        ]
        ws.append(row)
        
        # Formatar linha
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Adicionar linha de resumo
    ws.append([''] * 9)  # Linha vazia
    ws.append([f'Total de Listas: {listas.count()}'] + [''] * 8)
    
    # Freezar primeira linha (cabeçalho)
    ws.freeze_panes = 'A2'
    
    # Gerar nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'listas_presenca_{timestamp}.xlsx'
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    return response
