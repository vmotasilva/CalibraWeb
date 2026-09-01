# -*- coding: utf-8 -*-
from datetime import date, timedelta
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from collections import defaultdict
from procedures.models import RegistroTreinamento, ColaboradorPerfil, MatrizProcedimento, Procedimento
from rh.models import Colaborador
from organization.models import Setor

@login_required
def avaliacao_eficacia_list_view(request):
    """
    Lista os treinamentos de procedimentos críticos para controle da autoavaliação de eficácia (a partir de 30 dias).
    """
    # Filtrar apenas registros que possuem colaborador e procedimento,
    # onde o procedimento é crítico, o treinamento já ocorreu (data_treinamento is not null),
    # o colaborador está ativo e o registro de treinamento está ativo.
    qs = RegistroTreinamento.objects.select_related(
        'colaborador', 'procedimento', 'colaborador__setor', 'colaborador__lider', 'colaborador__supervisor', 'colaborador__gerente', 'gestor_responsavel'
    ).filter(
        colaborador__isnull=False,
        procedimento__isnull=False,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
        colaborador__is_active=True,
        colaborador__afastado=False,
        colaborador__em_ferias=False,
    )

    # Obter dados para os filtros
    setores_list = Setor.objects.order_by('nome')
    gestores_list = Colaborador.objects.filter(is_active=True).filter(
        Q(posto_lideranca__in=['LIDER', 'SUPERVISOR', 'GERENTE', 'QUALIDADE', 'PROCESSOS', 'MANUTENCAO', 'EHS']) |
        Q(id__in=Colaborador.objects.values_list('lider_id', flat=True)) |
        Q(id__in=Colaborador.objects.values_list('supervisor_id', flat=True)) |
        Q(id__in=Colaborador.objects.values_list('gerente_id', flat=True))
    ).distinct().order_by('nome_completo')
    procedimentos_list = Procedimento.objects.filter(criticidade='CRITICO').order_by('codigo', 'nome')
    matrizes_opcoes = MatrizProcedimento.objects.filter(ativo=True).order_by('nome')

    # Parâmetros de filtro
    busca = (request.GET.get('q') or request.GET.get('busca') or '').strip()
    status_filtro = (request.GET.get('status') or '').strip()
    gestor_id = (request.GET.get('gestor') or request.GET.get('lider') or '').strip()
    setor_id = (request.GET.get('setor') or '').strip()
    procedimento_id = (request.GET.get('procedimento') or request.GET.get('procedimento_id') or '').strip()
    dias_decorridos_filtro = (request.GET.get('dias') or request.GET.get('dias_decorridos') or '').strip()
    vinculo_filtro = (request.GET.get('vinculo') or '').strip()
    posterior_filtro = (request.GET.get('posterior') or '').strip()
    matriz_filtro = (request.GET.get('matriz') or '').strip()
    exibir = (request.GET.get('exibir') or '30').strip()

    # Filtro especial por dias decorridos (antes de filtrar status)
    today = date.today()
    carencia_date = today - timedelta(days=30)

    # Aplicar filtros no banco
    if busca:
        qs = qs.filter(
            Q(colaborador__nome_completo__icontains=busca) |
            Q(colaborador__matricula__icontains=busca) |
            Q(procedimento__codigo__icontains=busca) |
            Q(procedimento__nome__icontains=busca)
        )
    if gestor_id:
        qs = qs.filter(
            Q(gestor_responsavel_id=gestor_id) |
            (Q(gestor_responsavel__isnull=True) & (
                Q(colaborador__lider_id=gestor_id) |
                Q(colaborador__supervisor_id=gestor_id) |
                Q(colaborador__gerente_id=gestor_id)
            ))
        )
    if setor_id:
        qs = qs.filter(colaborador__setor_id=setor_id)
    if procedimento_id:
        qs = qs.filter(procedimento_id=procedimento_id)
    if matriz_filtro:
        qs = qs.filter(procedimento__matriz__iexact=matriz_filtro)

    if dias_decorridos_filtro:
        if dias_decorridos_filtro == 'menos_30':
            qs = qs.filter(data_treinamento__gt=carencia_date)
        elif dias_decorridos_filtro == '30_60':
            qs = qs.filter(data_treinamento__lte=carencia_date, data_treinamento__gte=today - timedelta(days=60))
        elif dias_decorridos_filtro == '60_90':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=60), data_treinamento__gte=today - timedelta(days=90))
        elif dias_decorridos_filtro == '90_mais':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=90))

    if status_filtro:
        if status_filtro == 'EFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='EFICAZ')
        elif status_filtro == 'INEFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='INEFICAZ')
        elif status_filtro == 'NAO_APLICA':
            qs = qs.filter(avaliacao_eficacia_status='NAO_APLICA')
        elif status_filtro == 'CARENCIA':
            # PENDENTE mas com menos de 30 dias desde o treinamento
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__gt=carencia_date
            )
        elif status_filtro == 'PENDENTE':
            # PENDENTE e com 30 dias ou mais desde o treinamento
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__lte=carencia_date
            )

    # Ordenação padrão: por data do treinamento decrescente
    qs = qs.order_by('-data_treinamento', 'colaborador__nome_completo')

    # Calcular totais para os cards do dashboard
    # Precisamos da query com filtros de busca, lider/gestor, setor, procedimento e dias (mas sem filtro de status) para os totais locais
    totals_qs = RegistroTreinamento.objects.filter(
        colaborador__isnull=False,
        procedimento__isnull=False,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
        colaborador__is_active=True,
        colaborador__afastado=False,
        colaborador__em_ferias=False,
    )
    if busca:
        totals_qs = totals_qs.filter(
            Q(colaborador__nome_completo__icontains=busca) |
            Q(colaborador__matricula__icontains=busca) |
            Q(procedimento__codigo__icontains=busca) |
            Q(procedimento__nome__icontains=busca)
        )
    if gestor_id:
        totals_qs = totals_qs.filter(
            Q(gestor_responsavel_id=gestor_id) |
            (Q(gestor_responsavel__isnull=True) & (
                Q(colaborador__lider_id=gestor_id) |
                Q(colaborador__supervisor_id=gestor_id) |
                Q(colaborador__gerente_id=gestor_id)
            ))
        )
    if setor_id:
        qs = qs.filter(colaborador__setor_id=setor_id)
    if procedimento_id:
        qs = qs.filter(procedimento_id=procedimento_id)
    if matriz_filtro:
        qs = qs.filter(procedimento__matriz__iexact=matriz_filtro)

    if dias_decorridos_filtro:
        if dias_decorridos_filtro == 'menos_30':
            qs = qs.filter(data_treinamento__gt=carencia_date)
        elif dias_decorridos_filtro == '30_60':
            qs = qs.filter(data_treinamento__lte=carencia_date, data_treinamento__gte=today - timedelta(days=60))
        elif dias_decorridos_filtro == '60_90':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=60), data_treinamento__gte=today - timedelta(days=90))
        elif dias_decorridos_filtro == '90_mais':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=90))

    if status_filtro:
        if status_filtro == 'EFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='EFICAZ')
        elif status_filtro == 'INEFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='INEFICAZ')
        elif status_filtro == 'NAO_APLICA':
            qs = qs.filter(avaliacao_eficacia_status='NAO_APLICA')
        elif status_filtro == 'CARENCIA':
            # PENDENTE mas com menos de 30 dias desde o treinamento
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__gt=carencia_date
            )
        elif status_filtro == 'PENDENTE':
            # PENDENTE e com 30 dias ou mais desde o treinamento
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__lte=carencia_date
            )

    # Ordenação padrão: por data do treinamento decrescente
    qs = qs.order_by('-data_treinamento', 'colaborador__nome_completo')

    # Calcular totais para os cards do dashboard
    # Precisamos da query com filtros de busca, lider/gestor, setor, procedimento e dias (mas sem filtro de status) para os totais locais
    totals_qs = RegistroTreinamento.objects.filter(
        colaborador__isnull=False,
        procedimento__isnull=False,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
        colaborador__is_active=True,
        colaborador__afastado=False,
        colaborador__em_ferias=False,
    )
    if busca:
        totals_qs = totals_qs.filter(
            Q(colaborador__nome_completo__icontains=busca) |
            Q(colaborador__matricula__icontains=busca) |
            Q(procedimento__codigo__icontains=busca) |
            Q(procedimento__nome__icontains=busca)
        )
    if gestor_id:
        totals_qs = totals_qs.filter(
            Q(colaborador__lider_id=gestor_id) |
            Q(colaborador__supervisor_id=gestor_id) |
            Q(colaborador__gerente_id=gestor_id)
        )
    if setor_id:
        totals_qs = totals_qs.filter(colaborador__setor_id=setor_id)
    if procedimento_id:
        totals_qs = totals_qs.filter(procedimento_id=procedimento_id)
    if matriz_filtro:
        totals_qs = totals_qs.filter(procedimento__matriz__iexact=matriz_filtro)
    if dias_decorridos_filtro:
        if dias_decorridos_filtro == 'menos_30':
            totals_qs = totals_qs.filter(data_treinamento__gt=carencia_date)
        elif dias_decorridos_filtro == '30_60':
            totals_qs = totals_qs.filter(data_treinamento__lte=carencia_date, data_treinamento__gte=today - timedelta(days=60))
        elif dias_decorridos_filtro == '60_90':
            totals_qs = totals_qs.filter(data_treinamento__lt=today - timedelta(days=60), data_treinamento__gte=today - timedelta(days=90))
        elif dias_decorridos_filtro == '90_mais':
            totals_qs = totals_qs.filter(data_treinamento__lt=today - timedelta(days=90))

    # Obter todos os registros de treinamentos críticos ativos para calcular se há posterior em aberto
    all_critical = list(RegistroTreinamento.objects.filter(
        colaborador__isnull=False,
        procedimento__isnull=False,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
        colaborador__is_active=True,
    ))

    # Group by (colaborador_id, procedimento_id)
    grouped = defaultdict(list)
    for t in all_critical:
        grouped[(t.colaborador_id, t.procedimento_id)].append(t)

    # Sort each group by data_treinamento ascending
    for key in grouped:
        grouped[key].sort(key=lambda x: x.data_treinamento)

    # Find which ones have a posterior training with status PENDENTE / None
    posterior_pending_ids = set()
    for key, group in grouped.items():
        n = len(group)
        for i in range(n):
            t = group[i]
            has_posterior = False
            for j in range(i + 1, n):
                t2 = group[j]
                if t2.avaliacao_eficacia_status in ['PENDENTE', None, '']:
                    has_posterior = True
                    break
            if has_posterior:
                posterior_pending_ids.add(t.id)

    # Mapeamento do vínculo de perfil dos colaboradores
    colab_procs_map = defaultdict(set)
    for cp in ColaboradorPerfil.objects.filter(ativo=True):
        proc_ids = cp.get_procedimentos_necessarios().values_list('id', flat=True)
        colab_procs_map[cp.colaborador_id].update(proc_ids)

    # Filtrar e anotar no python (vínculo de perfil e posterior pendente)
    registros_list = list(qs)
    for t in registros_list:
        t.tem_vinculo_perfil = t.procedimento_id in colab_procs_map[t.colaborador_id]
        t.has_posterior_pending = t.id in posterior_pending_ids

    if vinculo_filtro == 'COM_VINCULO':
        registros_list = [t for t in registros_list if t.tem_vinculo_perfil]
    elif vinculo_filtro == 'SEM_VINCULO':
        registros_list = [t for t in registros_list if not t.tem_vinculo_perfil]

    if posterior_filtro == 'COM_POSTERIOR':
        registros_list = [t for t in registros_list if t.has_posterior_pending]
    elif posterior_filtro == 'SEM_POSTERIOR':
        registros_list = [t for t in registros_list if not t.has_posterior_pending]

    totals_list = list(totals_qs)
    for t in totals_list:
        t.tem_vinculo_perfil = t.procedimento_id in colab_procs_map[t.colaborador_id]
        t.has_posterior_pending = t.id in posterior_pending_ids

    if vinculo_filtro == 'COM_VINCULO':
        totals_list = [t for t in totals_list if t.tem_vinculo_perfil]
    elif vinculo_filtro == 'SEM_VINCULO':
        totals_list = [t for t in totals_list if not t.tem_vinculo_perfil]

    if posterior_filtro == 'COM_POSTERIOR':
        totals_list = [t for t in totals_list if t.has_posterior_pending]
    elif posterior_filtro == 'SEM_POSTERIOR':
        totals_list = [t for t in totals_list if not t.has_posterior_pending]

    total_geral = len(totals_list)
    total_eficaz = sum(1 for t in totals_list if t.avaliacao_eficacia_status == 'EFICAZ')
    total_ineficaz = sum(1 for t in totals_list if t.avaliacao_eficacia_status == 'INEFICAZ')
    total_nao_aplica = sum(1 for t in totals_list if t.avaliacao_eficacia_status == 'NAO_APLICA')
    total_carencia = sum(1 for t in totals_list if (t.avaliacao_eficacia_status in ['PENDENTE', None, '']) and (today - t.data_treinamento).days < 30)
    total_pendente = sum(1 for t in totals_list if (t.avaliacao_eficacia_status in ['PENDENTE', None, '']) and (today - t.data_treinamento).days >= 30)

    ordenacao = (request.GET.get('ordenacao') or '').strip()
    ordenacao_niveis = []
    if ordenacao:
        for part in ordenacao.split(','):
            part = part.strip()
            if ':' in part:
                f_name, f_dir = part.split(':', 1)
                f_name = f_name.strip()
                f_dir = f_dir.strip().lower()
                if f_name and f_dir in ['asc', 'desc']:
                    ordenacao_niveis.append({'campo': f_name, 'direcao': f_dir})

    # Processar cada objeto para adicionar informações calculadas antes da ordenação
    for t in registros_list:
        t.data_elegibilidade = t.data_treinamento + timedelta(days=30) if t.data_treinamento else date.min
        dias_dec = (today - t.data_treinamento).days if t.data_treinamento else 0
        t.dias_decorridos = dias_dec
        t.dias_restantes = max(0, 30 - dias_dec)
        
        # Determinar status para exibição
        if t.avaliacao_eficacia_status == 'EFICAZ':
            t.status_display = 'Eficaz'
            t.status_class = 'success'
        elif t.avaliacao_eficacia_status == 'INEFICAZ':
            t.status_display = 'Ineficaz'
            t.status_class = 'danger'
        elif t.avaliacao_eficacia_status == 'NAO_APLICA':
            t.status_display = 'Não se Aplica'
            t.status_class = 'secondary'
        else:
            if dias_dec < 30:
                t.status_display = 'Carência'
                t.status_class = 'info'
            else:
                t.status_display = 'Pendente'
                t.status_class = 'warning'

        # Nome do gestor
        if t.gestor_responsavel:
            t.gestor_nome = t.gestor_responsavel.nome_completo
        elif t.colaborador.posto_lideranca == 'SUPERVISOR':
            t.gestor_nome = t.colaborador.gerente.nome_completo if t.colaborador.gerente else ''
        elif t.colaborador.posto_lideranca == 'LIDER':
            t.gestor_nome = t.colaborador.supervisor.nome_completo if t.colaborador.supervisor else ''
        else:
            t.gestor_nome = t.colaborador.lider.nome_completo if t.colaborador.lider else ''

    # Aplicar ordenação multi-nível estável (do último subnível para o 1º nível)
    if ordenacao_niveis:
        def get_sort_value(item, campo):
            if campo == 'colaborador':
                return (item.colaborador.nome_completo or '').strip().upper()
            elif campo == 'matricula':
                return (item.colaborador.matricula or '').strip().upper()
            elif campo in ['responsavel', 'gestor']:
                return (getattr(item, 'gestor_nome', '') or '').strip().upper()
            elif campo == 'setor':
                return (item.colaborador.setor.nome if item.colaborador.setor else '').strip().upper()
            elif campo in ['procedimento', 'codigo']:
                return (item.procedimento.codigo or '').strip().upper()
            elif campo in ['procedimento_nome', 'titulo', 'nome']:
                return (getattr(item.procedimento, 'titulo', None) or item.procedimento.nome or '').strip().upper()
            elif campo == 'data_treinamento':
                return item.data_treinamento or date.min
            elif campo in ['elegibilidade', 'data_elegibilidade']:
                return getattr(item, 'data_elegibilidade', date.min) or date.min
            elif campo in ['decorridos', 'dias_decorridos']:
                return getattr(item, 'dias_decorridos', 0) or 0
            elif campo == 'status':
                return (getattr(item, 'status_display', '') or '').strip().upper()
            return ''

        for nivel in reversed(ordenacao_niveis):
            is_desc = (nivel['direcao'] == 'desc')
            registros_list.sort(key=lambda x: get_sort_value(x, nivel['campo']), reverse=is_desc)
    else:
        # Padrão: mais recentes primeiro
        registros_list.sort(key=lambda x: (x.data_treinamento or date.min), reverse=True)

    # Paginação
    if exibir == 'todos':
        per_page = max(1, len(registros_list))  # Evita zero para Paginator
    else:
        try:
            per_page = int(exibir)
        except ValueError:
            per_page = 30

    paginator = Paginator(registros_list, per_page)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Manter a query string dos filtros para paginação
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'setores': setores_list,
        'setores_list': setores_list,
        'lideres': gestores_list,
        'gestores_list': gestores_list,
        'procedimentos_list': procedimentos_list,
        'matrizes_opcoes': matrizes_opcoes,
        'busca': busca,
        'query': busca,
        'status_filtro': status_filtro,
        'lider_id': gestor_id,
        'gestor_filtro': gestor_id,
        'setor_id': setor_id,
        'setor_filtro': setor_id,
        'procedimento_id': procedimento_id,
        'procedimento_filtro': procedimento_id,
        'dias_decorridos_filtro': dias_decorridos_filtro,
        'dias_filtro': dias_decorridos_filtro,
        'vinculo_filtro': vinculo_filtro,
        'posterior_filtro': posterior_filtro,
        'matriz_filtro': matriz_filtro,
        'exibir': exibir,
        'ordenacao': ordenacao,
        'ordenacao_niveis': ordenacao_niveis,
        'query_string': query_string,
        # Totais
        'total_geral': total_geral,
        'total_eficaz': total_eficaz,
        'total_ineficaz': total_ineficaz,
        'total_nao_aplica': total_nao_aplica,
        'total_carencia': total_carencia,
        'total_pendente': total_pendente,
        'today': today,
    }
    return render(request, "procedures/avaliacao_eficacia_lista.html", context)

@require_POST
@login_required
def avaliacao_eficacia_registrar_view(request, treinamento_id):
    """
    Grava ou atualiza os dados da autoavaliação de eficácia e permite ajustar o gestor responsável.
    """
    is_ajax = (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or request.POST.get('is_ajax') == '1'
        or 'application/json' in request.headers.get('Accept', '')
    )

    treinamento = get_object_or_404(RegistroTreinamento, id=treinamento_id)
    
    # Validar se o procedimento associado é crítico
    if not treinamento.procedimento or treinamento.procedimento.criticidade != 'CRITICO':
        msg = "Este treinamento não requer Avaliação de Eficácia."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    # Obter dados do formulário
    status = request.POST.get('status')
    data_avaliacao_str = request.POST.get('data_avaliacao')
    resultado_avaliacao = request.POST.get('resultado_avaliacao', '').strip()
    gestor_resp_id = request.POST.get('gestor_responsavel')

    # Validações básicas
    if status not in ['EFICAZ', 'INEFICAZ', 'NAO_APLICA']:
        msg = "Status inválido selecionado."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    if status == 'NAO_APLICA' and not resultado_avaliacao:
        msg = "A justificativa é obrigatória para a opção 'Não se Aplica'."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    if not data_avaliacao_str:
        msg = "A data da avaliação é obrigatória."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    from datetime import datetime
    try:
        data_avaliacao = datetime.strptime(data_avaliacao_str, '%Y-%m-%d').date()
    except ValueError:
        msg = "Formato de data inválido."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    # Salvar avaliação no registro do treinamento
    treinamento.avaliacao_eficacia_status = status
    treinamento.avaliacao_eficacia_data = data_avaliacao
    treinamento.resultado_avaliacao = resultado_avaliacao
    
    if gestor_resp_id:
        if gestor_resp_id == 'padrao' or gestor_resp_id == '':
            treinamento.gestor_responsavel = None
        else:
            try:
                treinamento.gestor_responsavel_id = int(gestor_resp_id)
            except (ValueError, TypeError):
                pass

    treinamento.save()

    if is_ajax:
        # Resolver nome do gestor formatado
        gestor_nome = ''
        is_direcionado = False
        if treinamento.gestor_responsavel:
            gestor_nome = treinamento.gestor_responsavel.nome_completo
            is_direcionado = True
        elif treinamento.colaborador.posto_lideranca == 'SUPERVISOR':
            gestor_nome = treinamento.colaborador.gerente.nome_completo if treinamento.colaborador.gerente else '-'
        elif treinamento.colaborador.posto_lideranca == 'LIDER':
            gestor_nome = treinamento.colaborador.supervisor.nome_completo if treinamento.colaborador.supervisor else '-'
        else:
            gestor_nome = treinamento.colaborador.lider.nome_completo if treinamento.colaborador.lider else '-'

        status_display_map = {
            'EFICAZ': ('Eficaz', 'success'),
            'INEFICAZ': ('Ineficaz', 'danger'),
            'NAO_APLICA': ('Não se Aplica', 'secondary'),
            'PENDENTE': ('Pendente', 'warning'),
            'CARENCIA': ('Carência', 'info'),
        }
        status_disp, status_cls = status_display_map.get(treinamento.avaliacao_eficacia_status, ('Pendente', 'warning'))

        return JsonResponse({
            'success': True,
            'message': f"Avaliação de eficácia registrada com sucesso para {treinamento.colaborador.nome_completo}.",
            'id': treinamento.id,
            'status': treinamento.avaliacao_eficacia_status,
            'status_display': status_disp,
            'status_class': status_cls,
            'data_avaliacao': treinamento.avaliacao_eficacia_data.strftime('%Y-%m-%d') if treinamento.avaliacao_eficacia_data else '',
            'data_avaliacao_formatada': treinamento.avaliacao_eficacia_data.strftime('%d/%m/%Y') if treinamento.avaliacao_eficacia_data else '',
            'resultado_avaliacao': treinamento.resultado_avaliacao or '',
            'gestor_nome': gestor_nome,
            'gestor_id': str(treinamento.gestor_responsavel_id or ''),
            'is_direcionado': is_direcionado,
            'colaborador_nome': treinamento.colaborador.nome_completo,
        })

    messages.success(request, f"Avaliação de eficácia registrada com sucesso para {treinamento.colaborador.nome_completo}.")
    return redirect("procedures:avaliacao_eficacia_list")

@require_POST
@login_required
def avaliacao_eficacia_registrar_massa_view(request):
    """
    Registra avaliação de eficácia em massa para vários treinamentos selecionados.
    """
    treinamento_ids = request.POST.getlist('treinamento_ids')
    if not treinamento_ids:
        messages.warning(request, "Nenhum treinamento selecionado para avaliação em massa.")
        return redirect("procedures:avaliacao_eficacia_list")

    status = request.POST.get('status')
    data_avaliacao_str = request.POST.get('data_avaliacao')
    resultado_avaliacao = request.POST.get('resultado_avaliacao', '').strip()
    gestor_resp_id = request.POST.get('gestor_responsavel')

    # Validações básicas
    if status not in ['EFICAZ', 'INEFICAZ', 'NAO_APLICA']:
        messages.error(request, "Status de eficácia inválido.")
        return redirect("procedures:avaliacao_eficacia_list")

    if status == 'NAO_APLICA' and not resultado_avaliacao:
        messages.error(request, "A justificativa é obrigatória para a opção 'Não se Aplica'.")
        return redirect("procedures:avaliacao_eficacia_list")

    if not data_avaliacao_str:
        messages.error(request, "A data da avaliação é obrigatória.")
        return redirect("procedures:avaliacao_eficacia_list")

    from datetime import datetime
    try:
        data_avaliacao = datetime.strptime(data_avaliacao_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Formato de data inválido.")
        return redirect("procedures:avaliacao_eficacia_list")

    # Filtrar os treinamentos que de fato necessitam e não estão em carência
    today = date.today()
    carencia_date = today - timedelta(days=30)
    
    treinamentos = RegistroTreinamento.objects.filter(
        id__in=treinamento_ids,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
    ).filter(
        # Apenas os elegíveis (data_treinamento + 30 <= hoje)
        data_treinamento__lte=carencia_date
    )

    count = 0
    if treinamentos.exists():
        from django.db import transaction
        with transaction.atomic():
            for t in treinamentos:
                t.avaliacao_eficacia_status = status
                t.avaliacao_eficacia_data = data_avaliacao
                t.resultado_avaliacao = resultado_avaliacao
                if gestor_resp_id and gestor_resp_id != 'manter':
                    if gestor_resp_id == 'padrao':
                        t.gestor_responsavel = None
                    else:
                        try:
                            t.gestor_responsavel_id = int(gestor_resp_id)
                        except (ValueError, TypeError):
                            pass
                t.save()
                count += 1

    messages.success(request, f"Avaliação em massa concluída! {count} treinamento(s) atualizado(s) com sucesso.")
    return redirect("procedures:avaliacao_eficacia_list")


@require_POST
@login_required
def avaliacao_eficacia_alterar_gestor_massa_view(request):
    """
    Permite alterar/redirecionar forçadamente o gestor responsável para um ou múltiplos treinamentos selecionados.
    """
    is_ajax = (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or request.POST.get('is_ajax') == '1'
        or 'application/json' in request.headers.get('Accept', '')
    )

    treinamento_ids = request.POST.getlist('treinamento_ids')
    if not treinamento_ids:
        msg = "Nenhum treinamento selecionado para direcionamento de gestor."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.warning(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    novo_gestor_id = request.POST.get('novo_gestor_id')
    if not novo_gestor_id:
        msg = "Nenhum gestor selecionado."
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=400)
        messages.warning(request, msg)
        return redirect("procedures:avaliacao_eficacia_list")

    treinamentos = RegistroTreinamento.objects.filter(
        id__in=treinamento_ids,
        procedimento__criticidade='CRITICO',
        ativo=True,
    )

    count = 0
    if treinamentos.exists():
        from django.db import transaction
        with transaction.atomic():
            for t in treinamentos:
                if novo_gestor_id == 'padrao':
                    t.gestor_responsavel = None
                else:
                    try:
                        t.gestor_responsavel_id = int(novo_gestor_id)
                    except (ValueError, TypeError):
                        continue
                t.save()
                count += 1

    if is_ajax:
        if count == 1:
            t = treinamentos.first()
            gestor_nome = ''
            is_direcionado = False
            if t and t.gestor_responsavel:
                gestor_nome = t.gestor_responsavel.nome_completo
                is_direcionado = True
            elif t and t.colaborador.posto_lideranca == 'SUPERVISOR':
                gestor_nome = t.colaborador.gerente.nome_completo if t.colaborador.gerente else '-'
            elif t and t.colaborador.posto_lideranca == 'LIDER':
                gestor_nome = t.colaborador.supervisor.nome_completo if t.colaborador.supervisor else '-'
            elif t:
                gestor_nome = t.colaborador.lider.nome_completo if t.colaborador.lider else '-'

            return JsonResponse({
                'success': True,
                'message': "Gestor responsável atualizado com sucesso!",
                'id': t.id if t else None,
                'gestor_nome': gestor_nome,
                'gestor_id': str(t.gestor_responsavel_id or '') if t else '',
                'is_direcionado': is_direcionado,
            })
        return JsonResponse({
            'success': True,
            'message': f"Gestor responsável atualizado com sucesso para {count} treinamento(s).",
            'count': count,
        })

    return_url = request.POST.get('return_url') or request.META.get('HTTP_REFERER') or "procedures:avaliacao_eficacia_list"
    if count == 1:
        messages.success(request, "Gestor responsável atualizado com sucesso!")
    else:
        messages.success(request, f"Gestor responsável atualizado com sucesso para {count} treinamento(s).")
    return redirect(return_url)


@login_required
def avaliacao_eficacia_export_excel_view(request):
    """
    Exporta em Excel as avaliações de eficácia dos colaboradores ativos de acordo com os filtros aplicados.
    """
    # Mesma query base
    qs = RegistroTreinamento.objects.select_related(
        'colaborador', 'procedimento', 'colaborador__setor', 'colaborador__lider', 'colaborador__supervisor', 'colaborador__gerente', 'gestor_responsavel'
    ).filter(
        colaborador__isnull=False,
        procedimento__isnull=False,
        procedimento__criticidade='CRITICO',
        data_treinamento__isnull=False,
        ativo=True,
        colaborador__is_active=True,
        colaborador__afastado=False,
        colaborador__em_ferias=False,
    )

    # Parâmetros de filtro
    busca = (request.GET.get('q') or request.GET.get('busca') or '').strip()
    status_filtro = (request.GET.get('status') or '').strip()
    gestor_id = (request.GET.get('gestor') or request.GET.get('lider') or '').strip()
    setor_id = (request.GET.get('setor') or '').strip()
    procedimento_id = (request.GET.get('procedimento') or request.GET.get('procedimento_id') or '').strip()
    dias_decorridos_filtro = (request.GET.get('dias') or request.GET.get('dias_decorridos') or '').strip()
    vinculo_filtro = (request.GET.get('vinculo') or '').strip()
    posterior_filtro = (request.GET.get('posterior') or '').strip()
    matriz_filtro = (request.GET.get('matriz') or '').strip()

    today = date.today()
    carencia_date = today - timedelta(days=30)

    # Aplicar filtros
    if busca:
        qs = qs.filter(
            Q(colaborador__nome_completo__icontains=busca) |
            Q(colaborador__matricula__icontains=busca) |
            Q(procedimento__codigo__icontains=busca) |
            Q(procedimento__nome__icontains=busca)
        )
    if gestor_id:
        qs = qs.filter(
            Q(gestor_responsavel_id=gestor_id) |
            (Q(gestor_responsavel__isnull=True) & (
                Q(colaborador__lider_id=gestor_id) |
                Q(colaborador__supervisor_id=gestor_id) |
                Q(colaborador__gerente_id=gestor_id)
            ))
        )
    if setor_id:
        qs = qs.filter(colaborador__setor_id=setor_id)
    if procedimento_id:
        qs = qs.filter(procedimento_id=procedimento_id)
    if matriz_filtro:
        qs = qs.filter(procedimento__matriz__iexact=matriz_filtro)

    if dias_decorridos_filtro:
        if dias_decorridos_filtro == 'menos_30':
            qs = qs.filter(data_treinamento__gt=carencia_date)
        elif dias_decorridos_filtro == '30_60':
            qs = qs.filter(data_treinamento__lte=carencia_date, data_treinamento__gte=today - timedelta(days=60))
        elif dias_decorridos_filtro == '60_90':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=60), data_treinamento__gte=today - timedelta(days=90))
        elif dias_decorridos_filtro == '90_mais':
            qs = qs.filter(data_treinamento__lt=today - timedelta(days=90))

    if status_filtro:
        if status_filtro == 'EFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='EFICAZ')
        elif status_filtro == 'INEFICAZ':
            qs = qs.filter(avaliacao_eficacia_status='INEFICAZ')
        elif status_filtro == 'NAO_APLICA':
            qs = qs.filter(avaliacao_eficacia_status='NAO_APLICA')
        elif status_filtro == 'CARENCIA':
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__gt=carencia_date
            )
        elif status_filtro == 'PENDENTE':
            qs = qs.filter(
                Q(avaliacao_eficacia_status='PENDENTE') | Q(avaliacao_eficacia_status__isnull=True),
                data_treinamento__lte=carencia_date
            )

    qs = qs.order_by('-data_treinamento', 'colaborador__nome_completo')
    registros_list = list(qs)

    # Vínculo ao perfil
    if vinculo_filtro:
        colab_ids = {t.colaborador_id for t in registros_list}
        perfis_colabs = ColaboradorPerfil.objects.filter(
            colaborador_id__in=colab_ids,
            ativo=True
        ).select_related('perfil').prefetch_related('perfil__grupos__subgrupos__procedimentos')
        
        colab_to_procs = defaultdict(set)
        for cp in perfis_colabs:
            for g in cp.perfil.grupos.all():
                for sg in g.subgrupos.all():
                    for proc in sg.procedimentos.all():
                        colab_to_procs[cp.colaborador_id].add(proc.id)
                        
        for t in registros_list:
            t.tem_vinculo_perfil = t.procedimento_id in colab_to_procs[t.colaborador_id]
            
        if vinculo_filtro == 'COM_VINCULO':
            registros_list = [t for t in registros_list if t.tem_vinculo_perfil]
        elif vinculo_filtro == 'SEM_VINCULO':
            registros_list = [t for t in registros_list if not t.tem_vinculo_perfil]

    # Treinamento Posterior
    if posterior_filtro or True:  # Sempre calcula para o excel
        all_critical = list(RegistroTreinamento.objects.filter(
            colaborador__isnull=False,
            procedimento__isnull=False,
            procedimento__criticidade='CRITICO',
            data_treinamento__isnull=False,
            ativo=True,
            colaborador__is_active=True,
        ))
        grouped = defaultdict(list)
        for t in all_critical:
            grouped[(t.colaborador_id, t.procedimento_id)].append(t)
        for key in grouped:
            grouped[key].sort(key=lambda x: x.data_treinamento)
        posterior_pending_ids = set()
        for key, group in grouped.items():
            n = len(group)
            for i in range(n):
                t = group[i]
                has_posterior = False
                for j in range(i + 1, n):
                    post = group[j]
                    if post.avaliacao_eficacia_status in ['PENDENTE', None, '']:
                        has_posterior = True
                        break
                if has_posterior:
                    posterior_pending_ids.add(t.id)

        for t in registros_list:
            t.has_posterior_pending = t.id in posterior_pending_ids

        if posterior_filtro == 'COM_POSTERIOR':
            registros_list = [t for t in registros_list if t.has_posterior_pending]
        elif posterior_filtro == 'SEM_POSTERIOR':
            registros_list = [t for t in registros_list if not t.has_posterior_pending]

    # Agora construímos o arquivo Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avaliacoes"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    data_font = Font(name="Arial", size=10)
    
    border_side = Side(style='thin', color='D9D9D9')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    headers = [
        "Colaborador",
        "Matrícula",
        "Cargo",
        "Setor",
        "Responsável",
        "Matriz",
        "Código Procedimento",
        "Nome Procedimento",
        "Data Treinamento",
        "Status Eficácia",
        "Data Avaliação",
        "Justificativa / Evidências / Observações"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
    ws.row_dimensions[1].height = 25

    status_mapping = {
        'EFICAZ': 'Eficaz',
        'INEFICAZ': 'Ineficaz',
        'NAO_APLICA': 'Não se Aplica',
        'PENDENTE': 'Pendente',
    }

    for row_idx, t in enumerate(registros_list, 2):
        # Determinar status para exportação
        status_val = t.avaliacao_eficacia_status
        if status_val in [None, '', 'PENDENTE']:
            days_diff = (today - t.data_treinamento).days
            status_text = 'Carência' if days_diff < 30 else 'Pendente'
        else:
            status_text = status_mapping.get(status_val, status_val)

        colab = t.colaborador
        responsavel_nome = '-'
        if t.gestor_responsavel:
            responsavel_nome = t.gestor_responsavel.nome_completo
        elif colab:
            if colab.posto_lideranca == 'SUPERVISOR':
                responsavel_nome = colab.gerente.nome_completo if colab.gerente else '-'
            elif colab.posto_lideranca == 'LIDER':
                responsavel_nome = colab.supervisor.nome_completo if colab.supervisor else '-'
            else:
                responsavel_nome = colab.lider.nome_completo if colab.lider else '-'

        row_data = [
            t.colaborador.nome_completo,
            t.colaborador.matricula,
            t.colaborador.cargo or '-',
            t.colaborador.setor.nome if t.colaborador.setor else '-',
            responsavel_nome,
            t.procedimento.matriz or '-',
            t.procedimento.codigo,
            t.procedimento.nome,
            t.data_treinamento.strftime('%d/%m/%Y') if t.data_treinamento else '-',
            status_text,
            t.avaliacao_eficacia_data.strftime('%d/%m/%Y') if t.avaliacao_eficacia_data else '-',
            t.resultado_avaliacao or ''
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            
            # Alinhamentos específicos
            if col_idx in [2, 9, 10, 11]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        ws.row_dimensions[row_idx].height = 20

    # Auto-ajustar colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # Preparar Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="avaliacao_eficacia_{today.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_avaliacao_eficacia_for142_view(request, treinamento_id):
    """
    Exporta o Formulário Individual de Avaliação de Eficácia do Treinamento (FOR.142.r01) em Excel (.xlsx).
    Calcula a data devida (+30 dias de carência) na tag {{DATA_EFICACIA_CALCULADA}}.
    """
    treinamento = get_object_or_404(
        RegistroTreinamento.objects.select_related('colaborador', 'procedimento'),
        id=treinamento_id
    )

    from procedures.services.treinamento_excel_export_service import gerar_avaliacao_eficacia_for142_xlsx

    try:
        excel_buffer = gerar_avaliacao_eficacia_for142_xlsx(treinamento.id)
    except Exception as e:
        messages.error(request, f"Erro ao gerar formulário de eficácia Excel (FOR.142): {str(e)}")
        return redirect('procedures:avaliacao_eficacia_list')

    colab_slug = (treinamento.colaborador.nome_completo.split()[0] if treinamento.colaborador else 'colaborador')
    filename = f"FOR.142.r01_Avaliacao_Eficacia_{colab_slug}_ID{treinamento.id}.xlsx"
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Access-Control-Expose-Headers"] = "Content-Disposition"
    return response


@login_required
def exportar_avaliacao_eficacia_for142_massa_view(request):
    """
    Exporta formulários de Avaliação de Eficácia (FOR.142.r01) em lote com Múltiplas Abas.
    Cada linha selecionada pelo usuário gera uma aba independente na mesma pasta de trabalho Excel.
    Aceita IDs via GET (ex: ?ids=1,2,3 ou ?treinamento_ids=1&treinamento_ids=2) ou via POST.
    """
    ids_raw = request.POST.getlist('treinamento_ids') or request.GET.getlist('treinamento_ids')
    if not ids_raw:
        ids_str = request.GET.get('ids', '')
        if ids_str:
            ids_raw = [x.strip() for x in ids_str.split(',') if x.strip().isdigit()]

    treinamento_ids = []
    for item in ids_raw:
        try:
            treinamento_ids.append(int(item))
        except (ValueError, TypeError):
            continue

    if not treinamento_ids:
        messages.warning(request, "Nenhum treinamento selecionado para exportação do FOR.142.")
        return redirect('procedures:avaliacao_eficacia_list')

    from procedures.services.treinamento_excel_export_service import gerar_avaliacao_eficacia_multiplas_abas_for142_xlsx

    try:
        excel_buffer = gerar_avaliacao_eficacia_multiplas_abas_for142_xlsx(treinamento_ids)
    except Exception as e:
        messages.error(request, f"Erro ao gerar formulários de eficácia com múltiplas abas: {str(e)}")
        return redirect('procedures:avaliacao_eficacia_list')

    total = len(treinamento_ids)
    data_hoje = timezone.now().strftime("%Y%m%d")
    filename = f"FOR.142.r01_Avaliacao_Eficacia_{total}_Abas_{data_hoje}.xlsx"

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Access-Control-Expose-Headers"] = "Content-Disposition"
    return response


@login_required
def avaliacao_eficacia_search_options_api(request):
    """
    Endpoint AJAX para Select2 dos filtros de Setor, Gestor e Procedimento.
    """
    tipo = request.GET.get('tipo', '').strip().lower()
    q = request.GET.get('q', '').strip()

    results = []
    if tipo == 'setor':
        qs = Setor.objects.all()
        if q:
            qs = qs.filter(nome__icontains=q)
        qs = qs.order_by('nome')[:50]
        results = [{'id': str(s.id), 'text': s.nome} for s in qs]

    elif tipo in ['gestor', 'lider', 'responsavel']:
        qs = Colaborador.objects.filter(is_active=True).filter(
            Q(posto_lideranca__in=['LIDER', 'SUPERVISOR', 'GERENTE', 'QUALIDADE', 'PROCESSOS', 'MANUTENCAO', 'EHS']) |
            Q(id__in=Colaborador.objects.values_list('lider_id', flat=True)) |
            Q(id__in=Colaborador.objects.values_list('supervisor_id', flat=True)) |
            Q(id__in=Colaborador.objects.values_list('gerente_id', flat=True))
        ).distinct()
        if q:
            qs = qs.filter(
                Q(nome_completo__icontains=q) |
                Q(matricula__icontains=q)
            )
        qs = qs.order_by('nome_completo')[:50]
        results = [{'id': str(g.id), 'text': f"{g.nome_completo} ({g.matricula})" if g.matricula else g.nome_completo} for g in qs]

    elif tipo == 'procedimento':
        qs = Procedimento.objects.filter(criticidade='CRITICO')
        if q:
            qs = qs.filter(
                Q(codigo__icontains=q) |
                Q(nome__icontains=q) |
                Q(titulo__icontains=q)
            )
        qs = qs.order_by('codigo', 'nome')[:50]
        results = [{'id': str(p.id), 'text': f"{p.codigo} - {getattr(p, 'titulo', None) or p.nome}"} for p in qs]

    return JsonResponse({'results': results})


