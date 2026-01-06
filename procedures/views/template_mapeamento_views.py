# -*- coding: utf-8 -*-
"""
Views para upload e mapeamento configurável de templates de listas de presença

Funcionalidades:
- Upload de arquivo Excel template
- Preview das abas e células do Excel
- Mapeamento visual (clique + referência)
- Salvar configuração de mapeamento
"""

import json
import os
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

from procedures.models import TemplateListaPresenca, MapeamentoCampoListaPresenca

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def upload_excel_template_view(request, pk=None):
    """
    Upload de arquivo Excel para usar como template de mapeamento
    
    GET: Formulário de upload
    POST: Processa arquivo Excel
    """
    if pk:
        template = get_object_or_404(TemplateListaPresenca, pk=pk)
    else:
        template = None
    
    if request.method == 'POST':
        arquivo_excel = request.FILES.get('arquivo_excel')
        
        if not arquivo_excel:
            messages.error(request, 'Nenhum arquivo foi enviado.')
            return render(request, 'procedures/upload_excel_template.html', {
                'template': template
            })
        
        # Validar extensão
        if not arquivo_excel.name.endswith('.xlsx'):
            messages.error(request, 'Apenas arquivos .xlsx são aceitos.')
            return render(request, 'procedures/upload_excel_template.html', {
                'template': template
            })
        
        # Validar tamanho (max 5 MB)
        if arquivo_excel.size > 5 * 1024 * 1024:
            messages.error(request, 'Arquivo muito grande (máximo 5 MB).')
            return render(request, 'procedures/upload_excel_template.html', {
                'template': template
            })
        
        try:
            # Se é atualização, remover arquivo anterior
            if template and template.arquivo_excel_template:
                if template.arquivo_excel_template.storage.exists(
                    template.arquivo_excel_template.name
                ):
                    template.arquivo_excel_template.storage.delete(
                        template.arquivo_excel_template.name
                    )
            
            # Salvar novo arquivo
            if template:
                template.arquivo_excel_template = arquivo_excel
                template.save()
            
            messages.success(
                request, 
                'Arquivo Excel enviado com sucesso! Clique em "Mapear Campos" para configurar.'
            )
            
            if template:
                return redirect('procedures:mapear_campos_template', pk=template.pk)
            else:
                # Se foi novo, redirecionar para criação
                messages.warning(
                    request,
                    'Por favor, crie um template primeiro.'
                )
                return redirect('admin:procedures_templatelistapresenca_add')
        
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
    
    return render(request, 'procedures/upload_excel_template.html', {
        'template': template
    })


@login_required
def mapear_campos_template_view(request, pk):
    """
    Interface para mapear campos do Excel para dados da lista de presença
    
    GET: Mostra interface de mapeamento com preview do Excel
    POST: Salva mapeamento configurado
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    if not template.arquivo_excel_template:
        messages.error(request, 'Template não tem arquivo Excel anexado.')
        return redirect('admin:procedures_templatelistapresenca_change', object_id=pk)
    
    # Campos que devem ser mapeados
    CAMPOS_OBRIGATORIOS = [
        'titulo_treinamento',
        'facilitador_fornecedor',
        'data_hora',
        'carga_horaria',
        'procedimentos_assuntos',
    ]
    
    # Mapeamento de campos para rótulos legíveis
    CAMPOS_LABELS = {
        'titulo_treinamento': 'Título do Treinamento',
        'facilitador_fornecedor': 'Facilitador/Fornecedor',
        'data_hora': 'Data/Hora',
        'carga_horaria': 'Carga Horária',
        'procedimentos_assuntos': 'Procedimentos/Assuntos',
    }
    
    # Criar lista de campos com labels
    campos_com_labels = [(campo, CAMPOS_LABELS.get(campo, campo)) for campo in CAMPOS_OBRIGATORIOS]
    
    # Obter mapeamentos existentes
    mapeamentos_existentes = {}
    mapeamentos_flattenados = {}
    for mapeamento in template.mapeamentos.all():
        mapeamentos_existentes[mapeamento.tipo_campo] = {
            'localizacao': mapeamento.localizacao,
            'metodo': mapeamento.metodo,
            'pagina': mapeamento.pagina,
            'obrigatorio': mapeamento.obrigatorio,
            'permite_imagem_marcacao': mapeamento.permite_imagem_marcacao,
        }
        # Criar chaves flattenadas para acesso fácil no template
        mapeamentos_flattenados[f'{mapeamento.tipo_campo}_localizacao'] = mapeamento.localizacao
        mapeamentos_flattenados[f'{mapeamento.tipo_campo}_metodo'] = mapeamento.metodo
        mapeamentos_flattenados[f'{mapeamento.tipo_campo}_pagina'] = mapeamento.pagina
        mapeamentos_flattenados[f'{mapeamento.tipo_campo}_obrigatorio'] = mapeamento.obrigatorio
        mapeamentos_flattenados[f'{mapeamento.tipo_campo}_permite_imagem_marcacao'] = mapeamento.permite_imagem_marcacao
    
    # Obter info do Excel se disponível
    excel_info = None
    if OPENPYXL_AVAILABLE and template.arquivo_excel_template:
        try:
            arquivo_path = template.arquivo_excel_template.path
            workbook = openpyxl.load_workbook(arquivo_path)
            
            excel_info = {
                'abas': workbook.sheetnames,
                'aba_ativa': workbook.active.title if workbook.active else 'Plan1',
                'celulas_preview': {}  # Será preenchida via AJAX
            }
        except Exception as e:
            messages.warning(request, f'Erro ao ler Excel: {str(e)}')
    
    if request.method == 'POST':
        # Receber mapeamento do formulário
        mapeamento_dados = {}
        
        # Processar dados do POST
        for tipo_campo in CAMPOS_OBRIGATORIOS:
            localizacao = request.POST.get(f'{tipo_campo}_localizacao', '').strip()
            metodo = request.POST.get(f'{tipo_campo}_metodo', 'referencia')
            pagina = int(request.POST.get(f'{tipo_campo}_pagina', 1))
            obrigatorio = request.POST.get(f'{tipo_campo}_obrigatorio') == 'on'
            permite_imagem = request.POST.get(f'{tipo_campo}_permite_imagem') == 'on'
            
            if localizacao:
                mapeamento_dados[tipo_campo] = {
                    'localizacao': localizacao,
                    'metodo': metodo,
                    'pagina': pagina,
                    'obrigatorio': obrigatorio,
                    'permite_imagem_marcacao': permite_imagem,
                }
        
        try:
            # Salvar mapeamentos no BD
            for tipo_campo, dados in mapeamento_dados.items():
                mapeamento, created = MapeamentoCampoListaPresenca.objects.update_or_create(
                    template=template,
                    tipo_campo=tipo_campo,
                    defaults={
                        'localizacao': dados['localizacao'],
                        'metodo': dados['metodo'],
                        'pagina': dados['pagina'],
                        'obrigatorio': dados['obrigatorio'],
                        'permite_imagem_marcacao': dados['permite_imagem_marcacao'],
                    }
                )
            
            # Salvar mapeamento em JSON também para rápido acesso
            template.mapeamento_campos = mapeamento_dados
            template.mapeamento_completo = len(mapeamento_dados) == len(CAMPOS_OBRIGATORIOS)
            template.save()
            
            messages.success(request, 'Mapeamento salvo com sucesso!')
            return redirect('admin:procedures_templatelistapresenca_change', object_id=pk)
        
        except Exception as e:
            messages.error(request, f'Erro ao salvar mapeamento: {str(e)}')
    
    context = {
        'template': template,
        'campos_obrigatorios': CAMPOS_OBRIGATORIOS,
        'campos_com_labels': campos_com_labels,
        'mapeamentos_existentes': mapeamentos_existentes,
        'mapeamentos_flattenados': mapeamentos_flattenados,
        'excel_info': excel_info,
        'metodo_choices': MapeamentoCampoListaPresenca._meta.get_field('metodo').choices,
        'openpyxl_available': OPENPYXL_AVAILABLE,
    }
    
    return render(request, 'procedures/mapear_campos_template.html', context)


@login_required
@require_http_methods(["GET"])
def preview_excel_abas_api(request, pk):
    """
    API REST para obter lista de abas do Excel
    Retorna JSON com nomes das abas
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'erro': 'Biblioteca openpyxl não está instalada'
        }, status=400)
    
    if not template.arquivo_excel_template:
        return JsonResponse({
            'erro': 'Template não tem arquivo Excel'
        }, status=400)
    
    try:
        arquivo_path = template.arquivo_excel_template.path
        workbook = openpyxl.load_workbook(arquivo_path, data_only=False)
        
        abas = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            abas.append({
                'nome': sheet_name,
                'linhas': worksheet.max_row,
                'colunas': worksheet.max_column,
            })
        
        return JsonResponse({
            'abas': abas,
            'aba_ativa': workbook.active.title if workbook.active else 'Plan1'
        })
    
    except Exception as e:
        return JsonResponse({
            'erro': f'Erro ao ler Excel: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def preview_excel_celulas_api(request, pk):
    """
    API REST para obter dados das células de uma aba específica
    
    Query params:
    - aba: nome da aba (default: primeira)
    - range: intervalo de células (ex: "A1:F50" ou "A1")
    
    Retorna JSON com valores e formatação das células
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'erro': 'openpyxl não instalado'}, status=400)
    
    if not template.arquivo_excel_template:
        return JsonResponse({'erro': 'Sem arquivo Excel'}, status=400)
    
    try:
        arquivo_path = template.arquivo_excel_template.path
        workbook = openpyxl.load_workbook(arquivo_path, data_only=False)
        
        aba_nome = request.GET.get('aba', workbook.sheetnames[0] if workbook.sheetnames else 'Plan1')
        worksheet = workbook[aba_nome]
        
        celulas = []
        
        # Obter range de células solicitado
        range_str = request.GET.get('range', 'A1:Z50')  # Default: primeiras 50 linhas
        
        # Processa range (suporta colunas com 1 ou 2 letras, ex: A1, Z10, AA3)
        if ':' in range_str:
            inicio, fim = range_str.split(':', 1)
        else:
            inicio = fim = range_str

        def _split_ref(ref: str) -> tuple[str, int]:
            ref = (ref or '').strip().upper()
            letters = ''.join([c for c in ref if c.isalpha()])
            numbers = ''.join([c for c in ref if c.isdigit()])
            if not letters:
                letters = 'A'
            row = int(numbers) if numbers else 1
            return letters, row

        def _col_letters_to_index(letters: str) -> int:
            # A -> 0, B -> 1, Z -> 25, AA -> 26...
            idx = 0
            for ch in letters:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx - 1

        inicio_letters, inicio_linha = _split_ref(inicio)
        fim_letters, fim_linha = _split_ref(fim)

        inicio_col = _col_letters_to_index(inicio_letters)
        fim_col = _col_letters_to_index(fim_letters)
        
        # Normalizar limites (caso range venha invertido)
        if fim_col < inicio_col:
            inicio_col, fim_col = fim_col, inicio_col
        if fim_linha < inicio_linha:
            inicio_linha, fim_linha = fim_linha, inicio_linha

        # Limitar para não sobrecarregar
        fim_col = min(fim_col, 50)  # Até AY (suficiente p/ layout)
        fim_linha = min(fim_linha, 200)  # Até linha 200

        def _argb_to_hex(argb: str | None) -> str | None:
            if not argb:
                return None
            # openpyxl geralmente usa ARGB, ex: 'FF112233'
            argb = str(argb)
            if len(argb) == 8:
                return f"#{argb[2:]}"
            if len(argb) == 6:
                return f"#{argb}"
            return None

        # Mapear merges dentro do range: start_cell -> {rowspan, colspan}
        merges = []
        merged_map: dict[str, dict[str, int]] = {}
        merged_skip: set[str] = set()
        for m in getattr(worksheet.merged_cells, 'ranges', []) or []:
            try:
                min_row, min_col, max_row, max_col = m.min_row, m.min_col, m.max_row, m.max_col
            except Exception:
                continue

            # Interseção com o range solicitado
            if max_row < inicio_linha or min_row > fim_linha:
                continue
            if max_col < (inicio_col + 1) or min_col > (fim_col + 1):
                continue

            start_ref = f"{get_column_letter(min_col)}{min_row}"
            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1
            merged_map[start_ref] = {'rowspan': rowspan, 'colspan': colspan}
            merges.append({'start': start_ref, 'rowspan': rowspan, 'colspan': colspan})

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ref = f"{get_column_letter(c)}{r}"
                    if ref != start_ref:
                        merged_skip.add(ref)
        
        for linha in range(inicio_linha, fim_linha + 1):
            for col in range(inicio_col, fim_col + 1):
                cell = worksheet.cell(row=linha, column=col + 1)
                col_letter = get_column_letter(col + 1)
                ref = f'{col_letter}{linha}'

                # Se a célula está dentro de um merge e não é a célula inicial, pode ser pulada no front-end
                is_merged = ref in merged_skip or ref in merged_map

                # Bordas
                b = getattr(cell, 'border', None)
                borders = None
                if b:
                    def _side_style(side):
                        st = getattr(side, 'style', None)
                        if not st:
                            return None
                        return str(st)

                    borders = {
                        't': _side_style(getattr(b, 'top', None)),
                        'r': _side_style(getattr(b, 'right', None)),
                        'b': _side_style(getattr(b, 'bottom', None)),
                        'l': _side_style(getattr(b, 'left', None)),
                    }

                # Fill
                fill_hex = None
                fill = getattr(cell, 'fill', None)
                if fill and getattr(fill, 'patternType', None) in ('solid', 'darkDown', 'darkGray', 'lightDown', 'lightGray'):
                    fg = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
                    fill_hex = _argb_to_hex(fg)

                # Fonte
                font = getattr(cell, 'font', None)
                bold = bool(getattr(font, 'bold', False)) if font else False
                font_color = _argb_to_hex(getattr(getattr(font, 'color', None), 'rgb', None)) if font else None

                # Alinhamento
                align = getattr(cell, 'alignment', None)
                alignment = None
                if align:
                    alignment = {
                        'h': getattr(align, 'horizontal', None),
                        'v': getattr(align, 'vertical', None),
                    }

                merge_span = merged_map.get(ref)

                celulas.append({
                    'ref': ref,
                    'row': linha,
                    'col': col + 1,
                    'valor': str(cell.value) if cell.value is not None else '',
                    'tipo': type(cell.value).__name__,
                    'borders': borders,
                    'fill': fill_hex,
                    'bold': bold,
                    'color': font_color,
                    'alignment': alignment,
                    'merge': merge_span,
                    'is_merged': is_merged,
                    'skip': ref in merged_skip,
                })
        
        return JsonResponse({
            'aba': aba_nome,
            'celulas': celulas,
            'merges': merges,
            'range': {
                'inicio': {'row': inicio_linha, 'col': inicio_col + 1},
                'fim': {'row': fim_linha, 'col': fim_col + 1},
            },
            'total_linhas': worksheet.max_row,
            'total_colunas': worksheet.max_column,
        })
    
    except Exception as e:
        return JsonResponse({
            'erro': f'Erro: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def atualizar_mapeamento_campo_api(request, pk):
    """
    API REST para atualizar mapeamento de um campo específico
    
    POST data (JSON):
    {
        "tipo_campo": "titulo_treinamento",
        "localizacao": "A1",
        "metodo": "referencia",
        "pagina": 1,
        "obrigatorio": true,
        "permite_imagem_marcacao": false
    }
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    try:
        dados = json.loads(request.body)
        tipo_campo = dados.get('tipo_campo')
        
        if not tipo_campo:
            return JsonResponse({'erro': 'tipo_campo obrigatório'}, status=400)
        
        # Atualizar ou criar mapeamento
        mapeamento, created = MapeamentoCampoListaPresenca.objects.update_or_create(
            template=template,
            tipo_campo=tipo_campo,
            defaults={
                'localizacao': dados.get('localizacao', 'A1'),
                'metodo': dados.get('metodo', 'referencia'),
                'pagina': dados.get('pagina', 1),
                'obrigatorio': dados.get('obrigatorio', True),
                'permite_imagem_marcacao': dados.get('permite_imagem_marcacao', False),
            }
        )
        
        # Atualizar JSON do template
        if template.mapeamento_campos is None:
            template.mapeamento_campos = {}
        
        template.mapeamento_campos[tipo_campo] = {
            'localizacao': dados.get('localizacao', 'A1'),
            'metodo': dados.get('metodo', 'referencia'),
            'pagina': dados.get('pagina', 1),
            'obrigatorio': dados.get('obrigatorio', True),
            'permite_imagem_marcacao': dados.get('permite_imagem_marcacao', False),
        }
        
        template.save()
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Campo "{tipo_campo}" mapeado para "{dados.get("localizacao")}"',
            'mapeamento': mapeamento.localizacao,
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'erro': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def remover_mapeamento_campo_api(request, pk):
    """
    API REST para remover mapeamento de um campo
    
    POST data (JSON):
    {
        "tipo_campo": "titulo_treinamento"
    }
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    try:
        dados = json.loads(request.body)
        tipo_campo = dados.get('tipo_campo')
        
        if not tipo_campo:
            return JsonResponse({'erro': 'tipo_campo obrigatório'}, status=400)
        
        # Remover mapeamento
        MapeamentoCampoListaPresenca.objects.filter(
            template=template,
            tipo_campo=tipo_campo
        ).delete()
        
        # Remover do JSON também
        if template.mapeamento_campos and tipo_campo in template.mapeamento_campos:
            del template.mapeamento_campos[tipo_campo]
            template.save()
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Mapeamento de "{tipo_campo}" removido'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'erro': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def status_mapeamento_api(request, pk):
    """
    API REST para obter status do mapeamento
    
    Retorna:
    - Quais campos já foram mapeados
    - Quais ainda precisam ser mapeados
    - Se está completo
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    CAMPOS_OBRIGATORIOS = [
        'titulo_treinamento',
        'categoria_treinamento',
        'metodologia',
        'area_conhecimento',
        'necessita_avaliacao',
        'facilitador_fornecedor',
        'data_hora',
        'carga_horaria',
        'procedimentos_assuntos',
    ]
    
    mapeados = set()
    mapeamentos_info = []
    
    for mapeamento in template.mapeamentos.all():
        mapeados.add(mapeamento.tipo_campo)
        mapeamentos_info.append({
            'tipo_campo': mapeamento.tipo_campo,
            'localizacao': mapeamento.localizacao,
            'metodo': mapeamento.get_metodo_display(),
            'obrigatorio': mapeamento.obrigatorio,
        })
    
    pendentes = [c for c in CAMPOS_OBRIGATORIOS if c not in mapeados]
    
    return JsonResponse({
        'total_campos': len(CAMPOS_OBRIGATORIOS),
        'campos_mapeados': len(mapeados),
        'completo': len(pendentes) == 0,
        'pendentes': pendentes,
        'mapeamentos': mapeamentos_info,
    })


@login_required
def mapear_placeholders_view(request, pk):
    """
    View para mapear placeholders do PDF template para campos de dados da lista de presença.
    
    Placeholders disponíveis:
    - {{titulo}}: Título do Treinamento
    - {{facilitador}}: Facilitador/Fornecedor
    - {{data}}: Data do Treinamento (dd/mm/yyyy)
    - {{hora_inicio}}: Hora de Início
    - {{hora_fim}}: Hora de Fim
    - {{carga_horaria}}: Carga Horária
    - {{local}}: Local do Treinamento
    - {{procedimentos}}: Procedimentos/Disciplinas
    - {{empresa}}: Empresa
    - {{departamento}}: Departamento
    """
    template = get_object_or_404(TemplateListaPresenca, pk=pk)
    
    # Campos disponíveis para mapeamento
    campos_disponiveis = [
        ('titulo', 'Título do Treinamento'),
        ('facilitador', 'Facilitador/Fornecedor'),
        ('data', 'Data (dd/mm/yyyy)'),
        ('hora_inicio', 'Hora de Início'),
        ('hora_fim', 'Hora de Fim'),
        ('carga_horaria', 'Carga Horária'),
        ('local', 'Local do Treinamento'),
        ('procedimentos', 'Procedimentos/Disciplinas'),
        ('empresa', 'Empresa'),
        ('departamento', 'Departamento'),
    ]
    
    # Extrair placeholders do PDF (simular detecção)
    # Em produção, isso deveria extrair do PDF real
    placeholders = [
        'titulo', 'facilitador', 'data', 'hora_inicio', 'hora_fim',
        'carga_horaria', 'local', 'procedimentos', 'empresa', 'departamento'
    ]
    
    if request.method == 'POST':
        # Processar mapeamento de placeholders
        try:
            mapeamentos_data = request.POST
            
            # Limpar mapeamentos antigos
            template.mapeamentos.all().delete()
            
            # Criar novos mapeamentos
            for placeholder in placeholders:
                campo_dados = mapeamentos_data.get(f'campo_dados[{placeholder}]')
                
                if campo_dados:
                    MapeamentoCampoListaPresenca.objects.create(
                        template=template,
                        placeholder=placeholder,
                        campo_dados=campo_dados,
                        tipo_campo='dinâmico',
                    )
            
            messages.success(request, 'Mapeamento de placeholders salvo com sucesso!')
            return redirect('procedures:gerenciar_templates_presenca')
            
        except Exception as e:
            messages.error(request, f'Erro ao salvar mapeamento: {str(e)}')
    
    # GET: Mostrar formulário de mapeamento
    mapeamentos_existentes = template.mapeamentos.all()
    
    return render(request, 'procedures/mapear_template_fields.html', {
        'template': template,
        'placeholders': placeholders,
        'campos_disponiveis': campos_disponiveis,
        'mapeamentos_existentes': mapeamentos_existentes,
    })
