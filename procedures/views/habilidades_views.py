# -*- coding: utf-8 -*-
"""
Views para Matriz de Habilidades
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Max, Count
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json

from procedures.models import Disciplina, MatrizHabilidade, AvaliacaoHabilidade, ColaboradorMatrizHabilidade
from procedures.forms import DisciplinaForm, MatrizHabilidadeForm, AvaliacaoHabilidadeForm
from rh.models import Colaborador


# ==============================================================================
# DISCIPLINAS
# ==============================================================================

@login_required
def disciplinas_list_view(request):
    """Lista de disciplinas com filtros."""
    termo = request.GET.get('q', '').strip()
    ativo = request.GET.get('ativo', '')
    
    qs = Disciplina.objects.all()
    
    if termo:
        qs = qs.filter(Q(codigo__icontains=termo) | Q(nome__icontains=termo))
    
    if ativo:
        qs = qs.filter(ativo=(ativo == 'true'))
    
    paginator = Paginator(qs.order_by('codigo'), 50)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    ctx = {
        'disciplinas': page_obj.object_list,
        'page_obj': page_obj,
        'termo': termo,
        'ativo': ativo,
    }
    return render(request, 'procedures/disciplina_lista.html', ctx)


@login_required
def nova_disciplina_view(request):
    """Cria nova disciplina."""
    matriz_id = request.GET.get('matriz')
    
    if request.method == 'POST':
        form = DisciplinaForm(request.POST)
        if form.is_valid():
            disc = form.save()
            messages.success(request, f"Disciplina {disc.codigo} criada com sucesso!")
            # Se veio de uma matriz, volta para ela
            if matriz_id:
                return redirect('procedures:detalhe_matriz', matriz_id=disc.matriz.id)
            return redirect('procedures:disciplinas_list')
    else:
        # Pré-preencher matriz se veio por parâmetro
        initial = {}
        if matriz_id:
            initial['matriz'] = matriz_id
        form = DisciplinaForm(initial=initial)
    
    return render(request, 'procedures/disciplina_form.html', {
        'form': form,
        'titulo': 'Nova Disciplina'
    })


@login_required
def editar_disciplina_view(request, disciplina_id):
    """Edita disciplina existente."""
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    if request.method == 'POST':
        form = DisciplinaForm(request.POST, instance=disc)
        if form.is_valid():
            form.save()
            messages.success(request, "Disciplina atualizada com sucesso!")
            return redirect('procedures:disciplinas_list')
    else:
        form = DisciplinaForm(instance=disc)
    
    return render(request, 'procedures/disciplina_form.html', {
        'form': form,
        'disc': disc,
        'titulo': f'Editar Disciplina: {disc.codigo}'
    })


@login_required
def detalhe_disciplina_view(request, disciplina_id):
    """Visualiza detalhes de disciplina com procedimentos associados."""
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    # Buscar procedimentos associados
    from procedures.models import DisciplinaProcedimento, Procedimento
    
    procedimentos_associados = DisciplinaProcedimento.objects.filter(
        disciplina=disc
    ).select_related('procedimento').order_by('ordem', 'procedimento__codigo')
    
    # Buscar procedimentos disponíveis para adicionar
    procedimentos_disponiveis = Procedimento.objects.exclude(
        id__in=procedimentos_associados.values_list('procedimento_id', flat=True)
    ).order_by('codigo')[:100]

    total_procedimentos_associados = procedimentos_associados.count()
    
    return render(request, 'procedures/disciplina_detalhe.html', {
        'disc': disc,
        'procedimentos_associados': procedimentos_associados,
        'procedimentos_disponiveis': procedimentos_disponiveis,
        'total_procedimentos_associados': total_procedimentos_associados,
        'pode_deletar_disciplina': total_procedimentos_associados == 0,
    })


@login_required
def adicionar_procedimento_disciplina_view(request, disciplina_id):
    """Adiciona um procedimento à disciplina."""
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    if request.method == 'POST':
        from procedures.models import DisciplinaProcedimento, Procedimento
        
        procedimento_id = request.POST.get('procedimento_id')
        ordem = request.POST.get('ordem', 0)
        obrigatorio = request.POST.get('obrigatorio') == 'on'
        
        try:
            procedimento = Procedimento.objects.get(id=procedimento_id)
            
            # Verificar se já está associado
            if DisciplinaProcedimento.objects.filter(
                disciplina=disc,
                procedimento=procedimento
            ).exists():
                messages.warning(request, f'O procedimento {procedimento.codigo} já está associado a esta disciplina.')
            else:
                DisciplinaProcedimento.objects.create(
                    disciplina=disc,
                    procedimento=procedimento,
                    ordem=int(ordem) if ordem else 0,
                    obrigatorio=obrigatorio
                )
                messages.success(request, f'Procedimento {procedimento.codigo} adicionado com sucesso!')
        except Procedimento.DoesNotExist:
            messages.error(request, 'Procedimento não encontrado.')
        except Exception as e:
            messages.error(request, f'Erro ao adicionar procedimento: {str(e)}')
    
    return redirect('procedures:detalhe_disciplina', disciplina_id=disciplina_id)


@login_required
def remover_procedimento_disciplina_view(request, disciplina_id, assoc_id):
    """Remove um procedimento da disciplina."""
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    if request.method == 'POST':
        from procedures.models import DisciplinaProcedimento
        
        try:
            assoc = DisciplinaProcedimento.objects.get(id=assoc_id, disciplina=disc)
            proc_codigo = assoc.procedimento.codigo
            assoc.delete()
            messages.success(request, f'Procedimento {proc_codigo} removido da disciplina.')
        except DisciplinaProcedimento.DoesNotExist:
            messages.error(request, 'Associação não encontrada.')
        except Exception as e:
            messages.error(request, f'Erro ao remover procedimento: {str(e)}')
    
    return redirect('procedures:detalhe_disciplina', disciplina_id=disciplina_id)


@login_required
def filtrar_procedimentos_view(request, disciplina_id):
    """API AJAX para filtrar procedimentos com base em critérios."""
    from procedures.models import DisciplinaProcedimento, Procedimento
    
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    # Procedimentos já associados
    associados = DisciplinaProcedimento.objects.filter(
        disciplina=disc
    ).values_list('procedimento_id', flat=True)
    
    # Query base
    query = Procedimento.objects.exclude(id__in=associados)
    
    # Filtros
    busca = request.GET.get('busca', '').strip()
    matriz = request.GET.get('matriz', '').strip()
    subarea = request.GET.get('subarea', '').strip()
    
    if busca:
        query = query.filter(
            Q(codigo__icontains=busca) | 
            Q(nome__icontains=busca) |
            Q(descricao__icontains=busca)
        )
    
    if matriz:
        query = query.filter(matriz=matriz)
    
    if subarea:
        query = query.filter(sub_area=subarea)
    
    # Ordenar e limitar
    query = query.order_by('codigo')[:200]
    
    # Serializar resultado
    resultado = [
        {
            'id': p.id,
            'codigo': p.codigo,
            'nome': p.nome or '',
            'matriz': p.matriz or '',
            'sub_area': p.sub_area or '',
        }
        for p in query
    ]
    
    return JsonResponse(resultado, safe=False)


@login_required
def obter_opcoes_filtro_view(request, disciplina_id):
    """API AJAX para obter opções de filtro (matrizes e subáreas)."""
    from django.http import JsonResponse
    from procedures.models import DisciplinaProcedimento, Procedimento
    
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    # Procedimentos já associados
    associados = DisciplinaProcedimento.objects.filter(
        disciplina=disc
    ).values_list('procedimento_id', flat=True)
    
    # Disponíveis
    disponiveis = Procedimento.objects.exclude(id__in=associados)
    
    # Matrizes únicas
    matrizes = sorted(
        set(disponiveis.filter(matriz__isnull=False).values_list('matriz', flat=True))
    )
    
    # Subáreas únicas
    subareas = sorted(
        set(disponiveis.filter(sub_area__isnull=False).values_list('sub_area', flat=True))
    )
    
    return JsonResponse({
        'matrizes': list(matrizes),
        'subareas': list(subareas),
    })


@login_required
def adicionar_multiplos_procedimentos_view(request, disciplina_id):
    """Adiciona múltiplos procedimentos à disciplina de uma vez."""
    from procedures.models import DisciplinaProcedimento, Procedimento
    
    disc = get_object_or_404(Disciplina, id=disciplina_id)
    
    if request.method == 'POST':
        procedimento_ids = request.POST.getlist('procedimento_ids[]')
        ordem_base = int(request.POST.get('ordem', 0))
        obrigatorio = request.POST.get('obrigatorio') == 'on'
        
        adicionados = []
        duplicatas = []
        erros = []
        
        try:
            # Obter maior ordem existente
            max_ordem = DisciplinaProcedimento.objects.filter(
                disciplina=disc
            ).aggregate(Max('ordem'))['ordem__max'] or 0
            
            ordem_atual = max(ordem_base, max_ordem + 1)
            
            for proc_id in procedimento_ids:
                try:
                    procedimento = Procedimento.objects.get(id=proc_id)
                    
                    # Verificar duplicata
                    if DisciplinaProcedimento.objects.filter(
                        disciplina=disc,
                        procedimento=procedimento
                    ).exists():
                        duplicatas.append(procedimento.codigo)
                    else:
                        DisciplinaProcedimento.objects.create(
                            disciplina=disc,
                            procedimento=procedimento,
                            ordem=ordem_atual,
                            obrigatorio=obrigatorio
                        )
                        adicionados.append(procedimento.codigo)
                        ordem_atual += 1
                except Procedimento.DoesNotExist:
                    erros.append(f'Procedimento ID {proc_id} não encontrado')
            
            # Montar mensagem de feedback
            if adicionados:
                messages.success(
                    request,
                    f'{len(adicionados)} procedimento(s) adicionado(s): {", ".join(adicionados)}'
                )
            
            if duplicatas:
                messages.warning(
                    request,
                    f'{len(duplicatas)} procedimento(s) já estava(m) associado(s): {", ".join(duplicatas)}'
                )
            
            if erros:
                messages.error(
                    request,
                    f'Erro(s) ao adicionar: {", ".join(erros)}'
                )
        
        except Exception as e:
            messages.error(request, f'Erro ao adicionar procedimentos: {str(e)}')
    
    return redirect('procedures:detalhe_disciplina', disciplina_id=disciplina_id)


# ==============================================================================
# MATRIZES DE HABILIDADE
# ==============================================================================

@login_required
def matrizes_list_view(request):
    """Lista de matrizes de habilidade com filtros."""
    termo = request.GET.get('q', '').strip()
    ativo = request.GET.get('ativo', '')
    
    qs = MatrizHabilidade.objects.all()
    
    if termo:
        qs = qs.filter(Q(codigo__icontains=termo) | Q(nome__icontains=termo))
    
    if ativo:
        qs = qs.filter(ativo=(ativo == 'true'))
    
    paginator = Paginator(qs.order_by('codigo'), 50)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    
    ctx = {
        'matrizes': page_obj.object_list,
        'page_obj': page_obj,
        'termo': termo,
        'ativo': ativo,
    }
    return render(request, 'procedures/matriz_lista.html', ctx)


@login_required
def nova_matriz_view(request):
    """Cria nova matriz de habilidade."""
    if request.method == 'POST':
        form = MatrizHabilidadeForm(request.POST)
        if form.is_valid():
            matriz = form.save()
            messages.success(request, f"Matriz {matriz.codigo} criada com sucesso!")
            return redirect('procedures:matrizes_list')
    else:
        form = MatrizHabilidadeForm()
    
    return render(request, 'procedures/matriz_form.html', {
        'form': form,
        'titulo': 'Nova Matriz de Habilidade'
    })


@login_required
def editar_matriz_view(request, matriz_id):
    """Edita matriz de habilidade existente."""
    matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
    
    if request.method == 'POST':
        form = MatrizHabilidadeForm(request.POST, instance=matriz)
        if form.is_valid():
            form.save()
            messages.success(request, "Matriz atualizada com sucesso!")
            return redirect('procedures:matrizes_list')
    else:
        form = MatrizHabilidadeForm(instance=matriz)
    
    return render(request, 'procedures/matriz_form.html', {
        'form': form,
        'matriz': matriz,
        'titulo': f'Editar Matriz: {matriz.codigo}'
    })


@login_required
def detalhe_matriz_view(request, matriz_id):
    """Visualiza detalhes de matriz de habilidade."""
    matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
    disciplinas = matriz.disciplinas_matriz.annotate(
        total_procedimentos_associados=Count('procedimentos_associados')
    ).order_by('codigo')
    
    return render(request, 'procedures/matriz_detalhe.html', {
        'matriz': matriz,
        'disciplinas': disciplinas
    })


# ==============================================================================
# API - ASSOCIAÇÃO DE COLABORADORES A MATRIZES
# ==============================================================================

@login_required
@require_http_methods(["GET", "POST"])
def matriz_colaboradores_api(request, matriz_id):
    """API para gerenciar colaboradores associados a uma matriz."""
    matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
    
    if request.method == 'GET':
        # Parâmetros de filtro da query string
        termo = request.GET.get('termo', '').strip()
        setor_id = request.GET.get('setor', '').strip()
        lider_id = request.GET.get('lider', '').strip()
        supervisor_id = request.GET.get('supervisor', '').strip()
        turno = request.GET.get('turno', '').strip()
        
        # Construir query base com colaboradores ativos
        qs = Colaborador.objects.all()
        
        # Aplicar filtros
        if termo:
            qs = qs.filter(
                Q(nome_completo__icontains=termo) | 
                Q(matricula__icontains=termo)
            )
        
        if setor_id:
            qs = qs.filter(setor_id=setor_id)
        
        if lider_id:
            qs = qs.filter(lider_id=lider_id)
        
        if supervisor_id:
            qs = qs.filter(supervisor_id=supervisor_id)
        
        if turno:
            qs = qs.filter(turno=turno)
        
        # Ordena por nome
        qs = qs.order_by('nome_completo')
        
        # Limite de resultados (máximo 100)
        qs = qs[:100]
        
        # IDs dos colaboradores já associados
        associados_ids = set(
            ColaboradorMatrizHabilidade.objects
            .filter(matriz=matriz, ativo=True)
            .values_list('colaborador_id', flat=True)
        )
        
        resultado = {
            'matriz': {
                'id': matriz.id,
                'codigo': matriz.codigo,
                'nome': matriz.nome
            },
            'colaboradores': [
                {
                    'id': colab.id,
                    'nome': colab.nome_completo,
                    'matricula': colab.matricula or '',
                    'setor': str(colab.setor) if colab.setor else '',
                    'associado': colab.id in associados_ids
                }
                for colab in qs
            ]
        }
        return JsonResponse(resultado)
    
    elif request.method == 'POST':
        # Salva associações de colaboradores (apenas adiciona)
        try:
            data = json.loads(request.body)
            colaboradores_ids = data.get('colaboradores', [])
            
            # Colaboradores que devem estar associados
            colaboradores_novos = set(colaboradores_ids)
            
            # Colaboradores que já estão associados
            colaboradores_existentes = set(
                ColaboradorMatrizHabilidade.objects
                .filter(matriz=matriz, ativo=True)
                .values_list('colaborador_id', flat=True)
            )
            
            # Apenas adicionar novas associações (não remover as antigas)
            para_adicionar = colaboradores_novos - colaboradores_existentes
            for colab_id in para_adicionar:
                colaborador = get_object_or_404(Colaborador, id=colab_id)
                ColaboradorMatrizHabilidade.objects.get_or_create(
                    colaborador=colaborador,
                    matriz=matriz,
                    defaults={'ativo': True}
                )
            
            return JsonResponse({
                'success': True,
                'message': f'{len(para_adicionar)} novo(s) colaborador(es) associado(s) com sucesso!'
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)


@login_required
@require_http_methods(["DELETE"])
def remover_colaborador_matriz(request, assoc_id):
    """Remove um colaborador de uma matriz de habilidade."""
    try:
        assoc = get_object_or_404(ColaboradorMatrizHabilidade, id=assoc_id)
        assoc.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Colaborador removido com sucesso!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET", "POST"])
def filtros_colaboradores_api(request):
    """Retorna opções para filtros de colaboradores."""
    try:
        # Setores únicos - usar o relacionamento correto
        setores = Colaborador.objects.filter(
            setor__isnull=False
        ).values('setor__id', 'setor__nome').distinct().order_by('setor__nome')
        
        setores_list = [
            {'id': item['setor__id'], 'nome': item['setor__nome']}
            for item in setores
        ]
        
        # Líderes únicos
        lideres = Colaborador.objects.filter(
            lider__isnull=False
        ).values('lider__id', 'lider__nome_completo').distinct().order_by('lider__nome_completo')
        
        lideres_list = [
            {'id': item['lider__id'], 'nome_completo': item['lider__nome_completo']}
            for item in lideres
        ]
        
        # Supervisores únicos
        supervisores = Colaborador.objects.filter(
            supervisor__isnull=False
        ).values('supervisor__id', 'supervisor__nome_completo').distinct().order_by('supervisor__nome_completo')
        
        supervisores_list = [
            {'id': item['supervisor__id'], 'nome_completo': item['supervisor__nome_completo']}
            for item in supervisores
        ]
        
        # Turnos únicos
        turnos = Colaborador.objects.filter(
            turno__isnull=False
        ).values_list('turno', flat=True).distinct()
        
        turnos_list = [
            {'valor': turno, 'label': turno}
            for turno in sorted(turnos) if turno
        ]
        
        return JsonResponse({
            'setores': setores_list,
            'lideres': lideres_list,
            'supervisores': supervisores_list,
            'turnos': turnos_list
        })
    except Exception as e:
        print(f"Erro em filtros_colaboradores_api: {str(e)}")
        return JsonResponse({
            'error': str(e),
            'setores': [],
            'lideres': [],
            'supervisores': [],
            'turnos': []
        }, status=400)


@login_required
def avaliar_matriz_view(request, matriz_id):
    """Página de avaliação de colaboradores por disciplinas em formato GRID."""
    matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
    
    # Pegar colaboradores associados à matriz
    colaboradores_asociacoes = ColaboradorMatrizHabilidade.objects.filter(
        matriz=matriz, ativo=True
    ).select_related('colaborador').order_by('colaborador__nome_completo')
    
    # Pegar disciplinas da matriz
    disciplinas = matriz.disciplinas_matriz.all().order_by('codigo')
    
    # Construir dados para o grid
    grid_data = []
    for assoc in colaboradores_asociacoes:
        colab = assoc.colaborador
        linha = {
            'colaborador_id': colab.id,
            'colaborador_nome': colab.nome_completo,
            'matricula': colab.matricula,
            'avaliacoes': {}
        }
        
        # Pegar avaliações para cada disciplina
        for disc in disciplinas:
            try:
                avaliacao = AvaliacaoHabilidade.objects.get(
                    colaborador=colab,
                    disciplina=disc,
                    matriz=matriz
                )
                linha['avaliacoes'][disc.id] = {
                    'id': avaliacao.id,
                    'nivel': avaliacao.nivel,
                    'data_avaliacao': avaliacao.data_avaliacao.isoformat() if avaliacao.data_avaliacao else '',
                    'observacoes': avaliacao.observacoes or '',
                    'avaliador_nome': avaliacao.avaliador.nome_completo if avaliacao.avaliador else 'Sistema',
                    'atualizado_em': avaliacao.atualizado_em.isoformat() if avaliacao.atualizado_em else ''
                }
            except AvaliacaoHabilidade.DoesNotExist:
                linha['avaliacoes'][disc.id] = {
                    'id': None,
                    'nivel': '',
                    'data_avaliacao': '',
                    'observacoes': '',
                    'avaliador_nome': '',
                    'atualizado_em': ''
                }
        
        grid_data.append(linha)
    
    ctx = {
        'matriz': matriz,
        'disciplinas': disciplinas,
        'grid_data': grid_data,
        'colaboradores_count': len(grid_data),
        'disciplinas_count': len(disciplinas)
    }
    
    return render(request, 'procedures/matriz_avaliacao_grid.html', ctx)


@login_required
@require_http_methods(["POST"])
def salvar_avaliacao_api(request):
    """
    API para salvar avaliação de um colaborador para uma disciplina.
    
    Recebe JSON:
    {
        'colaborador_id': int,
        'disciplina_id': int,
        'matriz_id': int,
        'nivel': str (1-5) ou null para remover,
        'data_avaliacao': str (YYYY-MM-DD) - opcional,
        'observacoes': str - opcional
    }
    """
    try:
        data = json.loads(request.body)
        
        colaborador_id = data.get('colaborador_id')
        disciplina_id = data.get('disciplina_id')
        matriz_id = data.get('matriz_id')
        nivel = data.get('nivel')
        data_avaliacao = data.get('data_avaliacao')
        observacoes = data.get('observacoes')
        
        # Validações básicas
        if not all([colaborador_id, disciplina_id, matriz_id]):
            return JsonResponse({
                'success': False,
                'message': 'Parâmetros obrigatórios faltando'
            }, status=400)
        
        # Verificar se colaborador está associado à matriz
        assoc = ColaboradorMatrizHabilidade.objects.filter(
            colaborador_id=colaborador_id,
            matriz_id=matriz_id,
            ativo=True
        ).first()
        
        if not assoc:
            return JsonResponse({
                'success': False,
                'message': 'Colaborador não está associado a esta matriz'
            }, status=403)
        
        # Obter objetos
        colaborador = get_object_or_404(Colaborador, id=colaborador_id)
        disciplina = get_object_or_404(Disciplina, id=disciplina_id)
        matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
        
        # Se nivel é vazio/null, deletar avaliação se existir
        if not nivel:
            AvaliacaoHabilidade.objects.filter(
                colaborador=colaborador,
                disciplina=disciplina,
                matriz=matriz
            ).delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Avaliação removida'
            })
        
        # Validar nivel
        if nivel not in ['1', '2', '3', '4', '5']:
            return JsonResponse({
                'success': False,
                'message': 'Nível inválido. Deve ser 1-5'
            }, status=400)
        
        # Criar ou atualizar avaliação
        from datetime import datetime
        update_dict = {
            'nivel': nivel,
            'avaliador': request.user,
        }
        
        # Adicionar observações se fornecidas
        if observacoes is not None:
            update_dict['observacoes'] = observacoes
        
        # Se data_avaliacao foi fornecida, usar ela, senão usar data de hoje
        if data_avaliacao:
            try:
                update_dict['data_avaliacao'] = datetime.strptime(data_avaliacao, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato de data inválido. Use YYYY-MM-DD'
                }, status=400)
        else:
            # Se não tiver data anterior e não forneceu data, usar hoje
            from datetime import date
            avaliacao_anterior = AvaliacaoHabilidade.objects.filter(
                colaborador=colaborador,
                disciplina=disciplina,
                matriz=matriz
            ).first()
            if avaliacao_anterior and avaliacao_anterior.data_avaliacao:
                update_dict['data_avaliacao'] = avaliacao_anterior.data_avaliacao
            else:
                update_dict['data_avaliacao'] = date.today()
        
        avaliacao, created = AvaliacaoHabilidade.objects.update_or_create(
            colaborador=colaborador,
            disciplina=disciplina,
            matriz=matriz,
            defaults=update_dict
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Avaliação salva com sucesso',
            'avaliacao_id': avaliacao.id,
            'created': created
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'JSON inválido'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao salvar: {str(e)}'
        }, status=500)


# ==============================================================================
# DELETE OPERATIONS
# ==============================================================================

@login_required
@require_http_methods(["POST"])
def deletar_matriz_view(request, matriz_id):
    """Deleta uma matriz de habilidade e todas as disciplinas/avaliações associadas."""
    try:
        matriz = get_object_or_404(MatrizHabilidade, id=matriz_id)
        
        # Deletar todas as disciplinas associadas (e suas avaliações em cascata)
        matriz.disciplinas_matriz.all().delete()
        
        # Deletar a matriz
        matriz.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Matriz e todas as disciplinas associadas foram deletadas com sucesso!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def deletar_disciplina_view(request, disciplina_id):
    """Deleta uma disciplina de habilidade somente quando não há procedimentos associados."""
    try:
        from procedures.models import DisciplinaProcedimento

        disciplina = get_object_or_404(Disciplina, id=disciplina_id)
        total_associacoes = DisciplinaProcedimento.objects.filter(disciplina=disciplina).count()

        if total_associacoes > 0:
            return JsonResponse({
                'success': False,
                'message': (
                    'Não é possível remover a disciplina porque existem '
                    f'{total_associacoes} procedimento(s) associado(s).'
                )
            }, status=400)

        disciplina.delete()

        return JsonResponse({
            'success': True,
            'message': 'Disciplina deletada com sucesso!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

# ==============================================================================
# IMPORTAÇÃO EM MASSA
# ==============================================================================

@login_required
def importacao_matriz_view(request):
    """Tela para importação em massa de matrizes, disciplinas e colaboradores."""
    from procedures.forms.forms import ImportacaoMatrizHabilidadeForm
    from procedures.utils.importacao_matriz import gerar_template_csv, gerar_template_excel
    
    if request.method == 'POST':
        form = ImportacaoMatrizHabilidadeForm(request.POST, request.FILES)
        if form.is_valid():
            return processar_importacao_matriz(request, form)
    else:
        form = ImportacaoMatrizHabilidadeForm()
    
    ctx = {
        'form': form,
        'template_csv': gerar_template_csv(),
        'template_excel': gerar_template_excel(),
    }
    return render(request, 'procedures/matriz_importacao.html', ctx)


def processar_importacao_matriz(request, form):
    """Processa o arquivo de importação."""
    from procedures.utils.importacao_matriz import ImportadorMatrizHabilidade, validar_arquivo_importacao
    
    try:
        arquivo = request.FILES.get('arquivo')
        formato = form.cleaned_data.get('formato')
        
        # Validar arquivo
        valido, erro = validar_arquivo_importacao(arquivo)
        if not valido:
            messages.error(request, f"Erro na validação: {erro}")
            return redirect('procedures:importacao_matriz')
        
        # Processar importação
        importador = ImportadorMatrizHabilidade()
        
        if formato == 'csv':
            sucesso = importador.processar_csv(arquivo)
        else:  # excel
            sucesso = importador.processar_excel(arquivo)
        
        # Gerar resumo
        resumo = importador.obter_resumo()
        
        # Armazenar na sessão para exibição
        request.session['importacao_resumo'] = resumo
        request.session['importacao_sucesso'] = sucesso
        
        return redirect('procedures:importacao_matriz_resultado')
    
    except Exception as e:
        messages.error(request, f"Erro ao processar arquivo: {str(e)}")
        return redirect('procedures:importacao_matriz')


@login_required
def importacao_matriz_resultado_view(request):
    """Exibe resultado da importação."""
    resumo = request.session.pop('importacao_resumo', {})
    sucesso = request.session.pop('importacao_sucesso', False)
    
    if not resumo:
        return redirect('procedures:importacao_matriz')
    
    ctx = {
        'resumo': resumo,
        'sucesso': sucesso,
        'total_erros': len(resumo.get('erros', [])),
        'total_avisos': len(resumo.get('avisos', [])),
    }
    return render(request, 'procedures/matriz_importacao_resultado.html', ctx)


@login_required
def baixar_template_importacao_view(request, formato='csv'):
    """Download de template para importação."""
    from django.http import HttpResponse
    from procedures.utils.importacao_matriz import gerar_template_csv
    
    if formato == 'csv':
        conteudo = gerar_template_csv()
        response = HttpResponse(conteudo, content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="template_matrizes.csv"'
        return response
    
    elif formato == 'excel':
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matrizes"
        
        # Headers
        headers = [
            'Matriz Código',
            'Matriz Nome',
            'Disciplina Código',
            'Disciplina Nome',
            'Colaborador Matrícula',
            'Colaborador Nome',
            'Nível de Competência',
            'Observações'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Dados de exemplo
        exemplo_dados = [
            ['MAT001', 'Operação', 'DISC001', 'Segurança', 'MAT001', 'João Silva', '2', 'Necessita aprimoramento'],
            ['MAT001', 'Operação', 'DISC002', 'Qualidade', 'MAT002', 'Maria Santos', '3', 'Em dia com treinamentos'],
            ['MAT002', 'Manutenção', 'DISC003', 'Manutenção Preventiva', 'MAT003', 'Pedro Costa', 'N/A', 'Não se aplica para esta disciplina'],
        ]
        
        for row_idx, row_data in enumerate(exemplo_dados, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_matrizes.xlsx"'
        return response
    
    else:
        return redirect('procedures:importacao_matriz')


# ==============================================================================
# EXPORTAÇÃO DE MATRIZES
# ==============================================================================

@login_required
@require_http_methods(["GET"])
def exportar_matrizes_view(request, formato='csv'):
    """
    Exporta todas as matrizes, disciplinas e colaboradores em CSV ou Excel.
    """
    from django.http import HttpResponse
    from procedures.utils.exportacao_matriz import ExportadorMatrizHabilidade
    
    try:
        exportador = ExportadorMatrizHabilidade()
        
        if formato == 'csv':
            output, filename = exportador.exportar_csv()
            response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        elif formato == 'excel':
            output, filename = exportador.exportar_excel()
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return redirect('procedures:matrizes_list')
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    except Exception as e:
        messages.error(request, f"Erro ao exportar matrizes: {str(e)}")
        return redirect('procedures:matrizes_list')