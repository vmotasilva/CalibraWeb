# -*- coding: utf-8 -*-
"""
Serviço de Geração de Lista de Presença em Excel
Baseado em substituição de tags dinâmicas no template FOR.033.r07
"""

import os
from io import BytesIO
from copy import copy
import openpyxl
from django.conf import settings
from procedures.models import PlanejamentoTreinamento

# Constantes Unicode para Marcação de Checkboxes / Radios
CHECK_ON = "●"   # \u25cf
CHECK_OFF = "○"  # \u25cb


def _obter_mapeamento_checkboxes(planejamento: PlanejamentoTreinamento, overrides: dict = None) -> dict:
    """
    Avalia os dados do planejamento, procedimentos e possíveis respostas do usuário
    para determinar os estados dos checkboxes (● ou ○).
    """
    mapping = {}
    overrides = overrides or {}

    # -------------------------------------------------------------
    # 1. Categoria do Treinamento (Treinamento, Reunião, Reciclagem)
    # -------------------------------------------------------------
    cat_override = overrides.get('categoria')
    if cat_override:
        cat_sel = cat_override.upper()
        mapping["{{CHK_TREIN}}"] = CHECK_ON if cat_sel in ['TREIN', 'TREINAMENTO'] else CHECK_OFF
        mapping["{{CHK_REUN}}"] = CHECK_ON if cat_sel in ['REUN', 'REUNIAO', 'REUNIÃO'] else CHECK_OFF
        mapping["{{CHK_RECIC}}"] = CHECK_ON if cat_sel in ['RECIC', 'RECICLAGEM'] else CHECK_OFF
        mapping["{{CHK_INTEGRACAO}}"] = CHECK_ON if cat_sel in ['INTEG', 'INTEGRACAO', 'INTEGRAÇÃO'] else CHECK_OFF
    else:
        origem = (getattr(planejamento, 'origem', '') or '').upper()
        titulo = (planejamento.titulo or '').upper()
        obs = (getattr(planejamento, 'observacoes', '') or '').upper()

        is_reuniao = "REUNIAO" in origem or "REUNIÃO" in titulo or "REUNIAO" in titulo
        is_reciclagem = "RECIC" in origem or "RECICLAGEM" in titulo or "RECICLAGEM" in obs
        is_integracao = "INTEGRACAO" in origem or "INTEGRAÇÃO" in titulo
        is_treinamento = not (is_reuniao or is_reciclagem or is_integracao)

        mapping["{{CHK_TREIN}}"] = CHECK_ON if is_treinamento else CHECK_OFF
        mapping["{{CHK_REUN}}"] = CHECK_ON if is_reuniao else CHECK_OFF
        mapping["{{CHK_RECIC}}"] = CHECK_ON if is_reciclagem else CHECK_OFF
        mapping["{{CHK_INTEGRACAO}}"] = CHECK_ON if is_integracao else CHECK_OFF

    # -------------------------------------------------------------
    # 2. Metodologia (Tradicional / Teórico vs LOFT / Prático)
    # -------------------------------------------------------------
    met_override = overrides.get('metodologia')
    if met_override:
        met_sel = met_override.upper()
        is_loft = met_sel in ['LOFT', 'PRATICA', 'PRÁTICA']
    else:
        metodologia = getattr(planejamento, 'metodologia', '')
        if metodologia:
            is_loft = "LOFT" in str(metodologia).upper() or "PRAT" in str(metodologia).upper()
        else:
            # Padrão: Tradicional
            is_loft = False

    mapping["{{CHK_LOFT}}"] = CHECK_ON if is_loft else CHECK_OFF
    mapping["{{CHK_TRAD}}"] = CHECK_OFF if is_loft else CHECK_ON

    # -------------------------------------------------------------
    # 3. Necessita Avaliação de Eficácia (SIM / NÃO)
    # Regra: Deve estar marcado SIM quando houver ao menos um procedimento crítico
    # -------------------------------------------------------------
    aval_override = overrides.get('necessita_avaliacao')
    if aval_override:
        aval_str = str(aval_override).upper()
        necessita_aval = aval_str in ['SIM', 'TRUE', '1', 'S']
    else:
        tem_procedimento_critico = planejamento.procedimentos.filter(criticidade='CRITICO').exists()
        necessita_aval = getattr(
            planejamento, 
            'necessita_avaliacao', 
            getattr(planejamento, 'avaliacao_eficacia', tem_procedimento_critico)
        )

    mapping["{{CHK_AVAL_SIM}}"] = CHECK_ON if bool(necessita_aval) else CHECK_OFF
    mapping["{{CHK_AVAL_NAO}}"] = CHECK_OFF if bool(necessita_aval) else CHECK_ON

    # -------------------------------------------------------------
    # 4. Área de Conhecimento (Administrativo, Qualidade, EHS, Estoque, Produção, Outros)
    # -------------------------------------------------------------
    area_override = overrides.get('area_conhecimento')
    if area_override:
        area_sel = area_override.upper()
        mapping["{{CHK_ADM}}"] = CHECK_ON if area_sel in ['ADM', 'ADMINISTRATIVO', 'RH'] else CHECK_OFF
        mapping["{{CHK_QUALIDADE}}"] = CHECK_ON if area_sel in ['QUALIDADE', 'SGQ', 'ISO'] else CHECK_OFF
        mapping["{{CHK_EHS}}"] = CHECK_ON if area_sel in ['EHS', 'SEGURANCA', 'MEIO_AMBIENTE'] else CHECK_OFF
        mapping["{{CHK_ESTOQUE}}"] = CHECK_ON if area_sel in ['ESTOQUE', 'ALMOXARIFADO', 'LOGISTICA'] else CHECK_OFF
        mapping["{{CHK_PRODUCAO}}"] = CHECK_ON if area_sel in ['PRODUCAO', 'PRODUÇÃO', 'FABRICA'] else CHECK_OFF
        mapping["{{CHK_OUTROS}}"] = CHECK_ON if area_sel in ['OUTROS', 'OUTRO'] else CHECK_OFF
    else:
        areas = [
            (p.area_conhecimento or '').upper() 
            for p in planejamento.procedimentos.all()
        ]
        areas_str = " ".join(areas) + " " + (planejamento.titulo or '').upper() + " " + (getattr(planejamento, 'observacoes', '') or '').upper()

        is_qualidade = "QUALIDADE" in areas_str or "SGQ" in areas_str or "ISO" in areas_str
        is_ehs = "EHS" in areas_str or "SEGURAN" in areas_str or "AMBIENTE" in areas_str
        is_estoque = "ESTOQUE" in areas_str or "ALMOXARIF" in areas_str or "LOGIST" in areas_str
        is_producao = "PRODU" in areas_str or "FABRICA" in areas_str
        is_adm = "ADM" in areas_str or "ADMINISTRATIV" in areas_str or "RH" in areas_str

        # Se nenhuma for detectada, padrão é Qualidade
        if not (is_qualidade or is_ehs or is_estoque or is_producao or is_adm):
            is_qualidade = True

    # -------------------------------------------------------------
    # 5. Tags Unificadas de Bloco Único (Recomendado - 1 tag por célula)
    # -------------------------------------------------------------
    cat_str = f"{mapping['{{CHK_TREIN}}']}Treinamento   {mapping['{{CHK_REUN}}']}Reunião   {mapping['{{CHK_RECIC}}']}Reciclagem"
    met_str = f"{mapping['{{CHK_LOFT}}']}LOFT   {mapping['{{CHK_TRAD}}']}Tradicional"
    area_str = f"{mapping['{{CHK_ADM}}']}Administrativo   {mapping['{{CHK_QUALIDADE}}']}Qualidade   {mapping['{{CHK_EHS}}']}EHS   {mapping['{{CHK_ESTOQUE}}']}Estoque   {mapping['{{CHK_PRODUCAO}}']}Produção   {mapping['{{CHK_OUTROS}}']}Outros: _______"
    aval_str = f"{mapping['{{CHK_AVAL_SIM}}']}Sim   {mapping['{{CHK_AVAL_NAO}}']}Não"

    mapping["{{CATEGORIA_TREINAMENTO}}"] = cat_str
    mapping["{{CATEGORIA_OPCOES}}"] = cat_str
    mapping["{{CATEGORIA}}"] = cat_str

    mapping["{{METODOLOGIA}}"] = met_str
    mapping["{{METODOLOGIA_OPCOES}}"] = met_str

    mapping["{{AREA_CONHECIMENTO}}"] = area_str
    mapping["{{AREA_OPCOES}}"] = area_str
    mapping["{{AREA}}"] = area_str

    mapping["{{NECESSITA_AVALIACAO}}"] = aval_str
    mapping["{{AVALIACAO_OPCOES}}"] = aval_str
    mapping["{{NECESSITA_AVAL}}"] = aval_str

    return mapping


def _search_and_replace_sheet(sheet, mapping: dict):
    """
    Substitui todas as tags de texto e checkboxes nas células da planilha.
    Suporta tags individuais, tags de bloco único e detecção inteligente de opções estáticas.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valor = str(cell.value).strip()

                # 1. Substituição direta por tags explícitas
                for tag, novo_valor in mapping.items():
                    if tag in valor:
                        valor = valor.replace(tag, str(novo_valor if novo_valor is not None else ''))

                # 2. Detecção inteligente se a célula tiver texto padrão estático com bolinhas (○ / ●)
                # Categoria
                if ('Treinamento' in valor or 'Trein' in valor) and ('Reunião' in valor or 'Reuniao' in valor) and any(b in valor for b in ['○', '●', '()', '[]']):
                    valor = mapping.get("{{CATEGORIA_TREINAMENTO}}", valor)
                # Metodologia
                elif 'LOFT' in valor and ('Tradicional' in valor or 'Trad' in valor) and any(b in valor for b in ['○', '●', '()', '[]']):
                    valor = mapping.get("{{METODOLOGIA}}", valor)
                # Área de Conhecimento
                elif ('Administrativo' in valor or 'Adm' in valor) and 'Qualidade' in valor and 'Produção' in valor and any(b in valor for b in ['○', '●', '()', '[]']):
                    valor = mapping.get("{{AREA_CONHECIMENTO}}", valor)
                # Necessita Avaliação
                elif ('Sim' in valor or 'SIM' in valor) and ('Não' in valor or 'Nao' in valor or 'NÃO' in valor) and any(b in valor for b in ['○', '●', '()', '[]']):
                    valor = mapping.get("{{NECESSITA_AVALIACAO}}", valor)

                cell.value = valor


def _copiar_estilo_celula(origem, destino):
    """Copia a formatação da célula âncora para as novas linhas."""
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)


def gerar_lista_presenca_xlsx(planejamento: PlanejamentoTreinamento, overrides: dict = None) -> BytesIO:
    """Carrega o template e preenche cabeçalhos, checkboxes e procedimentos em todas as abas."""
    import base64
    from procedures.models import TemplateDocumentoTreinamento

    wb = None

    # 1. Prioridade: Buscar template ativo configurado na sessão de Templates de Treinamento
    template_config = TemplateDocumentoTreinamento.objects.filter(
        funcao='LISTA_PRESENCA',
        tipo_arquivo='xlsx',
        ativo=True
    ).first()

    if not template_config:
        template_config = TemplateDocumentoTreinamento.objects.filter(
            codigo__icontains='FOR.033',
            tipo_arquivo='xlsx',
            ativo=True
        ).first()

    if not template_config:
        template_config = TemplateDocumentoTreinamento.objects.filter(
            funcao='LISTA_PRESENCA',
            tipo_arquivo='xlsx'
        ).first()

    if not template_config:
        template_config = TemplateDocumentoTreinamento.objects.filter(
            codigo__icontains='FOR.033',
            tipo_arquivo='xlsx'
        ).first()

    if template_config:
        # A. Tentar ler do Base64 gravado no banco de dados (100% persistente em nuvem)
        if getattr(template_config, 'arquivo_base64', None):
            try:
                arquivo_bytes = base64.b64decode(template_config.arquivo_base64)
                wb = openpyxl.load_workbook(BytesIO(arquivo_bytes))
            except Exception:
                wb = None

        # B. Tentar ler do FileField se não carregou por Base64
        if wb is None and template_config.arquivo:
            try:
                template_config.arquivo.seek(0)
                arquivo_bytes = BytesIO(template_config.arquivo.read())
                wb = openpyxl.load_workbook(arquivo_bytes)
            except Exception:
                wb = None

    # 2. Fallback: Diretórios padrão do sistema de arquivos
    if wb is None:
        import glob
        candidatos = [
            os.path.join(settings.BASE_DIR, "FOR.033.r07_Lista_de_Presenca_de_Treinamento.xlsx"),
            os.path.join(settings.BASE_DIR, "templates", "FOR.033.r07_Lista_de_Presenca_de_Treinamento.xlsx"),
            os.path.join(settings.BASE_DIR, "procedures", "templates", "FOR.033.r07_Lista_de_Presenca_de_Treinamento.xlsx"),
            os.path.join(settings.BASE_DIR, "static", "templates", "FOR.033.r07_Lista_de_Presenca_de_Treinamento.xlsx"),
            os.path.join(settings.MEDIA_ROOT, "templates_treinamento_docs", "FOR.033.r07_Lista_de_Presenca_de_Treinamento.xlsx"),
        ]
        candidatos.extend(glob.glob(os.path.join(settings.BASE_DIR, "*FOR.033*.xlsx")))
        candidatos.extend(glob.glob(os.path.join(settings.BASE_DIR, "*Lista*Presen*.xlsx")))
        candidatos.extend(glob.glob(os.path.join(settings.BASE_DIR, "**", "*FOR.033*.xlsx"), recursive=True))

        for path in candidatos:
            if os.path.exists(path) and os.path.isfile(path):
                try:
                    wb = openpyxl.load_workbook(path)
                    if wb:
                        break
                except Exception:
                    pass

    if wb is None:
        raise FileNotFoundError("O arquivo do template ativo precisa ser recadastrado. Acesse 'Templates de Documentos' e faça o upload do arquivo para salvar na nuvem.")

    # 1. Montar substituição de texto com formato "DD/MM/YYYY às HH:MM"
    if planejamento.horario_previsto:
        data_hora_str = planejamento.horario_previsto.strftime("%d/%m/%Y às %H:%M")
    elif planejamento.data_prevista and hasattr(planejamento, 'horario_inicio') and planejamento.horario_inicio:
        data_hora_str = f"{planejamento.data_prevista.strftime('%d/%m/%Y')} às {planejamento.horario_inicio.strftime('%H:%M')}"
    elif planejamento.data_prevista:
        data_hora_str = planejamento.data_prevista.strftime("%d/%m/%Y")
    else:
        data_hora_str = ""

    carga_horaria_str = f"{planejamento.carga_horaria} Minutos" if planejamento.carga_horaria else ""
    instrutor_nome = planejamento.instrutor.nome_completo if planejamento.instrutor else ""

    substituicoes = {
        "{{TITULO}}": planejamento.titulo or "",
        "{{INSTRUTOR}}": instrutor_nome,
        "{{DATA_HORA}}": data_hora_str,
        "{{CARGA_HORARIA}}": carga_horaria_str,
        # Limpar quaisquer tags de participantes para deixar a folha limpa para preenchimento/assinatura manual
        "{{NOME_PARTICIPANTE}}": "",
        "{{NOME_COLABORADOR}}": "",
        "{{CPF_PARTICIPANTE}}": "",
        "{{MATRICULA_PARTICIPANTE}}": "",
        "{{CARGO_PARTICIPANTE}}": "",
        "{{SETOR_PARTICIPANTE}}": "",
        "{{DEPARTAMENTO_PARTICIPANTE}}": "",
        "{{CPF}}": "",
        "{{CARGO}}": "",
        "{{DEPARTAMENTO}}": "",
    }

    # 2. Incorporar os Checkboxes calculados com base nas perguntas/respostas
    checkbox_mapping = _obter_mapeamento_checkboxes(planejamento, overrides=overrides)
    substituicoes.update(checkbox_mapping)

    procedimentos = list(planejamento.procedimentos.all())

    # 3. Processar TODAS as planilhas do arquivo Excel
    for ws in wb.worksheets:
        # A. Substituição de tags gerais de cabeçalhos, datas e checkboxes
        _search_and_replace_sheet(ws, substituicoes)

        # B. Preenchimento de Procedimentos (Conteúdo do Treinamento)
        proc_anchor_row = None
        proc_col = None

        # 1. Procurar tag explícita {{PROCEDIMENTOS}} ou variações
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val_upper = cell.value.upper()
                    if "{{PROCEDIMENTOS}}" in val_upper or "{{PROCEDIMENTO}}" in val_upper or "{{CONTEUDO_PROGRAMATICO}}" in val_upper:
                        proc_anchor_row = cell.row
                        proc_col = cell.column
                        break
            if proc_anchor_row:
                break

        # 2. Se não achou tag, procurar pela seção "CONTEÚDO DO TREINAMENTO"
        if not proc_anchor_row:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_lower = cell.value.strip().lower()
                        if "conteúdo do treinamento" in val_lower or "conteudo do treinamento" in val_lower:
                            # A linha de procedimentos fica tipicamente 2 linhas abaixo (após o título)
                            proc_anchor_row = cell.row + 2
                            proc_col = cell.column
                            break
                if proc_anchor_row:
                    break

        if proc_anchor_row and proc_col:
            if procedimentos:
                p1 = procedimentos[0]
                ws.cell(row=proc_anchor_row, column=proc_col, value=f"{p1.codigo or ''} - {p1.nome or ''}".strip(" -"))

                for idx, p in enumerate(procedimentos[1:], start=1):
                    t_row = proc_anchor_row + idx
                    texto_proc = f"{p.codigo or ''} - {p.nome or ''}".strip(" -")
                    cell_p = ws.cell(row=t_row, column=proc_col, value=texto_proc)
                    _copiar_estilo_celula(ws.cell(row=proc_anchor_row, column=proc_col), cell_p)
            else:
                ws.cell(row=proc_anchor_row, column=proc_col, value="")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
