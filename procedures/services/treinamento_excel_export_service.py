# -*- coding: utf-8 -*-
"""
Serviço de Exportação de Formulários Oficiais em Excel (.xlsx)
Preenche automaticamente os templates:
1. FOR.133.r01 - Planejamento de Treinamento (Matriz relacional de cronograma)
2. FOR.141.r02 - Auto-Avaliação de Treinamento Crítico (5 perguntas)
3. FOR.142.r01 - Avaliação de Eficácia do Treinamento (Cálculo matemático de elegibilidade + 30 dias)
"""

import os
import re
import glob
import base64
import zipfile
from io import BytesIO
from datetime import date, timedelta
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.utils import timezone
from procedures.models import (
    PlanejamentoTreinamento,
    RegistroTreinamento,
    TemplateDocumentoTreinamento,
    PerguntaAvaliacao
)


# ==============================================================================
# FUNÇÕES AUXILIARES DE MANIPULAÇÃO DE WORKBOOK E TAGS
# ==============================================================================

def _obter_raw_bytes_template(funcao: str, codigo_busca: str, nome_padrao: str):
    """
    Recupera os bytes brutos do arquivo .xlsx original (do banco Neon ou do disco local),
    preservando 100% dos arquivos internos (drawings, vml, media, shapes, charts, rels).
    """
    template_config = TemplateDocumentoTreinamento.objects.filter(
        funcao=funcao,
        ativo=True
    ).first()

    if not template_config and codigo_busca:
        template_config = TemplateDocumentoTreinamento.objects.filter(
            codigo__icontains=codigo_busca,
            ativo=True
        ).first()

    raw_bytes = None

    if template_config:
        if getattr(template_config, 'arquivo_base64', None):
            try:
                raw_bytes = base64.b64decode(template_config.arquivo_base64)
            except Exception:
                raw_bytes = None

        if raw_bytes is None and template_config.arquivo:
            try:
                template_config.arquivo.seek(0)
                raw_bytes = template_config.arquivo.read()
            except Exception:
                raw_bytes = None

    if raw_bytes is None and nome_padrao:
        candidatos = [
            os.path.join(settings.BASE_DIR, nome_padrao),
            os.path.join(settings.BASE_DIR, "templates", nome_padrao),
            os.path.join(settings.BASE_DIR, "procedures", "templates", nome_padrao),
            os.path.join(settings.BASE_DIR, "static", "templates", nome_padrao),
            os.path.join(settings.MEDIA_ROOT, "templates_treinamento_docs", nome_padrao),
        ]
        if codigo_busca:
            candidatos.extend(glob.glob(os.path.join(settings.BASE_DIR, f"*{codigo_busca}*.xlsx")))
            candidatos.extend(glob.glob(os.path.join(settings.BASE_DIR, "**", f"*{codigo_busca}*.xlsx"), recursive=True))

        for path in candidatos:
            if os.path.exists(path) and os.path.isfile(path):
                try:
                    with open(path, 'rb') as f_orig:
                        raw_bytes = f_orig.read()
                    if raw_bytes:
                        break
                except Exception:
                    pass

    return raw_bytes, template_config


def _preencher_template_xlsx_preservando_formas(orig_bytes: bytes, substituicoes: dict, cell_updates: dict = None) -> BytesIO:
    """
    Substitui tags e valores de células diretamente na estrutura XML do arquivo .xlsx,
    mantendo intactas todas as formas gráficas, caixas de texto, imagens, gráficos radar e formatações.
    """
    in_zip = zipfile.ZipFile(BytesIO(orig_bytes), 'r')
    out_buf = BytesIO()
    out_zip = zipfile.ZipFile(out_buf, 'w', compression=zipfile.ZIP_DEFLATED)

    for item in in_zip.infolist():
        data = in_zip.read(item.filename)
        if item.filename.endswith('.xml') or item.filename.endswith('.vml') or item.filename.endswith('.rels'):
            text = data.decode('utf-8', errors='ignore')

            # Atualização pontual de células em planilhas (ex: D3, D4, D5, etc.)
            if item.filename.startswith('xl/worksheets/') and cell_updates:
                for cell_ref, val in cell_updates.items():
                    val_str = str(val if val is not None else '')
                    val_clean = val_str.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

                    pattern_self = rf'(<x:c\s+r="{cell_ref}"(?:\s+s="(\d+)")?[^/]*/>)'
                    pattern_full = rf'(<x:c\s+r="{cell_ref}"(?:\s+s="(\d+)")?[^>]*>)(.*?)(</x:c>)'

                    def repl_self(m):
                        s_attr = f' s="{m.group(2)}"' if m.group(2) else ''
                        return f'<x:c r="{cell_ref}"{s_attr} t="inlineStr"><x:is><x:t>{val_clean}</x:t></x:is></x:c>'

                    def repl_full(m):
                        s_attr = f' s="{m.group(2)}"' if m.group(2) else ''
                        return f'<x:c r="{cell_ref}"{s_attr} t="inlineStr"><x:is><x:t>{val_clean}</x:t></x:is></x:c>'

                    if re.search(pattern_self, text):
                        text = re.sub(pattern_self, repl_self, text)
                    elif re.search(pattern_full, text):
                        text = re.sub(pattern_full, repl_full, text)

            # Substituição de tags gerais {{TAG}} em todo o XML (células, textboxes, shapes)
            for k, v in substituicoes.items():
                v_str = str(v if v is not None else '')
                v_clean = v_str.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                text = text.replace(k, v_clean)

            data = text.encode('utf-8')
        out_zip.writestr(item, data)

    in_zip.close()
    out_zip.close()
    out_buf.seek(0)
    return out_buf


def _carregar_workbook_template(funcao: str, codigo_busca: str, nome_padrao: str):
    """
    Busca e carrega o arquivo openpyxl.Workbook como fallback.
    """
    raw_bytes, template_config = _obter_raw_bytes_template(funcao, codigo_busca, nome_padrao)
    wb = None
    if raw_bytes:
        try:
            wb = openpyxl.load_workbook(BytesIO(raw_bytes))
        except Exception:
            wb = None
    return wb, template_config


def _search_and_replace_sheet(sheet, mapping: dict):
    """
    Substitui todas as tags de texto nas células de uma planilha Excel.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                valor = str(cell.value)
                modificado = False
                for tag, novo_valor in mapping.items():
                    if tag in valor:
                        valor = valor.replace(tag, str(novo_valor if novo_valor is not None else ''))
                        modificado = True
                if modificado:
                    cell.value = valor


def _copiar_estilo_celula(origem, destino):
    """Copia fontes, bordas, preenchimento e alinhamento de uma célula para outra."""
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)


# ==============================================================================
# 1. FOR.133.r01 - PLANEJAMENTO DE TREINAMENTO (MATRIZ DE CRONOGRAMA)
# ==============================================================================

def gerar_planejamento_matriz_for133_xlsx(planejamentos=None, ano_referencia=None, planejamento_unico: PlanejamentoTreinamento = None) -> BytesIO:
    """
    Gera a Matriz de Planejamento de Treinamentos / Cronograma Anual (FOR.133.r01).
    Extrai todos os treinamentos cadastrados (respeitando os filtros aplicados).
    Preenche a matriz de Jan a Dez com 'P' (Planejado) e 'R' (Realizado).
    """
    # Normalizar lista de planejamentos
    if planejamento_unico:
        lista_planejamentos = [planejamento_unico]
    elif isinstance(planejamentos, PlanejamentoTreinamento):
        lista_planejamentos = [planejamentos]
    elif planejamentos is not None:
        lista_planejamentos = list(planejamentos)
    else:
        lista_planejamentos = list(
            PlanejamentoTreinamento.objects.select_related('instrutor')
            .prefetch_related('colaboradores', 'procedimentos')
            .order_by('data_prevista')
        )

    wb, template_obj = _carregar_workbook_template(
        funcao='PLANEJAMENTO_MATRIZ',
        codigo_busca='133',
        nome_padrao='FOR.133.r01_Planejamento_de_Treinamento.xlsx'
    )

    if not ano_referencia:
        if lista_planejamentos and lista_planejamentos[0].data_prevista:
            ano_referencia = lista_planejamentos[0].data_prevista.year
        else:
            ano_referencia = timezone.now().year
    ano_str = str(ano_referencia)

    substituicoes = {
        "{{ANO}}": ano_str,
        "{{TOTAL_TREINAMENTOS}}": str(len(lista_planejamentos)),
        "{{DATA_EMISSAO}}": timezone.now().strftime("%d/%m/%Y"),
    }

    # Se carregou um template existente (.xlsx)
    if wb is not None:
        ws = wb.active
        for sheet in wb.worksheets:
            _search_and_replace_sheet(sheet, substituicoes)

        # Verificar se é o layout oficial do FOR.133 (Cabeçalho nas primeiras linhas)
        eh_layout_oficial_133 = False
        for r in range(1, 6):
            for c in range(1, 10):
                v = str(ws.cell(r, c).value or '').upper()
                if "NOME DO FORMANDO" in v or "TITULO DO TREINAMENTO" in v or "TÍTULO DO TREINAMENTO" in v:
                    eh_layout_oficial_133 = True
                    break

        if eh_layout_oficial_133:
            # Preencher a partir da linha 4
            row_idx = 4
            for plan in lista_planejamentos:
                procs = list(plan.procedimentos.all())
                colabs = list(plan.colaboradores.all())
                instrutor_nome = plan.instrutor.nome_completo if plan.instrutor else "-"
                proc_str = ", ".join([f"{p.codigo} - {p.nome}" for p in procs]) if procs else (plan.titulo or "-")
                
                mes_planejado = plan.data_prevista.month if plan.data_prevista else None
                mes_realizado = plan.data_realizada.month if plan.data_realizada else (plan.data_prevista.month if plan.status == 'REALIZADO' and plan.data_prevista else None)
                ch_min = plan.carga_horaria or 60

                colabs_iter = colabs if colabs else [None]
                for colab in colabs_iter:
                    colab_nome = colab.nome_completo if colab else "A Definir"
                    
                    ws.cell(row=row_idx, column=2, value=proc_str) # Col B: Título
                    ws.cell(row=row_idx, column=3, value=instrutor_nome) # Col C: Formador
                    ws.cell(row=row_idx, column=4, value=colab_nome) # Col D: Formando
                    ws.cell(row=row_idx, column=5, value="Treinamento Técnico / Operacional") # Col E: Metodologia
                    
                    # Carga horária
                    ws.cell(row=row_idx, column=6, value=f"{ch_min//60:02d}:{ch_min%60:02d}") # Col F: hh:mm
                    ws.cell(row=row_idx, column=7, value=round(ch_min/60, 2)) # Col G: h
                    
                    # Colunas de Meses: H (Jan, col 8) a S (Dez, col 19)
                    for m in range(1, 13):
                        col_mes = 7 + m
                        if mes_realizado == m:
                            ws.cell(row=row_idx, column=col_mes, value="R")
                        elif mes_planejado == m:
                            ws.cell(row=row_idx, column=col_mes, value="P")
                        else:
                            ws.cell(row=row_idx, column=col_mes, value="")

                    ws.cell(row=row_idx, column=20, value=f"Status: {plan.status}") # Col T: Observação
                    row_idx += 1
        else:
            # Localizar âncora da tabela de colaboradores/procedimentos
            anchor_row = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        v_up = cell.value.upper()
                        if "{{COLABORADOR}}" in v_up or "{{TABELA_MATRIZ}}" in v_up or "NOME DO COLABORADOR" in v_up:
                            anchor_row = cell.row
                            break
                if anchor_row:
                    break

            if anchor_row:
                linha_atual = anchor_row + 1 if "NOME DO COLABORADOR" in str(ws.cell(anchor_row, 1).value or '').upper() else anchor_row
                for plan in lista_planejamentos:
                    procs = list(plan.procedimentos.all())
                    colabs = list(plan.colaboradores.all())
                    instrutor_nome = plan.instrutor.nome_completo if plan.instrutor else "-"
                    proc_str = ", ".join([f"{p.codigo} - {p.nome}" for p in procs]) if procs else (plan.titulo or "-")
                    
                    mes_planejado = plan.data_prevista.month if plan.data_prevista else None
                    mes_realizado = plan.data_realizada.month if plan.data_realizada else (plan.data_prevista.month if plan.status == 'REALIZADO' and plan.data_prevista else None)

                    colabs_iter = colabs if colabs else [None]
                    for colab in colabs_iter:
                        ws.cell(row=linha_atual, column=1, value=colab.nome_completo if colab else "A Definir")
                        ws.cell(row=linha_atual, column=2, value=colab.matricula or "-")
                        ws.cell(row=linha_atual, column=3, value=colab.cargo or "-")
                        ws.cell(row=linha_atual, column=4, value=proc_str)
                        ws.cell(row=linha_atual, column=5, value=instrutor_nome)
                        
                        # Colunas de Meses (assumindo colunas 6 a 17 para Jan-Dez)
                        for m in range(1, 13):
                            col_mes = 5 + m
                            if col_mes <= ws.max_column:
                                if mes_realizado == m:
                                    ws.cell(row=linha_atual, column=col_mes, value="R")
                                elif mes_planejado == m:
                                    ws.cell(row=linha_atual, column=col_mes, value="P")
                        linha_atual += 1

    else:
        # Gerador nativo de alta fidelidade para FOR.133.r01
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cronograma FOR.133"

        # Estilos Oficiais
        font_header_doc = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_th = Font(name="Arial", size=8.5, bold=True, color="FFFFFF")
        font_td = Font(name="Arial", size=8.5, color="000000")
        font_meta = Font(name="Arial", size=8.5, bold=True, color="334155")
        font_val = Font(name="Arial", size=8.5, color="0F172A")

        fill_header_doc = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_meta = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_th = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        fill_meses_th = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        fill_planejado = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Azul claro
        fill_realizado = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Verde claro

        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Cabeçalho do Formulário
        ws.merge_cells("A1:R1")
        cell_top = ws.cell(row=1, column=1, value=f"FOR.133.r01 - MATRIZ DE PLANEJAMENTO DE TREINAMENTOS / CRONOGRAMA ANUAL ({ano_str})")
        cell_top.font = font_header_doc
        cell_top.fill = fill_header_doc
        cell_top.alignment = align_center
        ws.row_dimensions[1].height = 28

        # Metadados Gerais
        meta_rows = [
            ("Ano de Referência:", ano_str, "Total de Treinamentos:", str(len(lista_planejamentos))),
            ("Data de Emissão:", timezone.now().strftime("%d/%m/%Y"), "Status Geral:", "Filtro Aplicado"),
        ]

        for idx, (lbl1, val1, lbl2, val2) in enumerate(meta_rows, start=2):
            ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=10)
            ws.merge_cells(start_row=idx, start_column=12, end_row=idx, end_column=18)

            ws.cell(row=idx, column=1, value=lbl1)
            ws.cell(row=idx, column=2, value=val1)
            ws.cell(row=idx, column=11, value=lbl2)
            ws.cell(row=idx, column=12, value=val2)

            for col in range(1, 19):
                c = ws.cell(row=idx, column=col)
                c.border = border_thin
                if col in [1, 11]:
                    c.font = font_meta
                    c.fill = fill_meta
                    c.alignment = align_left
                else:
                    c.font = font_val
                    c.alignment = align_left
            ws.row_dimensions[idx].height = 18

        # Linha em branco
        ws.row_dimensions[4].height = 6

        # Cabeçalho da Matriz Relacional
        headers_base = [
            "Nº", "Procedimento / Treinamento", "Colaborador (Formando)", "Matrícula / Cargo",
            "Facilitador", "Carga Horária"
        ]
        meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

        for col_idx, h_text in enumerate(headers_base, start=1):
            c = ws.cell(row=5, column=col_idx, value=h_text)
            c.font = font_th
            c.fill = fill_th
            c.alignment = align_center
            c.border = border_thin

        for col_idx, m_text in enumerate(meses_nomes, start=7):
            c = ws.cell(row=5, column=col_idx, value=m_text)
            c.font = font_th
            c.fill = fill_meses_th
            c.alignment = align_center
            c.border = border_thin

        ws.row_dimensions[5].height = 22

        # Linhas de Dados
        row_num = 6
        item_counter = 1

        for plan in lista_planejamentos:
            procs = list(plan.procedimentos.all())
            colabs = list(plan.colaboradores.all())
            instrutor_nome = plan.instrutor.nome_completo if plan.instrutor else "-"
            proc_str = ", ".join([f"{p.codigo} - {p.nome}" for p in procs]) if procs else (plan.titulo or "-")
            ch_str = f"{plan.carga_horaria} min" if plan.carga_horaria else "-"
            
            mes_planejado = plan.data_prevista.month if plan.data_prevista else None
            mes_realizado = plan.data_realizada.month if plan.data_realizada else (plan.data_prevista.month if plan.status == 'REALIZADO' and plan.data_prevista else None)

            colabs_iter = colabs if colabs else [None]
            for colab in colabs_iter:
                colab_nome = colab.nome_completo if colab else "A Definir"
                colab_info = f"{colab.matricula or '-'} / {colab.cargo or '-'}" if colab else "-"

                row_vals = [
                    item_counter,
                    proc_str,
                    colab_nome,
                    colab_info,
                    instrutor_nome,
                    ch_str,
                ]

                for col_idx, val in enumerate(row_vals, start=1):
                    c = ws.cell(row=row_num, column=col_idx, value=val)
                    c.font = font_td
                    c.border = border_thin
                    c.alignment = align_center if col_idx in [1, 6] else align_left

                # Preenchimento das 12 colunas de meses
                for m_idx in range(1, 13):
                    col_m = 6 + m_idx
                    c_mes = ws.cell(row=row_num, column=col_m)
                    c_mes.border = border_thin
                    c_mes.alignment = align_center
                    c_mes.font = font_td

                    if mes_realizado == m_idx:
                        c_mes.value = "R"
                        c_mes.fill = fill_realizado
                        c_mes.font = Font(name="Arial", size=9, bold=True, color="166534")
                    elif mes_planejado == m_idx:
                        c_mes.value = "P"
                        c_mes.fill = fill_planejado
                        c_mes.font = Font(name="Arial", size=9, bold=True, color="1E40AF")
                    else:
                        c_mes.value = ""

                ws.row_dimensions[row_num].height = 19
                row_num += 1
                item_counter += 1

        # Legenda no Rodapé
        ws.row_dimensions[row_num].height = 8
        row_legenda = row_num + 1

        ws.cell(row=row_legenda, column=2, value="Legenda:")
        ws.cell(row=row_legenda, column=2).font = font_meta

        c_leg_p = ws.cell(row=row_legenda, column=3, value="P = Planejado")
        c_leg_p.fill = fill_planejado
        c_leg_p.font = Font(name="Arial", size=8.5, bold=True, color="1E40AF")
        c_leg_p.border = border_thin
        c_leg_p.alignment = align_center

        c_leg_r = ws.cell(row=row_legenda, column=4, value="R = Realizado")
        c_leg_r.fill = fill_realizado
        c_leg_r.font = Font(name="Arial", size=8.5, bold=True, color="166534")
        c_leg_r.border = border_thin
        c_leg_r.alignment = align_center

        # Ajuste de larguras das colunas
        larguras = {
            'A': 5, 'B': 32, 'C': 26, 'D': 22, 'E': 24, 'F': 14,
            'G': 6, 'H': 6, 'I': 6, 'J': 6, 'K': 6, 'L': 6,
            'M': 6, 'N': 6, 'O': 6, 'P': 6, 'Q': 6, 'R': 6
        }
        for col_letter, width in larguras.items():
            ws.column_dimensions[col_letter].width = width

    raw_output = BytesIO()
    wb.save(raw_output)
    raw_bytes = raw_output.getvalue()

    output = _substituir_tags_no_arquivo_zip(raw_bytes, substituicoes)
    output.seek(0)
    return output


# ==============================================================================
# 2. FOR.141.r02 - AUTO-AVALIAÇÃO DE TREINAMENTO CRÍTICO (5 PERGUNTAS)
# ==============================================================================

def _substituir_tags_no_arquivo_zip(raw_bytes: bytes, mapping: dict) -> BytesIO:
    """
    Descompacta o arquivo Excel em memória e substitui quaisquer tags dinâmicas
    (ex: {{PER_1}}, {{PER_2}}, {{COLABORADOR}}) em todos os arquivos XML internos,
    incluindo desenhos, gráficos radar (drawing1.xml), caixas de texto e relacionamentos.
    """
    import zipfile
    in_buf = BytesIO(raw_bytes)
    out_buf = BytesIO()

    with zipfile.ZipFile(in_buf, 'r') as zin:
        with zipfile.ZipFile(out_buf, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                # Aplicar substituição em todos os arquivos xml, vml e rels
                if item.filename.endswith('.xml') or item.filename.endswith('.vml') or item.filename.endswith('.rels'):
                    try:
                        text = content.decode('utf-8')
                        modificado = False
                        for k, v in mapping.items():
                            if k in text:
                                text = text.replace(k, str(v if v is not None else ''))
                                modificado = True
                        if modificado:
                            content = text.encode('utf-8')
                    except Exception:
                        pass
                zout.writestr(item, content)

    out_buf.seek(0)
    return out_buf


def _obter_perguntas_treinamento(planejamento: PlanejamentoTreinamento, perguntas_selecionadas: list = None) -> list:
    """
    Recupera as 5 perguntas de autoavaliação para o treinamento crítico.
    Se perguntas_selecionadas for informada (como lista de strings ou lista de IDs), prioriza essas.
    Prioridade: Perguntas Selecionadas > Procedimento > Matriz > Perguntas Padrão SGQ.
    """
    perguntas_texto = []

    # 0. Se foram passadas perguntas selecionadas explicitamente
    if perguntas_selecionadas:
        for item in perguntas_selecionadas:
            if isinstance(item, int) or (isinstance(item, str) and item.isdigit()):
                p_obj = PerguntaAvaliacao.objects.filter(id=int(item), ativo=True).first()
                if p_obj and p_obj.enunciado:
                    perguntas_texto.append(p_obj.enunciado)
            elif isinstance(item, str) and item.strip():
                perguntas_texto.append(item.strip())

    if len(perguntas_texto) < 5:
        if hasattr(planejamento, '_mock_procs') and planejamento._mock_procs:
            procs = list(planejamento._mock_procs)
        elif hasattr(planejamento, 'procedimentos'):
            procs = list(planejamento.procedimentos.all())
        else:
            procs = []

        perguntas = []

        # 1. Buscar por Procedimento
        if procs:
            proc_ids = [p.id for p in procs]
            perguntas = list(
                PerguntaAvaliacao.objects.filter(procedimento_id__in=proc_ids, ativo=True)
                .order_by('ordem')[:5 - len(perguntas_texto)]
            )

        # 2. Buscar por Matriz
        if (len(perguntas_texto) + len(perguntas)) < 5 and procs:
            matrizes_nomes = [p.matriz for p in procs if p.matriz]
            if matrizes_nomes:
                perguntas_matriz = list(
                    PerguntaAvaliacao.objects.filter(matriz__nome__in=matrizes_nomes, ativo=True)
                    .order_by('ordem')[:5 - len(perguntas_texto) - len(perguntas)]
                )
                perguntas.extend(perguntas_matriz)

        for p in perguntas:
            if p.enunciado not in perguntas_texto:
                perguntas_texto.append(p.enunciado)

    # 3. Fallback: 5 Perguntas Padrão Técnicas e de Qualidade
    perguntas_padrao = [
        "Qual o objetivo principal deste procedimento operacional e quais os impactos de eventuais não conformidades no processo?",
        "Quais os equipamentos de proteção individual (EPIs), ferramentas e requisitos de segurança obrigatórios para esta atividade?",
        "Descreva a sequência padrão de execução das etapas e os principais parâmetros operacionais a serem rigorosamente controlados.",
        "Quais são os pontos críticos de controle (PCC), tolerâncias permitidas e critérios de aceitação do produto/serviço?",
        "Em caso de desvio, defeito ou falha identificada durante a operação, qual é o fluxo correto de contenção e comunicação imediata?"
    ]

    for p_padrao in perguntas_padrao:
        if len(perguntas_texto) >= 5:
            break
        if p_padrao not in perguntas_texto:
            perguntas_texto.append(p_padrao)

    return perguntas_texto[:5]


def _extrair_dados_colaborador_avaliado(colaborador):
    """
    Extrai todos os dados do colaborador que será avaliado, incluindo:
    - Nome Completo, Matrícula, Cargo/Função
    - Setor / Departamento / Área / Laboratório
    - Gestor / Líder / Supervisor / Gerente direto
    """
    if not colaborador:
        return {
            'nome': "________________________________________",
            'matricula': "_________",
            'cargo': "____________________",
            'setor': "____________________",
            'gestor': "____________________",
            'lider': "____________________",
            'supervisor': "____________________",
            'gerente': "____________________",
            'posto_trabalho': "____________________",
            'centro_custo': "____________________",
        }

    colab_nome = colaborador.nome_completo or ""
    colab_mat = colaborador.matricula or "-"
    colab_cargo = colaborador.cargo or "-"

    # 1. Setor do Colaborador Avaliado
    colab_setor = ""
    if getattr(colaborador, 'setor', None):
        colab_setor = getattr(colaborador.setor, 'nome', '') or getattr(colaborador.setor, 'codigo', '')
    if not colab_setor and getattr(colaborador, 'grupo', None):
        colab_setor = colaborador.grupo
    if not colab_setor and getattr(colaborador, 'posto_trabalho', None):
        colab_setor = colaborador.posto_trabalho
    colab_setor = colab_setor or "-"

    # 2. Gestor do Colaborador Avaliado (Hierarquia / Liderança Direta)
    gestor_obj = None
    posto_lideranca = getattr(colaborador, 'posto_lideranca', None)
    if posto_lideranca == 'SUPERVISOR':
        gestor_obj = getattr(colaborador, 'gerente', None) or getattr(colaborador, 'supervisor', None) or getattr(colaborador, 'lider', None)
    elif posto_lideranca == 'LIDER':
        gestor_obj = getattr(colaborador, 'supervisor', None) or getattr(colaborador, 'gerente', None) or getattr(colaborador, 'lider', None)
    else:
        gestor_obj = getattr(colaborador, 'lider', None) or getattr(colaborador, 'supervisor', None) or getattr(colaborador, 'gerente', None)

    if not gestor_obj and hasattr(colaborador, 'get_chefia'):
        try:
            gestor_obj = colaborador.get_chefia()
        except Exception:
            pass

    if not gestor_obj and getattr(colaborador, 'setor', None) and hasattr(colaborador.setor, 'responsavel') and colaborador.setor.responsavel:
        gestor_obj = colaborador.setor.responsavel

    gestor_nome = gestor_obj.nome_completo if hasattr(gestor_obj, 'nome_completo') else (str(gestor_obj) if gestor_obj else "-")
    lider_nome = colaborador.lider.nome_completo if getattr(colaborador, 'lider', None) else gestor_nome
    supervisor_nome = colaborador.supervisor.nome_completo if getattr(colaborador, 'supervisor', None) else gestor_nome
    gerente_nome = colaborador.gerente.nome_completo if getattr(colaborador, 'gerente', None) else gestor_nome
    posto_trabalho = getattr(colaborador, 'posto_trabalho', None) or colab_setor
    centro_custo = str(colaborador.centro_custo) if getattr(colaborador, 'centro_custo', None) else "-"

    return {
        'nome': colab_nome,
        'matricula': colab_mat,
        'cargo': colab_cargo,
        'setor': colab_setor,
        'gestor': gestor_nome,
        'lider': lider_nome,
        'supervisor': supervisor_nome,
        'gerente': gerente_nome,
        'posto_trabalho': posto_trabalho,
        'centro_custo': centro_custo,
    }


def gerar_auto_avaliacao_for141_xlsx(planejamento: PlanejamentoTreinamento, colaborador_id: int = None, perguntas_selecionadas: list = None) -> BytesIO:
    """
    Gera o Formulário de Auto-Avaliação de Treinamento Crítico (FOR.141.r02).
    Preserva 100% das formas gráficas, do gráfico radar pentagonal e das caixas de texto oficiais.
    Injeta as 5 perguntas delimitadas nas caixas dos 5 vértices ({{PER_1}} a {{PER_5}})
    e preenche os campos do cabeçalho nas células delimitadas.
    """
    raw_bytes, template_obj = _obter_raw_bytes_template(
        funcao='AUTO_AVALIACAO',
        codigo_busca='141',
        nome_padrao='FOR.141.r02_Auto_Avaliacao.xlsx'
    )

    perguntas = _obter_perguntas_treinamento(planejamento, perguntas_selecionadas=perguntas_selecionadas)
    if hasattr(planejamento, '_mock_procs') and planejamento._mock_procs:
        procs = list(planejamento._mock_procs)
    elif hasattr(planejamento, 'procedimentos'):
        procs = list(planejamento.procedimentos.all())
    else:
        procs = []

    proc_str = ", ".join([f"{p.codigo} - {p.nome}" for p in procs]) if procs else (planejamento.titulo or "-")
    instrutor_nome = (planejamento.instrutor.nome_completo if getattr(planejamento, 'instrutor', None) else "-")
    data_str = planejamento.data_prevista.strftime("%d/%m/%Y") if getattr(planejamento, 'data_prevista', None) else timezone.now().strftime("%d/%m/%Y")

    # Colaborador específico ou lista
    colaborador = None
    if hasattr(planejamento, 'colaboradores'):
        if colaborador_id:
            colaborador = planejamento.colaboradores.select_related('setor', 'lider', 'supervisor', 'gerente').filter(id=colaborador_id).first()
        if not colaborador and planejamento.colaboradores.exists():
            colaborador = planejamento.colaboradores.select_related('setor', 'lider', 'supervisor', 'gerente').first()

    d_colab = _extrair_dados_colaborador_avaliado(colaborador)

    # Garantir que a lista de perguntas tenha 5 posições
    p_padded = list(perguntas)
    while len(p_padded) < 5:
        p_padded.append("")

    substituicoes = {
        "{{TITULO}}": planejamento.titulo or "-",
        "{{PROCEDIMENTO}}": proc_str,
        "{{CODIGO_PROCEDIMENTO}}": procs[0].codigo if procs else "-",
        "{{NOME_PROCEDIMENTO}}": procs[0].nome if procs else (planejamento.titulo or "-"),

        # Tags do Colaborador Avaliado
        "{{COLABORADOR}}": d_colab['nome'],
        "{{NOME_COLABORADOR}}": d_colab['nome'],
        "{{MATRICULA}}": d_colab['matricula'],
        "{{CARGO}}": d_colab['cargo'],
        "{{POSTO_TRABALHO}}": d_colab['posto_trabalho'],
        "{{CENTRO_CUSTO}}": d_colab['centro_custo'],

        # Tags de Setor do Colaborador Avaliado
        "{{SETOR}}": d_colab['setor'],
        "{{SETOR_COLABORADOR}}": d_colab['setor'],
        "{{NOME_SETOR}}": d_colab['setor'],
        "{{DEPARTAMENTO}}": d_colab['setor'],
        "{{AREA}}": d_colab['setor'],
        "{{LABORATORIO}}": d_colab['setor'],

        # Tags de Gestor do Colaborador Avaliado
        "{{GESTOR}}": d_colab['gestor'],
        "{{GESTOR_COLABORADOR}}": d_colab['gestor'],
        "{{NOME_GESTOR}}": d_colab['gestor'],
        "{{LIDER}}": d_colab['lider'],
        "{{SUPERVISOR}}": d_colab['supervisor'],
        "{{GERENTE}}": d_colab['gerente'],
        "{{CHEFIA}}": d_colab['gestor'],
        "{{RESPONSAVEL}}": d_colab['gestor'],

        # Facilitador / Instrutor
        "{{INSTRUTOR}}": instrutor_nome,
        "{{FACILITADOR}}": instrutor_nome,
        "{{DATA}}": data_str,
        "{{DATA_HORA}}": data_str,
        "{{CARGA_HORARIA}}": f"{planejamento.carga_horaria} Minutos" if getattr(planejamento, 'carga_horaria', None) else "-",

        # Perguntas 1 a 5 delimitadas nos vértices do radar
        "{{PER_1}}": p_padded[0],
        "{{PER_2}}": p_padded[1],
        "{{PER_3}}": p_padded[2],
        "{{PER_4}}": p_padded[3],
        "{{PER_5}}": p_padded[4],
        "{{PERGUNTA_1}}": p_padded[0],
        "{{PERGUNTA_2}}": p_padded[1],
        "{{PERGUNTA_3}}": p_padded[2],
        "{{PERGUNTA_4}}": p_padded[3],
        "{{PERGUNTA_5}}": p_padded[4],
    }

    # 1. Se possuímos o arquivo template original, preenchemos de forma 100% LOSSLESS preservando formas e radar
    if raw_bytes is not None:
        cell_updates = {
            'D3': d_colab['nome'],
            'D4': d_colab['setor'],
            'D5': proc_str,
            'D6': data_str,
            'D7': instrutor_nome,
        }
        return _preencher_template_xlsx_preservando_formas(raw_bytes, substituicoes, cell_updates)

    # 2. Fallback nativo openpyxl caso o template não exista
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auto-Avaliação FOR.141"

    font_header_doc = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_section = Font(name="Arial", size=9.5, bold=True, color="FFFFFF")
    font_label = Font(name="Arial", size=8.5, bold=True, color="334155")
    font_val = Font(name="Arial", size=8.5, color="0F172A")
    font_pergunta = Font(name="Arial", size=9, bold=True, color="1E3A8A")

    fill_header_doc = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_section = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_meta = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_pergunta = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    border_box = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Cabeçalho Principal
    ws.merge_cells("A1:G1")
    c_top = ws.cell(row=1, column=1, value="FOR.141.r02 - AUTO-AVALIAÇÃO DE TREINAMENTO CRÍTICO")
    c_top.font = font_header_doc
    c_top.fill = fill_header_doc
    c_top.alignment = align_center
    ws.row_dimensions[1].height = 28

    dados_header = [
        ("Colaborador:", d_colab['nome'], "Matrícula:", d_colab['matricula']),
        ("Cargo / Função:", d_colab['cargo'], "Setor:", d_colab['setor']),
        ("Procedimento / Treinamento:", proc_str, "Data:", data_str),
        ("Instrutor / Facilitador:", instrutor_nome, "Carga Horária:", f"{planejamento.carga_horaria} Minutos" if getattr(planejamento, 'carga_horaria', None) else "-"),
    ]

    for idx, (l1, v1, l2, v2) in enumerate(dados_header, start=2):
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
        ws.merge_cells(start_row=idx, start_column=6, end_row=idx, end_column=7)

        ws.cell(row=idx, column=1, value=l1)
        ws.cell(row=idx, column=2, value=v1)
        ws.cell(row=idx, column=5, value=l2)
        ws.cell(row=idx, column=6, value=v2)

        for col in range(1, 8):
            c = ws.cell(row=idx, column=col)
            c.border = border_box
            if col in [1, 5]:
                c.font = font_label
                c.fill = fill_meta
                c.alignment = align_left
            else:
                c.font = font_val
                c.alignment = align_left
        ws.row_dimensions[idx].height = 18

    ws.row_dimensions[6].height = 6

    # Seção de Perguntas
    ws.merge_cells("A7:G7")
    c_sec = ws.cell(row=7, column=1, value="QUESTIONÁRIO DE AUTOAVALIAÇÃO TÉCNICA (5 PERGUNTAS OBRIGATÓRIAS)")
    c_sec.font = font_section
    c_sec.fill = fill_section
    c_sec.alignment = align_center
    ws.row_dimensions[7].height = 20

    current_row = 8
    for num_p, p_texto in enumerate(perguntas, start=1):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        c_p = ws.cell(row=current_row, column=1, value=f"Questão {num_p}: {p_texto}")
        c_p.font = font_pergunta
        c_p.fill = fill_pergunta
        c_p.alignment = align_left
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = border_box
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=7)
        c_resp = ws.cell(row=current_row, column=1, value="Resposta do Colaborador:\n\n")
        c_resp.font = Font(name="Arial", size=8, italic=True, color="64748B")
        c_resp.alignment = align_top_left
        for r_sub in range(current_row, current_row + 3):
            for col in range(1, 8):
                ws.cell(row=r_sub, column=col).border = border_box
            ws.row_dimensions[r_sub].height = 16
        current_row += 3

    # Bloco de Assinaturas
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)

    c_ass_colab = ws.cell(row=current_row, column=1, value="Assinatura do Colaborador")
    c_ass_colab.font = font_label
    c_ass_colab.alignment = align_center

    c_ass_inst = ws.cell(row=current_row, column=5, value="Assinatura do Instrutor / Avaliador")
    c_ass_inst.font = font_label
    c_ass_inst.alignment = align_center

    larguras_141 = {'A': 16, 'B': 24, 'C': 16, 'D': 16, 'E': 14, 'F': 22, 'G': 18}
    for col_letter, width in larguras_141.items():
        ws.column_dimensions[col_letter].width = width

    raw_output = BytesIO()
    wb.save(raw_output)
    raw_bytes_fb = raw_output.getvalue()
    return _substituir_tags_no_arquivo_zip(raw_bytes_fb, substituicoes)


# ==============================================================================
# 3. FOR.142.r01 - AVALIAÇÃO DE EFICÁCIA DO TREINAMENTO (+30 DIAS)
# ==============================================================================

def gerar_avaliacao_eficacia_for142_xlsx(treinamento_id: int) -> BytesIO:
    """
    Gera o Formulário de Avaliação de Eficácia do Treinamento (FOR.142.r01).
    Preserva 100% das formas, checkboxes e estrutura oficial do template.
    Calcula a Data Devida da Eficácia (Data do Treinamento + 30 dias).
    """
    treinamento = RegistroTreinamento.objects.select_related(
        'colaborador', 'procedimento', 'colaborador__setor', 'colaborador__lider',
        'colaborador__supervisor', 'colaborador__gerente'
    ).get(id=treinamento_id)

    raw_bytes, template_obj = _obter_raw_bytes_template(
        funcao='AVALIACAO_EFICACIA',
        codigo_busca='142',
        nome_padrao='FOR.142.r01_Avaliacao_de_Eficacia_do_Treinamento.xlsx'
    )

    colab = treinamento.colaborador
    proc = treinamento.procedimento
    d_colab = _extrair_dados_colaborador_avaliado(colab)

    data_treinamento = treinamento.data_treinamento or timezone.now().date()
    data_eficacia_calculada = data_treinamento + timedelta(days=30)
    data_treinamento_str = data_treinamento.strftime("%d/%m/%Y")
    data_eficacia_str = data_eficacia_calculada.strftime("%d/%m/%Y")
    data_avaliacao_str = treinamento.avaliacao_eficacia_data.strftime("%d/%m/%Y") if treinamento.avaliacao_eficacia_data else "-"

    status_str = treinamento.avaliacao_eficacia_status or "PENDENTE"
    status_map = {
        'EFICAZ': 'Eficaz',
        'INEFICAZ': 'Ineficaz',
        'NAO_APLICA': 'Não se Aplica',
        'PENDENTE': 'Pendente'
    }
    status_display = status_map.get(status_str, status_str)

    substituicoes = {
        # Tags do Colaborador Avaliado
        "{{COLABORADOR}}": d_colab['nome'],
        "{{NOME_COLABORADOR}}": d_colab['nome'],
        "{{MATRICULA}}": d_colab['matricula'],
        "{{CARGO}}": d_colab['cargo'],
        "{{POSTO_TRABALHO}}": d_colab['posto_trabalho'],
        "{{CENTRO_CUSTO}}": d_colab['centro_custo'],

        # Tags de Setor do Colaborador Avaliado
        "{{SETOR}}": d_colab['setor'],
        "{{SETOR_COLABORADOR}}": d_colab['setor'],
        "{{NOME_SETOR}}": d_colab['setor'],
        "{{DEPARTAMENTO}}": d_colab['setor'],
        "{{AREA}}": d_colab['setor'],
        "{{LABORATORIO}}": d_colab['setor'],

        # Tags de Gestor do Colaborador Avaliado
        "{{GESTOR}}": d_colab['gestor'],
        "{{GESTOR_COLABORADOR}}": d_colab['gestor'],
        "{{NOME_GESTOR}}": d_colab['gestor'],
        "{{LIDER}}": d_colab['lider'],
        "{{SUPERVISOR}}": d_colab['supervisor'],
        "{{GERENTE}}": d_colab['gerente'],
        "{{CHEFIA}}": d_colab['gestor'],
        "{{RESPONSAVEL}}": d_colab['gestor'],
        "{{AVALIADOR}}": d_colab['gestor'],

        # Procedimento e Dados do Treinamento
        "{{PROCEDIMENTO}}": f"{proc.codigo} - {proc.nome}" if proc else "-",
        "{{CODIGO_PROCEDIMENTO}}": proc.codigo if proc else "-",
        "{{NOME_PROCEDIMENTO}}": proc.nome if proc else "-",
        "{{DATA_TREINAMENTO}}": data_treinamento_str,
        "{{DATA_EFICACIA_CALCULADA}}": data_eficacia_str,
        "{{DATA_ELEGIBILIDADE}}": data_eficacia_str,
        "{{DATA_AVALIACAO}}": data_avaliacao_str,
        "{{STATUS_EFICACIA}}": status_display,
        "{{OBSERVACOES}}": treinamento.resultado_avaliacao or "",
        "{{JUSTIFICATIVA}}": treinamento.resultado_avaliacao or "",
        "{{EVIDENCIAS}}": treinamento.resultado_avaliacao or "",
        "{{CHK_EFICAZ}}": "●" if status_str == 'EFICAZ' else "○",
        "{{CHK_INEFICAZ}}": "●" if status_str == 'INEFICAZ' else "○",
        "{{CHK_NAO_APLICA}}": "●" if status_str == 'NAO_APLICA' else "○",
    }

    # 1. Se possuímos o arquivo template original, preenchemos de forma 100% LOSSLESS preservando formas e layout
    if raw_bytes is not None:
        cell_updates = {
            'C4': f"APLICAR APÓS {data_eficacia_str} (CARÊNCIA DE 30 DIAS CALCULADA)",
            'C5': f"{proc.codigo} - {proc.nome}" if proc else "-",
            'W5': data_treinamento_str,
            'C6': d_colab['nome'],
            'W6': d_colab['setor'],
            'C7': d_colab['gestor'],
            'B30': treinamento.resultado_avaliacao or "",
            'P38': "[ X ]" if status_str == 'EFICAZ' else "[   ]",
            'V38': "[ X ]" if status_str == 'INEFICAZ' else "[   ]",
            'B42': f"Colaborador: {d_colab['nome']}",
            'Z42': f"Data: {data_treinamento_str}",
            'B43': f"Gestor: {d_colab['gestor']}",
            'Z43': f"Data: {data_avaliacao_str}",
        }
        return _preencher_template_xlsx_preservando_formas(raw_bytes, substituicoes, cell_updates)

    # 2. Fallback nativo openpyxl caso o template não exista
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Eficácia FOR.142"

        font_header_doc = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_section = Font(name="Arial", size=9.5, bold=True, color="FFFFFF")
        font_label = Font(name="Arial", size=8.5, bold=True, color="334155")
        font_val = Font(name="Arial", size=8.5, color="0F172A")
        font_destaque = Font(name="Arial", size=9, bold=True, color="1E3A8A")

        fill_header_doc = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_section = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        fill_meta = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_calculada = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amarelo suave

        border_box = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 1. Cabeçalho Principal
        ws.merge_cells("A1:G1")
        c_top = ws.cell(row=1, column=1, value="FOR.142.r01 - AVALIAÇÃO DE EFICÁCIA DO TREINAMENTO")
        c_top.font = font_header_doc
        c_top.fill = fill_header_doc
        c_top.alignment = align_center
        ws.row_dimensions[1].height = 28

        # 2. Dados do Colaborador e Treinamento
        dados_header = [
            ("Colaborador:", d_colab['nome'], "Matrícula:", d_colab['matricula']),
            ("Cargo / Função:", d_colab['cargo'], "Setor:", d_colab['setor']),
            ("Responsável / Gestor:", d_colab['gestor'], "Procedimento:", f"{proc.codigo} - {proc.nome}" if proc else "-"),
            ("Data do Treinamento:", data_treinamento_str, "Data Devida Eficácia (+30d):", data_eficacia_str),
        ]

        for idx, (l1, v1, l2, v2) in enumerate(dados_header, start=2):
            ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
            ws.merge_cells(start_row=idx, start_column=6, end_row=idx, end_column=7)

            ws.cell(row=idx, column=1, value=l1)
            ws.cell(row=idx, column=2, value=v1)
            ws.cell(row=idx, column=5, value=l2)
            c_v2 = ws.cell(row=idx, column=6, value=v2)

            for col in range(1, 8):
                c = ws.cell(row=idx, column=col)
                c.border = border_box
                if col in [1, 5]:
                    c.font = font_label
                    c.fill = fill_meta
                    c.alignment = align_left
                else:
                    c.font = font_val
                    c.alignment = align_left
            
            # Destacar a Data Devida Calculada
            if idx == 5:
                c_v2.fill = fill_calculada
                c_v2.font = Font(name="Arial", size=9, bold=True, color="92400E")

            ws.row_dimensions[idx].height = 18

        ws.row_dimensions[6].height = 6

        # 3. Seção de Avaliação
        ws.merge_cells("A7:G7")
        c_sec = ws.cell(row=7, column=1, value="PARECER TÉCNICO E RESULTADO DA EFICÁCIA (APÓS 30 DIAS)")
        c_sec.font = font_section
        c_sec.fill = fill_section
        c_sec.alignment = align_center
        ws.row_dimensions[7].height = 20

        # Linha de Status de Eficácia
        ws.cell(row=8, column=1, value="Resultado da Eficácia:")
        ws.cell(row=8, column=1).font = font_label
        ws.cell(row=8, column=1).fill = fill_meta
        ws.cell(row=8, column=1).border = border_box

        ws.merge_cells("B8:D8")
        chk_str = f"{'●' if status_str == 'EFICAZ' else '○'} Eficaz     {'●' if status_str == 'INEFICAZ' else '○'} Ineficaz     {'●' if status_str == 'NAO_APLICA' else '○'} Não se Aplica"
        c_chk = ws.cell(row=8, column=2, value=chk_str)
        c_chk.font = font_destaque
        c_chk.alignment = align_left
        for col in range(2, 5):
            ws.cell(row=8, column=col).border = border_box

        ws.cell(row=8, column=5, value="Data da Avaliação:")
        ws.cell(row=8, column=5).font = font_label
        ws.cell(row=8, column=5).fill = fill_meta
        ws.cell(row=8, column=5).border = border_box

        ws.merge_cells("F8:G8")
        c_dt_av = ws.cell(row=8, column=6, value=data_avaliacao_str)
        c_dt_av.font = font_val
        c_dt_av.alignment = align_left
        for col in range(6, 8):
            ws.cell(row=8, column=col).border = border_box
        ws.row_dimensions[8].height = 20

        # Evidências / Justificativas
        ws.merge_cells("A9:G9")
        c_ev_title = ws.cell(row=9, column=1, value="Evidências Objetivas Observadas / Justificativa:")
        c_ev_title.font = font_label
        c_ev_title.fill = fill_meta
        c_ev_title.alignment = align_left
        for col in range(1, 8):
            ws.cell(row=9, column=col).border = border_box
        ws.row_dimensions[9].height = 18

        ws.merge_cells("A10:G15")
        c_ev_body = ws.cell(row=10, column=1, value=treinamento.resultado_avaliacao or "Nenhuma observação ou evidência registrada.")
        c_ev_body.font = font_val
        c_ev_body.alignment = align_top_left
        for r_ev in range(10, 16):
            for col in range(1, 8):
                ws.cell(row=r_ev, column=col).border = border_box
            ws.row_dimensions[r_ev].height = 16

        # Bloco de Assinaturas
        ws.row_dimensions[16].height = 12

        ws.merge_cells("A17:C17")
        ws.merge_cells("E17:G17")

        c_ass_col = ws.cell(row=17, column=1, value="Assinatura do Colaborador Avaliado")
        c_ass_col.font = font_label
        c_ass_col.alignment = align_center

        c_ass_lid = ws.cell(row=17, column=5, value="Assinatura do Líder / Avaliador Responsável")
        c_ass_lid.font = font_label
        c_ass_lid.alignment = align_center

        larguras_142 = {'A': 18, 'B': 22, 'C': 16, 'D': 16, 'E': 20, 'F': 20, 'G': 16}
        for col_letter, width in larguras_142.items():
            ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
