# -*- coding: utf-8 -*-
"""
Utilidades para importação em massa de matrizes, disciplinas e colaboradores
"""

import csv
import io
from decimal import Decimal
from typing import Dict, List, Tuple
import openpyxl
from django.db import transaction

from procedures.models import MatrizHabilidade, Disciplina, ColaboradorMatrizHabilidade
from rh.models import Colaborador


class ImportadorMatrizHabilidade:
    """Classe para processar importações de matrizes, disciplinas e colaboradores."""
    
    def __init__(self):
        self.erros = []
        self.avisos = []
        self.resumo = {
            'matrizes_criadas': 0,
            'matrizes_atualizadas': 0,
            'disciplinas_criadas': 0,
            'disciplinas_atualizadas': 0,
            'colaboradores_associados': 0,
            'colaboradores_nao_encontrados': 0,
            'avaliacoes_criadas': 0,
            'avaliacoes_atualizadas': 0,
        }
    
    def adicionar_erro(self, linha: int, mensagem: str):
        """Adiciona erro à lista de erros."""
        self.erros.append(f"Linha {linha}: {mensagem}")
    
    def adicionar_aviso(self, linha: int, mensagem: str):
        """Adiciona aviso à lista de avisos."""
        self.avisos.append(f"Linha {linha}: {mensagem}")
    
    def processar_csv(self, arquivo_csv) -> bool:
        """
        Processa arquivo CSV com formato:
        matriz_codigo | matriz_nome | disciplina_codigo | disciplina_nome | 
        disciplina_descricao | disciplina_prioridade | disciplina_obrigatoriedade | 
        colaborador_matricula | colaborador_nome | colaborador_email
        """
        try:
            conteudo = arquivo_csv.read()
            
            # Detectar encoding
            try:
                texto = conteudo.decode('utf-8')
            except UnicodeDecodeError:
                texto = conteudo.decode('latin-1')
            
            # Processar como CSV
            reader = csv.DictReader(io.StringIO(texto), delimiter='|')
            
            if not reader.fieldnames:
                self.adicionar_erro(0, "Arquivo CSV vazio ou inválido")
                return False
            
            linha_num = 1
            with transaction.atomic():
                for row in reader:
                    linha_num += 1
                    if not any(row.values()):
                        continue  # Pular linhas vazias
                    
                    if not self._processar_linha(row, linha_num):
                        continue
            
            return len(self.erros) == 0
        
        except Exception as e:
            self.adicionar_erro(0, f"Erro ao processar arquivo: {str(e)}")
            return False
    
    def processar_excel(self, arquivo_excel) -> bool:
        """
        Processa arquivo Excel com colunas:
        - Matriz Código | Matriz Nome | Disciplina Código | Disciplina Nome | 
        - Disciplina Descrição | Disciplina Prioridade | Disciplina Obrigatoriedade | 
        - Colaborador Matrícula | Colaborador Nome | Colaborador Email
        """
        try:
            wb = openpyxl.load_workbook(arquivo_excel)
            ws = wb.active
            
            if not ws:
                self.adicionar_erro(0, "Arquivo Excel vazio")
                return False
            
            # Extrair headers da primeira linha
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if not any(headers):
                self.adicionar_erro(0, "Cabeçalho não encontrado")
                return False
            
            linha_num = 1
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue  # Pular linhas vazias
                    
                    # Converter para dicionário
                    row_dict = {}
                    for header_idx, header in enumerate(headers):
                        if header_idx < len(row):
                            row_dict[header] = row[header_idx]
                    
                    if not self._processar_linha(row_dict, row_idx):
                        continue
            
            return len(self.erros) == 0
        
        except Exception as e:
            self.adicionar_erro(0, f"Erro ao processar arquivo Excel: {str(e)}")
            return False
    
    def _processar_linha(self, row: Dict, linha_num: int) -> bool:
        """Processa uma linha de dados."""
        try:
            # Extrair dados da linha
            matriz_codigo = str(row.get('Matriz Código') or '').strip()
            matriz_nome = str(row.get('Matriz Nome') or '').strip()
            disc_codigo = str(row.get('Disciplina Código') or '').strip()
            disc_nome = str(row.get('Disciplina Nome') or '').strip()
            colab_matricula = str(row.get('Colaborador Matrícula') or '').strip()
            colab_nome = str(row.get('Colaborador Nome') or '').strip()
            nivel_str = str(row.get('Nível de Competência') or '').strip()
            observacoes = str(row.get('Observações') or '').strip()
            
            # Validações básicas
            if not matriz_codigo or not matriz_nome:
                self.adicionar_erro(linha_num, "Matriz código e nome são obrigatórios")
                return False
            
            if not disc_nome:
                self.adicionar_erro(linha_num, "Disciplina nome é obrigatória")
                return False
            
            # Criar ou atualizar matriz
            matriz, created = MatrizHabilidade.objects.get_or_create(
                codigo=matriz_codigo,
                defaults={'nome': matriz_nome, 'ativo': True}
            )
            
            if created:
                self.resumo['matrizes_criadas'] += 1
            else:
                if matriz.nome != matriz_nome:
                    matriz.nome = matriz_nome
                    matriz.save()
                    self.resumo['matrizes_atualizadas'] += 1
            
            # Criar ou atualizar disciplina
            disciplina, created = Disciplina.objects.get_or_create(
                matriz=matriz,
                nome=disc_nome,
                defaults={
                    'codigo': disc_codigo or self._gerar_codigo_disciplina(),
                    'ativo': True
                }
            )
            
            if created:
                self.resumo['disciplinas_criadas'] += 1
            else:
                self.resumo['disciplinas_atualizadas'] += 1
            
            # Associar colaborador se fornecido
            if colab_matricula or colab_nome:
                self._associar_colaborador(
                    matriz, disciplina, colab_matricula, colab_nome, nivel_str, observacoes, linha_num
                )
            
            return True
        
        except Exception as e:
            self.adicionar_erro(linha_num, f"Erro ao processar linha: {str(e)}")
            return False
    
    def _associar_colaborador(
        self, 
        matriz: MatrizHabilidade, 
        disciplina: Disciplina,
        matricula: str, 
        nome: str,
        nivel_str: str,
        observacoes: str,
        linha_num: int
    ) -> bool:
        """Associa um colaborador à matriz e cria/atualiza a avaliação."""
        try:
            colaborador = None
            
            # Tentar encontrar por matrícula
            if matricula:
                try:
                    colaborador = Colaborador.objects.get(matricula=matricula)
                except Colaborador.DoesNotExist:
                    pass
            
            # Tentar encontrar por nome
            if not colaborador and nome:
                try:
                    colaborador = Colaborador.objects.get(nome_completo__iexact=nome)
                except (Colaborador.DoesNotExist, Colaborador.MultipleObjectsReturned):
                    pass
            
            if not colaborador:
                self.adicionar_aviso(
                    linha_num, 
                    f"Colaborador não encontrado: {matricula or nome}"
                )
                self.resumo['colaboradores_nao_encontrados'] += 1
                return False
            
            # Criar associação se não existir
            _, created = ColaboradorMatrizHabilidade.objects.get_or_create(
                colaborador=colaborador,
                matriz=matriz,
                defaults={'ativo': True}
            )
            
            if created:
                self.resumo['colaboradores_associados'] += 1
            
            # Processar nível de competência se fornecido
            if nivel_str:
                try:
                    # Mapear "n/a", "N/A" para -1
                    nivel_str_lower = nivel_str.lower().strip()
                    if nivel_str_lower in ['n/a', 'na', '-1']:
                        nivel = -1
                    else:
                        nivel = int(nivel_str)
                    
                    # Validar nível
                    niveis_validos = [-1, 0, 1, 2, 3]
                    if nivel not in niveis_validos:
                        self.adicionar_aviso(
                            linha_num,
                            f"Nível de competência inválido: {nivel}. Valores válidos: {niveis_validos} ou 'N/A'"
                        )
                        return False
                    
                    # Criar ou atualizar avaliação
                    from datetime import date
                    from procedures.models import AvaliacaoHabilidade
                    
                    avaliacao, created = AvaliacaoHabilidade.objects.update_or_create(
                        colaborador=colaborador,
                        disciplina=disciplina,
                        matriz=matriz,
                        defaults={
                            'nivel': nivel,
                            'observacoes': observacoes or None,
                            'data_avaliacao': date.today()
                        }
                    )
                    
                    if created:
                        self.resumo['avaliacoes_criadas'] += 1
                    else:
                        self.resumo['avaliacoes_atualizadas'] += 1
                
                except ValueError:
                    self.adicionar_aviso(
                        linha_num,
                        f"Nível de competência não é um número válido: {nivel_str}. Use: -1, 0, 1, 2, 3 ou 'N/A'"
                    )
                    return False
            
            return True
        
        except Exception as e:
            self.adicionar_aviso(linha_num, f"Erro ao associar colaborador: {str(e)}")
            self.resumo['colaboradores_nao_encontrados'] += 1
            return False
    
    def _gerar_codigo_disciplina(self) -> str:
        """Gera código único para disciplina."""
        ultimo = Disciplina.objects.all().order_by('id').last()
        if ultimo:
            try:
                ultimo_num = int(ultimo.codigo.replace('DISC', ''))
                return f'DISC{str(ultimo_num + 1).zfill(3)}'
            except:
                return 'DISC001'
        return 'DISC001'
    
    def obter_resumo(self) -> Dict:
        """Retorna resumo da importação."""
        return {
            'resumo': self.resumo,
            'erros': self.erros,
            'avisos': self.avisos,
            'sucesso': len(self.erros) == 0
        }


def validar_arquivo_importacao(arquivo) -> Tuple[bool, str]:
    """
    Valida se o arquivo é CSV ou Excel.
    Retorna: (valido, mensagem_erro)
    """
    try:
        nome = arquivo.name.lower()
        
        if nome.endswith('.csv'):
            return True, ""
        elif nome.endswith(('.xls', '.xlsx')):
            return True, ""
        else:
            return False, "Arquivo deve ser CSV ou Excel (.xlsx, .xls)"
    
    except Exception as e:
        return False, f"Erro ao validar arquivo: {str(e)}"


def gerar_template_csv() -> str:
    """Gera conteúdo de template CSV para importação."""
    template = """Matriz Código|Matriz Nome|Disciplina Código|Disciplina Nome|Colaborador Matrícula|Colaborador Nome|Nível de Competência|Observações
MAT001|Operação|DISC001|Segurança|MAT001|João Silva|2|Necessita aprimoramento
MAT001|Operação|DISC002|Qualidade|MAT002|Maria Santos|3|Em dia com treinamentos
MAT002|Manutenção|DISC003|Manutenção Preventiva|MAT003|Pedro Costa|N/A|Não se aplica para esta disciplina"""
    return template


def gerar_template_excel() -> str:
    """Retorna instruções para criar template Excel."""
    return """
Colunas esperadas no Excel (primeira linha é cabeçalho):
1. Matriz Código - Código único da matriz (ex: MAT001)
2. Matriz Nome - Nome descritivo da matriz (ex: Operação)
3. Disciplina Código - Código da disciplina (ex: DISC001)
4. Disciplina Nome - Nome da disciplina (ex: Segurança)
5. Colaborador Matrícula - Matrícula do colaborador
6. Colaborador Nome - Nome completo do colaborador
7. Nível de Competência - Nível de competência (-1, 0, 1, 2, 3, ou "N/A")
    -1 ou N/A: Não se Aplica
    0: Há Intenção de Treinar
    1: Colaborador em Treinamento
    2: Treinado
    3: Treinado na Plataforma LOFT
8. Observações - Observações ou notas sobre a avaliação

Cada linha representa uma associação entre matriz, disciplina e colaborador.
Matrizes e disciplinas duplicadas serão atualizadas, não recriadas.
Avaliações duplicadas serão atualizadas com o novo nível e observações.
"""
