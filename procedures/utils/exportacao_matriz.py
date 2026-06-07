"""
Módulo para exportação de Matrizes de Habilidades em CSV ou Excel
"""
import csv
import io
from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from procedures.models import MatrizHabilidade, Disciplina, ColaboradorMatrizHabilidade


class ExportadorMatrizHabilidade:
    """
    Responsável por exportar dados de Matrizes em CSV ou Excel
    """

    def __init__(self):
        """Inicializa o exportador"""
        self.data_export = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def exportar_csv(self) -> Tuple[io.StringIO, str]:
        """
        Exporta matrizes em formato CSV

        Returns:
            Tuple contendo (StringIO com dados, nome do arquivo)
        """
        output = io.StringIO()
        writer = csv.writer(output, delimiter='|')

        # Header
        writer.writerow([
            'Matriz',
            'Disciplina',
            'Colaborador',
            'Turno',
            'Nota',
            'Data'
        ])

        from procedures.models import AvaliacaoHabilidade, ColaboradorMatrizHabilidade
        from core.models import TURNOS_CHOICES

        turno_dict = dict(TURNOS_CHOICES)

        # Pre-fetch all evaluations to lookup by (matriz_id, disciplina_id, colaborador_id)
        avaliacoes = AvaliacaoHabilidade.objects.select_related('colaborador', 'disciplina', 'matriz').all()
        av_map = {}
        for av in avaliacoes:
            key = (av.matriz_id, av.disciplina_id, av.colaborador_id)
            av_map[key] = av

        # Dados - Matriz -> Disciplina -> Colaborador (via ColaboradorMatrizHabilidade)
        for matriz in MatrizHabilidade.objects.prefetch_related(
            'disciplinas_matriz'
        ).all():
            disciplinas = matriz.disciplinas_matriz.all()

            if not disciplinas.exists():
                # Matriz sem disciplinas
                writer.writerow([
                    f"{matriz.codigo} - {matriz.nome}",
                    '', '', '', '', ''
                ])
            else:
                # Obter todos os colaboradores associados a esta matriz
                colaboradores_assoc = ColaboradorMatrizHabilidade.objects.filter(
                    matriz=matriz,
                    ativo=True
                ).select_related('colaborador').order_by('colaborador__nome_completo')

                if not colaboradores_assoc.exists():
                    # Disciplinas existem, mas nenhum colaborador associado
                    for disciplina in disciplinas:
                        writer.writerow([
                            f"{matriz.codigo} - {matriz.nome}",
                            f"{disciplina.codigo} - {disciplina.nome}",
                            '', '', '', ''
                        ])
                else:
                    for colaborador_assoc in colaboradores_assoc:
                        colaborador = colaborador_assoc.colaborador
                        turno_label = turno_dict.get(colaborador.turno, colaborador.turno)

                        for disciplina in disciplinas:
                            key = (matriz.id, disciplina.id, colaborador.id)
                            av = av_map.get(key)

                            nota = av.get_nivel_display() if av else ''
                            data_avaliacao = av.data_avaliacao.strftime("%d/%m/%Y") if (av and av.data_avaliacao) else ''

                            writer.writerow([
                                f"{matriz.codigo} - {matriz.nome}",
                                f"{disciplina.codigo} - {disciplina.nome}",
                                colaborador.nome_completo,
                                turno_label,
                                nota,
                                data_avaliacao
                            ])

        filename = f"exportacao_matrizes_{self.timestamp}.csv"
        return output, filename

    def exportar_excel(self) -> Tuple[bytes, str]:
        """
        Exporta matrizes em formato Excel

        Returns:
            Tuple contendo (bytes do arquivo, nome do arquivo)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Matrizes"

        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Matriz',
            'Disciplina',
            'Colaborador',
            'Turno',
            'Nota',
            'Data'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Dados
        from procedures.models import AvaliacaoHabilidade, ColaboradorMatrizHabilidade
        from core.models import TURNOS_CHOICES

        turno_dict = dict(TURNOS_CHOICES)

        # Pre-fetch all evaluations to lookup by (matriz_id, disciplina_id, colaborador_id)
        avaliacoes = AvaliacaoHabilidade.objects.select_related('colaborador', 'disciplina', 'matriz').all()
        av_map = {}
        for av in avaliacoes:
            key = (av.matriz_id, av.disciplina_id, av.colaborador_id)
            av_map[key] = av

        row_num = 2
        for matriz in MatrizHabilidade.objects.prefetch_related(
            'disciplinas_matriz'
        ).all():
            disciplinas = matriz.disciplinas_matriz.all()

            if not disciplinas.exists():
                # Matriz sem disciplinas
                ws.cell(row=row_num, column=1).value = f"{matriz.codigo} - {matriz.nome}"
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = border
                row_num += 1
            else:
                # Obter todos os colaboradores associados a esta matriz
                colaboradores_assoc = ColaboradorMatrizHabilidade.objects.filter(
                    matriz=matriz,
                    ativo=True
                ).select_related('colaborador').order_by('colaborador__nome_completo')

                if not colaboradores_assoc.exists():
                    # Disciplinas existem, mas nenhum colaborador associado
                    for disciplina in disciplinas:
                        ws.cell(row=row_num, column=1).value = f"{matriz.codigo} - {matriz.nome}"
                        ws.cell(row=row_num, column=2).value = f"{disciplina.codigo} - {disciplina.nome}"
                        for col in range(1, 7):
                            ws.cell(row=row_num, column=col).border = border
                        row_num += 1
                else:
                    for colaborador_assoc in colaboradores_assoc:
                        colaborador = colaborador_assoc.colaborador
                        turno_label = turno_dict.get(colaborador.turno, colaborador.turno)

                        for disciplina in disciplinas:
                            key = (matriz.id, disciplina.id, colaborador.id)
                            av = av_map.get(key)

                            nota = av.get_nivel_display() if av else ''
                            data_avaliacao = av.data_avaliacao.strftime("%d/%m/%Y") if (av and av.data_avaliacao) else ''

                            ws.cell(row=row_num, column=1).value = f"{matriz.codigo} - {matriz.nome}"
                            ws.cell(row=row_num, column=2).value = f"{disciplina.codigo} - {disciplina.nome}"
                            ws.cell(row=row_num, column=3).value = colaborador.nome_completo
                            ws.cell(row=row_num, column=4).value = turno_label
                            ws.cell(row=row_num, column=5).value = nota
                            ws.cell(row=row_num, column=6).value = data_avaliacao

                            # Aplicar border a todas as 6 células
                            for col in range(1, 7):
                                ws.cell(row=row_num, column=col).border = border

                            row_num += 1

        # Ajustar largura das colunas
        column_widths = [30, 30, 30, 15, 25, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Congelar primeira linha
        ws.freeze_panes = "A2"

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"exportacao_matrizes_{self.timestamp}.xlsx"
        return output.getvalue(), filename

    def gerar_relatorio_exportacao(self) -> Dict:
        """
        Gera relatório de dados para exportação

        Returns:
            Dict com estatísticas
        """
        matrizes = MatrizHabilidade.objects.all()
        disciplinas = Disciplina.objects.all()
        colaboradores = ColaboradorMatrizHabilidade.objects.all()

        return {
            'total_matrizes': matrizes.count(),
            'total_disciplinas': disciplinas.count(),
            'total_associacoes': colaboradores.count(),
            'data_export': self.data_export,
            'status': 'Pronto para exportar'
        }
