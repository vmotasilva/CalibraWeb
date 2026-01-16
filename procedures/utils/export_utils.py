"""
Utilitários de exportação para Excel
Responsável por gerar arquivos Excel com dados de planejamentos
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class PlanejamentoExcelExporter:
    """Exporta planejamentos de treinamento para Excel"""
    
    HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.ws = self.workbook.active
        self.ws.title = "Planejamentos"
    
    def export_lista_planejamentos(self, planejamentos):
        """
        Exporta lista de planejamentos para Excel
        
        Args:
            planejamentos: QuerySet de PlanejamentoTreinamento
        
        Returns:
            HttpResponse com arquivo Excel
        """
        # Configurar cabeçalhos
        headers = [
            "ID",
            "Título",
            "Status",
            "Origem",
            "Data Prevista",
            "Data Realizada",
            "Instrutor",
            "Carga Horária",
            "Procedimentos",
            "Colaboradores",
            "Local",
            "Observações"
        ]
        
        # Adicionar cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER
        
        # Adicionar dados
        for row_num, planejamento in enumerate(planejamentos, 2):
            dados = [
                planejamento.id,
                planejamento.titulo,
                planejamento.get_status_display() if hasattr(planejamento, 'get_status_display') else planejamento.status,
                planejamento.get_origem_display() if hasattr(planejamento, 'get_origem_display') else planejamento.origem,
                planejamento.data_prevista.strftime("%d/%m/%Y") if planejamento.data_prevista else "",
                planejamento.data_realizada.strftime("%d/%m/%Y") if planejamento.data_realizada else "",
                planejamento.instrutor.nome_completo if planejamento.instrutor else "",
                f"{planejamento.carga_horaria} min" if planejamento.carga_horaria else "",
                ", ".join([p.codigo for p in planejamento.procedimentos.all()]),
                ", ".join([c.nome_completo for c in planejamento.colaboradores.all()]),
                planejamento.local or "",
                planejamento.observacoes or ""
            ]
            
            for col_num, valor in enumerate(dados, 1):
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = valor
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = self.BORDER
        
        # Ajustar largura das colunas
        self._auto_adjust_columns()
        
        # Congelar a primeira linha
        self.ws.freeze_panes = "A2"
        
        return self._generate_response("planejamentos_lista.xlsx")
    
    def export_detalhe_planejamento(self, planejamento):
        """
        Exporta detalhes de um planejamento específico para Excel
        
        Args:
            planejamento: Instância de PlanejamentoTreinamento
        
        Returns:
            HttpResponse com arquivo Excel
        """
        # ===== ABA 1: Informações Gerais =====
        ws_info = self.ws
        ws_info.title = "Informações"
        
        # Cabeçalho
        self._adicionar_titulo(ws_info, planejamento.titulo, 1)
        
        row = 3
        # Bloco de informações principais
        dados_info = [
            ("ID", planejamento.id),
            ("Título", planejamento.titulo),
            ("Status", self._get_status_display(planejamento.status)),
            ("Origem", self._get_origem_display(planejamento.origem)),
            ("Data Prevista", planejamento.data_prevista.strftime("%d/%m/%Y") if planejamento.data_prevista else ""),
            ("Hora Prevista", planejamento.horario_previsto.strftime("%H:%M") if planejamento.horario_previsto else ""),
            ("Data Realizada", planejamento.data_realizada.strftime("%d/%m/%Y") if planejamento.data_realizada else ""),
            ("Instrutor", planejamento.instrutor.nome_completo if planejamento.instrutor else "Não definido"),
            ("Local", planejamento.local or ""),
            ("Carga Horária", f"{planejamento.carga_horaria} minutos" if planejamento.carga_horaria else ""),
            ("Observações", planejamento.observacoes or ""),
        ]
        
        for label, valor in dados_info:
            self._adicionar_linha_info(ws_info, row, label, valor)
            row += 1
        
        # ===== ABA 2: Procedimentos =====
        ws_proc = self.workbook.create_sheet("Procedimentos")
        self._adicionar_procedimentos(ws_proc, planejamento)
        
        # ===== ABA 3: Colaboradores =====
        ws_colab = self.workbook.create_sheet("Colaboradores")
        self._adicionar_colaboradores(ws_colab, planejamento)
        
        # ===== ABA 4: Registros de Treinamento =====
        ws_registros = self.workbook.create_sheet("Registros de Treinamento")
        self._adicionar_registros_treinamento(ws_registros, planejamento)
        
        # Ajustar largura de todas as abas
        for ws in self.workbook.sheetnames:
            self._auto_adjust_columns(self.workbook[ws])
        
        return self._generate_response(f"planejamento_{planejamento.id}.xlsx")
    
    def _adicionar_titulo(self, ws, titulo, row):
        """Adiciona um título formatado na planilha"""
        cell = ws.cell(row=row, column=1)
        cell.value = f"📋 {titulo}"
        cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f"A{row}:L{row}")
        ws.row_dimensions[row].height = 25
    
    def _adicionar_linha_info(self, ws, row, label, valor):
        """Adiciona uma linha de informação formatada"""
        # Label
        cell_label = ws.cell(row=row, column=1)
        cell_label.value = label
        cell_label.font = Font(bold=True, size=10)
        cell_label.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        cell_label.border = self.BORDER
        
        # Valor
        cell_valor = ws.cell(row=row, column=2)
        cell_valor.value = valor
        cell_valor.alignment = Alignment(wrap_text=True, vertical='top')
        cell_valor.border = self.BORDER
        
        ws.merge_cells(f"B{row}:L{row}")
    
    def _adicionar_procedimentos(self, ws, planejamento):
        """Adiciona lista de procedimentos em uma aba separada"""
        self._adicionar_titulo(ws, f"Procedimentos - {planejamento.titulo}", 1)
        
        # Cabeçalhos
        headers = ["Código", "Nome", "Descrição", "Disciplina"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        # Dados
        for row_num, procedimento in enumerate(planejamento.procedimentos.all(), 4):
            dados = [
                procedimento.codigo,
                procedimento.nome,
                procedimento.descricao or "",
                ", ".join([d.nome for d in procedimento.disciplinas.all()]) or ""
            ]
            
            for col_num, valor in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.BORDER
    
    def _adicionar_colaboradores(self, ws, planejamento):
        """Adiciona lista de colaboradores em uma aba separada"""
        self._adicionar_titulo(ws, f"Colaboradores - {planejamento.titulo}", 1)
        
        # Cabeçalhos
        headers = ["Nome", "Matrícula", "Cargo", "Setor", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        # Dados
        for row_num, colaborador in enumerate(planejamento.colaboradores.all(), 4):
            dados = [
                colaborador.nome_completo,
                colaborador.matricula or "",
                colaborador.cargo or "",
                colaborador.setor or "",
                "Ativo" if colaborador.is_active else "Inativo"
            ]
            
            for col_num, valor in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.BORDER
    
    def _adicionar_registros_treinamento(self, ws, planejamento):
        """Adiciona registros de treinamento realizados"""
        from procedures.models import RegistroTreinamento
        
        self._adicionar_titulo(ws, f"Registros de Treinamento - {planejamento.titulo}", 1)
        
        # Cabeçalhos
        headers = ["Colaborador", "Procedimento", "Data", "Hora", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        # Buscar registros
        registros = RegistroTreinamento.objects.filter(
            procedimento__in=planejamento.procedimentos.all(),
            colaborador__in=planejamento.colaboradores.all()
        ).select_related('colaborador', 'procedimento').order_by('-data_treinamento')
        
        # Dados
        for row_num, registro in enumerate(registros, 4):
            dados = [
                registro.colaborador.nome_completo,
                registro.procedimento.codigo,
                registro.data_treinamento.strftime("%d/%m/%Y") if registro.data_treinamento else "",
                registro.data_treinamento.strftime("%H:%M") if registro.data_treinamento else "",
                "Concluído" if registro.concluido else "Pendente"
            ]
            
            for col_num, valor in enumerate(dados, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.BORDER
        
        if not registros.exists():
            cell = ws.cell(row=4, column=1)
            cell.value = "Nenhum registro de treinamento encontrado"
            ws.merge_cells("A4:E4")
    
    def _auto_adjust_columns(self, ws=None):
        """Ajusta automaticamente a largura das colunas"""
        if ws is None:
            ws = self.ws
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _get_status_display(self, status):
        """Converte código de status para exibição"""
        status_map = {
            'PLANEJADO': 'Planejado',
            'CONFIRMADO': 'Confirmado',
            'REALIZADO': 'Realizado',
            'CANCELADO': 'Cancelado'
        }
        return status_map.get(status, status)
    
    def _get_origem_display(self, origem):
        """Converte código de origem para exibição"""
        origem_map = {
            'PROCEDIMENTO': 'Procedimento Operacional',
            'MATRIZ': 'Matriz de Habilidades',
            'DEMANDA': 'Demanda de Treinamento',
            'LIVRE': 'Planejamento Livre'
        }
        return origem_map.get(origem, origem)
    
    def _generate_response(self, filename):
        """Gera HttpResponse com o arquivo Excel"""
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
