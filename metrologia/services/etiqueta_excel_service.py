# -*- coding: utf-8 -*-
"""
Serviço para Renderização e Geração de Etiquetas de Instrumentos em Excel (.xlsx)
Suporta upload de templates customizados, preservação de estilos e substituição de tags dinâmicas.
"""

import io
import re
import base64
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TAGS_ETIQUETAS_METROLOGIA = [
    {
        'tag': '{{TAG}}',
        'sinonimos': ['{{CODIGO}}'],
        'campo': 'tag',
        'descricao': 'TAG / Identificação única do instrumento',
        'exemplo': 'ST-01'
    },
    {
        'tag': '{{CODIGO_INTERNO}}',
        'sinonimos': [],
        'campo': 'codigo',
        'descricao': 'Código interno de controle',
        'exemplo': 'CAL-0042'
    },
    {
        'tag': '{{DESCRICAO}}',
        'sinonimos': ['{{NOME}}'],
        'campo': 'descricao',
        'descricao': 'Descrição ou nome do instrumento',
        'exemplo': 'Stabillon Granville-Phillips'
    },
    {
        'tag': '{{FABRICANTE}}',
        'sinonimos': ['{{MARCA}}'],
        'campo': 'fabricante',
        'descricao': 'Fabricante / Marca do instrumento',
        'exemplo': 'Granville-Phillips'
    },
    {
        'tag': '{{MODELO}}',
        'sinonimos': [],
        'campo': 'modelo',
        'descricao': 'Modelo do instrumento',
        'exemplo': 'mks 275'
    },
    {
        'tag': '{{SERIE}}',
        'sinonimos': ['{{NUMERO_SERIE}}', '{{SN}}'],
        'campo': 'serie',
        'descricao': 'Número de série do instrumento',
        'exemplo': 'SN-8849102'
    },
    {
        'tag': '{{SETOR}}',
        'sinonimos': ['{{DEPARTAMENTO}}', '{{AREA}}'],
        'campo': 'setor',
        'descricao': 'Setor / Departamento onde o instrumento está alocado',
        'exemplo': 'HMC'
    },
    {
        'tag': '{{LOCALIZACAO}}',
        'sinonimos': ['{{LOCAL}}'],
        'campo': 'localizacao',
        'descricao': 'Localização física do instrumento',
        'exemplo': 'Bancada 02'
    },
    {
        'tag': '{{RESPONSAVEL}}',
        'sinonimos': ['{{GESTOR}}', '{{USUARIO}}'],
        'campo': 'responsavel',
        'descricao': 'Nome do colaborador responsável pelo instrumento',
        'exemplo': 'Vinícius Mota'
    },
    {
        'tag': '{{CATEGORIA}}',
        'sinonimos': ['{{TIPO}}'],
        'campo': 'categoria',
        'descricao': 'Categoria metrológica do instrumento',
        'exemplo': 'Termohigrômetro'
    },
    {
        'tag': '{{TRATATIVA}}',
        'sinonimos': ['{{TIPO_CALIBRACAO}}'],
        'campo': 'tratativa_calibracao',
        'descricao': 'Tratativa de calibração (Interna / Externa)',
        'exemplo': 'Interna'
    },
    {
        'tag': '{{ACAO}}',
        'sinonimos': [],
        'campo': 'acao',
        'descricao': 'Ação padrão (Calibração ou Verificação)',
        'exemplo': 'Calibração'
    },
    {
        'tag': '{{DATA_CALIBRACAO}}',
        'sinonimos': ['{{ULTIMA_CALIBRACAO}}', '{{CALIBRADO_EM}}'],
        'campo': 'data_ultima_calibracao',
        'descricao': 'Data da última calibração realizada (DD/MM/AAAA)',
        'exemplo': '19/03/2024'
    },
    {
        'tag': '{{PROXIMA_CALIBRACAO}}',
        'sinonimos': ['{{VALIDADE}}', '{{VENCIMENTO}}', '{{PROXIMA}}'],
        'campo': 'data_proxima_calibracao',
        'descricao': 'Data da próxima calibração / validade (DD/MM/AAAA)',
        'exemplo': '19/03/2025'
    },
    {
        'tag': '{{FREQUENCIA}}',
        'sinonimos': ['{{PERIODICIDADE}}'],
        'campo': 'frequencia_meses',
        'descricao': 'Frequência de calibração em meses',
        'exemplo': '12 meses'
    },
    {
        'tag': '{{STATUS}}',
        'sinonimos': ['{{SITUACAO}}'],
        'campo': 'status_calculado',
        'descricao': 'Status / Situação metrológica atual do instrumento',
        'exemplo': 'Vigente'
    },
    {
        'tag': '{{DATA_HOJE}}',
        'sinonimos': ['{{DATA_ATUAL}}', '{{EMISSAO}}'],
        'campo': 'hoje',
        'descricao': 'Data da emissão da etiqueta (DD/MM/AAAA)',
        'exemplo': '01/09/2026'
    },
    {
        'tag': '{{CERTIFICADO}}',
        'sinonimos': ['{{NUMERO_CERTIFICADO}}'],
        'campo': 'certificado',
        'descricao': 'Número do último certificado de calibração registrado',
        'exemplo': 'CERT-2024-089'
    },
    {
        'tag': '{{LABORATORIO}}',
        'sinonimos': ['{{ORGANISMO}}', '{{FORNECEDOR}}'],
        'campo': 'laboratorio',
        'descricao': 'Laboratório ou fornecedor que executou a calibração',
        'exemplo': 'Laboratório Interno'
    },
]


def format_date_br(d):
    """Formata data no padrão brasileiro DD/MM/AAAA."""
    if not d:
        return ""
    if isinstance(d, (datetime, date)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def sanitize_sheet_title(title):
    """Sanitiza o nome de uma aba para o Excel (máx 31 caracteres, sem caracteres inválidos)."""
    if not title:
        return "Etiqueta"
    clean = re.sub(r'[\\/*?:\[\]]', '_', str(title)).strip()
    return clean[:31] or "Etiqueta"


def get_instrumento_tag_dict(inst, index=None):
    """
    Constrói um dicionário com todos os valores de tags disponíveis para um instrumento.
    """
    hoje = date.today()
    
    # Status calculado
    status = "Ativo"
    if not inst.ativo:
        status = "Inativo"
    elif inst.data_proxima_calibracao:
        if inst.data_proxima_calibracao < hoje:
            status = "Vencido"
        elif (inst.data_proxima_calibracao - hoje).days <= 30:
            status = "Vence em 30d"
        else:
            status = "Vigente"

    # Último certificado / calibração
    cert_num = ""
    lab_nome = ""
    try:
        if hasattr(inst, 'calibracoes') and inst.calibracoes.exists():
            last_cal = inst.calibracoes.order_by('-data_calibracao').first()
            if last_cal:
                cert_num = last_cal.numero_certificado or ""
                if last_cal.laboratorio:
                    lab_nome = last_cal.laboratorio.nome
                elif last_cal.fornecedor:
                    lab_nome = last_cal.fornecedor.nome
    except Exception:
        pass

    # Categoria e ação
    cat_nome = inst.categoria.nome if inst.categoria else ""
    acao_str = "Calibração"
    if inst.categoria and hasattr(inst.categoria, 'acao'):
        acao_str = inst.categoria.get_acao_display() if hasattr(inst.categoria, 'get_acao_display') else str(inst.categoria.acao)

    setor_nome = inst.setor.nome if inst.setor else ""
    resp_nome = inst.responsavel.nome_completo if inst.responsavel else ""
    freq_str = f"{inst.frequencia_meses} meses" if inst.frequencia_meses else ""

    data_map = {
        '{{TAG}}': inst.tag or "",
        '{{CODIGO}}': inst.tag or "",
        '{{CODIGO_INTERNO}}': inst.codigo or "",
        '{{DESCRICAO}}': inst.descricao or "",
        '{{NOME}}': inst.descricao or "",
        '{{FABRICANTE}}': inst.fabricante or "",
        '{{MARCA}}': inst.fabricante or "",
        '{{MODELO}}': inst.modelo or "",
        '{{SERIE}}': inst.serie or "",
        '{{NUMERO_SERIE}}': inst.serie or "",
        '{{SN}}': inst.serie or "",
        '{{SETOR}}': setor_nome,
        '{{DEPARTAMENTO}}': setor_nome,
        '{{AREA}}': setor_nome,
        '{{LOCALIZACAO}}': inst.localizacao or "",
        '{{LOCAL}}': inst.localizacao or "",
        '{{RESPONSAVEL}}': resp_nome,
        '{{GESTOR}}': resp_nome,
        '{{USUARIO}}': resp_nome,
        '{{CATEGORIA}}': cat_nome,
        '{{TIPO}}': cat_nome,
        '{{TRATATIVA}}': inst.get_tratativa_calibracao_display() if hasattr(inst, 'get_tratativa_calibracao_display') else (inst.tratativa_calibracao or ""),
        '{{TIPO_CALIBRACAO}}': inst.get_tratativa_calibracao_display() if hasattr(inst, 'get_tratativa_calibracao_display') else (inst.tratativa_calibracao or ""),
        '{{ACAO}}': acao_str,
        '{{DATA_CALIBRACAO}}': format_date_br(inst.data_ultima_calibracao),
        '{{ULTIMA_CALIBRACAO}}': format_date_br(inst.data_ultima_calibracao),
        '{{CALIBRADO_EM}}': format_date_br(inst.data_ultima_calibracao),
        '{{PROXIMA_CALIBRACAO}}': format_date_br(inst.data_proxima_calibracao),
        '{{VALIDADE}}': format_date_br(inst.data_proxima_calibracao),
        '{{VENCIMENTO}}': format_date_br(inst.data_proxima_calibracao),
        '{{PROXIMA}}': format_date_br(inst.data_proxima_calibracao),
        '{{FREQUENCIA}}': freq_str,
        '{{PERIODICIDADE}}': freq_str,
        '{{STATUS}}': status,
        '{{SITUACAO}}': status,
        '{{DATA_HOJE}}': format_date_br(hoje),
        '{{DATA_ATUAL}}': format_date_br(hoje),
        '{{EMISSAO}}': format_date_br(hoje),
        '{{CERTIFICADO}}': cert_num,
        '{{NUMERO_CERTIFICADO}}': cert_num,
        '{{LABORATORIO}}': lab_nome or ("Laboratório Interno" if inst.tratativa_calibracao == 'INTERNA' else "Fornecedor Externo"),
        '{{ORGANISMO}}': lab_nome or ("Laboratório Interno" if inst.tratativa_calibracao == 'INTERNA' else "Fornecedor Externo"),
        '{{FORNECEDOR}}': lab_nome or ("Laboratório Interno" if inst.tratativa_calibracao == 'INTERNA' else "Fornecedor Externo"),
    }

    # Se informado um índice (ex: 1, 2, 3), adiciona também as tags indexadas {{TAG_1}}, {{DESCRICAO_1}}, etc.
    if index is not None:
        idx = str(index)
        for tag_k, val in list(data_map.items()):
            # Ex: {{TAG}} -> {{TAG_1}}
            tag_name = tag_k.strip('{}')
            indexed_tag = f"{{{{{tag_name}_{idx}}}}}"
            data_map[indexed_tag] = val

    return data_map


def replace_tags_in_sheet(sheet, tag_dict):
    """
    Substitui todas as tags presentes nas células de uma aba do Excel preservando a formatação.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                changed = False
                for tag, repl in tag_dict.items():
                    if tag in val:
                        val = val.replace(tag, str(repl))
                        changed = True
                if changed:
                    cell.value = val


def generate_default_etiquetas_workbook(instrumentos):
    """
    Gera uma planilha estilizada padrão de etiquetas caso não haja template customizado cadastrado.
    Layout elegante com cards de identificação metrológica.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etiquetas Metrológicas"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    card_bg = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    title_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=8, bold=True, color="64748B")
    val_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    tag_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    alert_font = Font(name="Calibri", size=11, bold=True, color="DC2626")
    ok_font = Font(name="Calibri", size=11, bold=True, color="16A34A")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Configuração de larguras
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 3

    current_row = 2

    # Título do Relatório
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
    t_cell = ws.cell(row=current_row, column=2, value="CALIBRAWEB • ETIQUETAS DE IDENTIFICAÇÃO METROLÓGICA")
    t_cell.font = title_font
    t_cell.fill = header_fill
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 2

    hoje = date.today()

    for idx, inst in enumerate(instrumentos, 1):
        d_cal = format_date_br(inst.data_ultima_calibracao) or "Pendente"
        d_prox = format_date_br(inst.data_proxima_calibracao) or "Pendente"
        setor_nome = inst.setor.nome if inst.setor else "-"
        resp_nome = inst.responsavel.nome_completo if inst.responsavel else "-"
        
        is_vencido = inst.data_proxima_calibracao and inst.data_proxima_calibracao < hoje

        # Header do Card da Etiqueta
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        c_tag = ws.cell(row=current_row, column=2, value=f"TAG: {inst.tag}")
        c_tag.font = tag_font
        c_tag.fill = light_blue_fill
        c_tag.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
        c_tipo = ws.cell(row=current_row, column=4, value=f"{inst.categoria.nome if inst.categoria else 'Instrumento'}")
        c_tipo.font = Font(name="Calibri", size=10, bold=True, color="475569")
        c_tipo.fill = light_blue_fill
        c_tipo.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Descrição
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        c_desc = ws.cell(row=current_row, column=2, value=f"{inst.descricao}")
        c_desc.font = val_font
        c_desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Setor e Fabricante
        ws.cell(row=current_row, column=2, value="SETOR / LOCAL:").font = label_font
        ws.cell(row=current_row, column=3, value=f"{setor_nome}").font = val_font
        ws.cell(row=current_row, column=4, value="FABRICANTE / MODELO:").font = label_font
        ws.cell(row=current_row, column=5, value=f"{inst.fabricante or ''} {inst.modelo or ''}".strip() or "-").font = val_font
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Última e Próxima Calibração
        ws.cell(row=current_row, column=2, value="ÚLTIMA CALIBRAÇÃO:").font = label_font
        ws.cell(row=current_row, column=3, value=d_cal).font = val_font
        ws.cell(row=current_row, column=4, value="PRÓXIMA CALIBRAÇÃO:").font = label_font
        
        c_prox = ws.cell(row=current_row, column=5, value=d_prox)
        c_prox.font = alert_font if is_vencido else ok_font
        ws.row_dimensions[current_row].height = 22

        # Bordas no bloco
        for r in range(current_row - 3, current_row + 1):
            for c in range(2, 6):
                ws.cell(row=r, column=c).border = thin_border

        current_row += 2

    return wb


def render_etiquetas_excel(template_obj, instrumentos):
    """
    Renderiza as etiquetas em um arquivo Excel (.xlsx) e retorna os bytes do arquivo.
    Suporta variações:
      - INDIVIDUAL: 1 etiqueta com o instrumento principal (ou o primeiro da seleção)
      - MULTI_ABA: Cria uma aba para cada instrumento selecionado a partir da aba modelo
      - GRADE_TABELA: Preenche múltiplos instrumentos na mesma planilha (tags normais e indexadas)
    """
    if not instrumentos:
        raise ValueError("Nenhum instrumento selecionado para geração de etiquetas.")

    # Se não houver template customizado, gerar com o layout padrão do sistema
    if not template_obj or (not template_obj.arquivo_base64 and not template_obj.arquivo):
        wb = generate_default_etiquetas_workbook(instrumentos)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # Carregar template a partir de Base64 ou Arquivo
    if template_obj.arquivo_base64:
        content = base64.b64decode(template_obj.arquivo_base64)
    else:
        template_obj.arquivo.seek(0)
        content = template_obj.arquivo.read()

    wb = openpyxl.load_workbook(io.BytesIO(content))
    var_tipo = template_obj.tipo_variacao or 'INDIVIDUAL'

    if var_tipo == 'INDIVIDUAL':
        # Gera a etiqueta para o primeiro instrumento (ou folha única)
        inst = instrumentos[0]
        tag_dict = get_instrumento_tag_dict(inst)
        for sheet in wb.worksheets:
            replace_tags_in_sheet(sheet, tag_dict)

    elif var_tipo == 'MULTI_ABA':
        # Clona a aba modelo para cada instrumento selecionado
        template_sheet = wb.active
        orig_sheet_name = template_sheet.title

        for idx, inst in enumerate(instrumentos):
            tag_dict = get_instrumento_tag_dict(inst)
            sheet_title = sanitize_sheet_title(inst.tag or f"Etiqueta_{idx+1}")
            
            # Garantir nome único de aba
            unique_title = sheet_title
            counter = 1
            while unique_title in wb.sheetnames:
                unique_title = f"{sheet_title[:28]}_{counter}"
                counter += 1

            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = unique_title
            replace_tags_in_sheet(new_sheet, tag_dict)

        # Remove a aba original de modelo para deixar apenas as abas dos instrumentos
        if orig_sheet_name in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb[orig_sheet_name])

    elif var_tipo == 'GRADE_TABELA':
        # Preenchimento em grade na mesma folha
        # Monta tags indexadas {{TAG_1}}, {{TAG_2}}, etc., e também preenche o primeiro como base
        all_tags = {}
        for idx, inst in enumerate(instrumentos, 1):
            inst_tags = get_instrumento_tag_dict(inst, index=idx)
            all_tags.update(inst_tags)
        
        # Tags sem índice pegam o primeiro instrumento como fallback
        first_tags = get_instrumento_tag_dict(instrumentos[0])
        for k, v in first_tags.items():
            if k not in all_tags:
                all_tags[k] = v

        for sheet in wb.worksheets:
            replace_tags_in_sheet(sheet, all_tags)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
