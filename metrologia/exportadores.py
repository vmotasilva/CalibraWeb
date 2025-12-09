# -*- coding: utf-8 -*-
"""
Módulo de exportação de dados - Fase 5
Suporta Excel, CSV e PDF
"""

from io import BytesIO
from datetime import date, timedelta
from django.http import HttpResponse
from django.db.models import Q, Count

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

import csv


class ExportadorInstrumentos:
    """Exportador de instrumentos em múltiplos formatos"""
    
    def __init__(self, queryset, filtros_aplicados=None):
        """
        Args:
            queryset: QuerySet de instrumentos
            filtros_aplicados: Dict com filtros aplicados (para logging)
        """
        self.queryset = queryset
        self.filtros_aplicados = filtros_aplicados or {}
        self.hoje = date.today()
    
    # =========================================================================
    # EXPORTAÇÃO EXCEL
    # =========================================================================
    
    def exportar_excel(self):
        """
        Exporta instrumentos para Excel com formatação
        
        Returns:
            HttpResponse com arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Instrumentos"
        
        # Definir largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Tag', 'Descrição', 'Categoria', 'Setor', 'Próx. Calib.', 'Status', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dados
        row = 2
        for instrumento in self.queryset:
            status = self._get_status_text(instrumento)
            situacao = "Ativo" if instrumento.ativo else "Inativo"
            
            ws.cell(row=row, column=1, value=instrumento.tag)
            ws.cell(row=row, column=2, value=instrumento.descricao)
            ws.cell(row=row, column=3, value=str(instrumento.categoria) if instrumento.categoria else "")
            ws.cell(row=row, column=4, value=str(instrumento.setor) if instrumento.setor else "")
            
            if instrumento.data_proxima_calibracao:
                ws.cell(row=row, column=5, value=instrumento.data_proxima_calibracao.strftime("%d/%m/%Y"))
            
            ws.cell(row=row, column=6, value=status)
            ws.cell(row=row, column=7, value=situacao)
            
            # Aplicar bordas
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
            
            row += 1
        
        # Adicionar resumo
        resumo_row = row + 2
        ws.cell(row=resumo_row, column=1, value="RESUMO")
        ws.cell(row=resumo_row, column=1).font = Font(bold=True, size=11)
        
        resumo_row += 1
        ws.cell(row=resumo_row, column=1, value="Total de Instrumentos:")
        ws.cell(row=resumo_row, column=2, value=self.queryset.count())
        
        resumo_row += 1
        ativo_count = self.queryset.filter(ativo=True).count()
        ws.cell(row=resumo_row, column=1, value="Ativos:")
        ws.cell(row=resumo_row, column=2, value=ativo_count)
        
        resumo_row += 1
        vencido_count = self.queryset.filter(
            data_proxima_calibracao__lt=self.hoje,
            ativo=True
        ).count()
        ws.cell(row=resumo_row, column=1, value="Vencidos:")
        ws.cell(row=resumo_row, column=2, value=vencido_count)
        
        # Criar resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="instrumentos.xlsx"'
        
        wb.save(response)
        return response
    
    # =========================================================================
    # EXPORTAÇÃO CSV
    # =========================================================================
    
    def exportar_csv(self):
        """
        Exporta instrumentos para CSV
        
        Returns:
            HttpResponse com arquivo CSV
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="instrumentos.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Tag',
            'Descrição',
            'Categoria',
            'Setor',
            'Próxima Calibração',
            'Status',
            'Situação'
        ])
        
        # Dados
        for instrumento in self.queryset:
            status = self._get_status_text(instrumento)
            situacao = "Ativo" if instrumento.ativo else "Inativo"
            
            writer.writerow([
                instrumento.tag,
                instrumento.descricao,
                str(instrumento.categoria) if instrumento.categoria else "",
                str(instrumento.setor) if instrumento.setor else "",
                instrumento.data_proxima_calibracao.strftime("%d/%m/%Y") if instrumento.data_proxima_calibracao else "",
                status,
                situacao
            ])
        
        return response
    
    # =========================================================================
    # EXPORTAÇÃO PDF
    # =========================================================================
    
    def exportar_pdf(self):
        """
        Exporta instrumentos para PDF
        
        Returns:
            HttpResponse com arquivo PDF
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab não está instalado. Execute: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="instrumentos.pdf"'
        
        # Criar PDF
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("RELATÓRIO DE INSTRUMENTOS", title_style))
        elements.append(Spacer(1, 12))
        
        # Info de geração
        info_text = f"Gerado em: {self.hoje.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Tabela de dados
        data = [['Tag', 'Descrição', 'Categoria', 'Status']]
        
        for instrumento in self.queryset[:100]:  # Limitar a 100 para não ficar muito grande
            status = self._get_status_text(instrumento)
            data.append([
                instrumento.tag,
                instrumento.descricao[:30],  # Truncar descrição
                str(instrumento.categoria) if instrumento.categoria else "",
                status
            ])
        
        # Criar tabela
        table = Table(data, colWidths=[1*inch, 2.5*inch, 1.2*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Adicionar resumo ao final
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("<b>RESUMO</b>", styles['Heading2']))
        
        total = self.queryset.count()
        ativos = self.queryset.filter(ativo=True).count()
        vencidos = self.queryset.filter(data_proxima_calibracao__lt=self.hoje, ativo=True).count()
        
        resumo_text = f"""
        Total de Instrumentos: {total}<br/>
        Instrumentos Ativos: {ativos}<br/>
        Instrumentos Vencidos: {vencidos}
        """
        elements.append(Paragraph(resumo_text, styles['Normal']))
        
        # Gerar PDF
        doc.build(elements)
        
        return response
    
    # =========================================================================
    # HELPERS
    # =========================================================================
    
    def _get_status_text(self, instrumento):
        """
        Retorna texto de status do instrumento
        
        Args:
            instrumento: Instância de Instrumento
        
        Returns:
            str: Status em texto
        """
        if not instrumento.ativo:
            return "Inativo"
        
        if not instrumento.data_proxima_calibracao:
            return "Sem data"
        
        if instrumento.data_proxima_calibracao < self.hoje:
            return "Vencido"
        
        data_limite = self.hoje + timedelta(days=30)
        if instrumento.data_proxima_calibracao <= data_limite:
            return "A Vencer"
        
        return "Vigente"


class ExportadorEstatisticas:
    """Exportador de relatórios de estatísticas"""
    
    def __init__(self, data_stats):
        """
        Args:
            data_stats: Dict com dados de estatísticas (from estatisticas_calibracao_view)
        """
        self.data = data_stats
        self.hoje = date.today()
    
    def exportar_excel(self):
        """Exporta estatísticas para Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
        
        wb = Workbook()
        
        # Sheet 1: KPIs
        ws_kpis = wb.active
        ws_kpis.title = "KPIs"
        
        self._preench_kpis_excel(ws_kpis)
        
        # Sheet 2: Por Categoria
        if self.data.get('por_categoria'):
            ws_cat = wb.create_sheet("Por Categoria")
            self._preench_categoria_excel(ws_cat)
        
        # Sheet 3: Por Setor
        if self.data.get('por_setor'):
            ws_setor = wb.create_sheet("Por Setor")
            self._preench_setor_excel(ws_setor)
        
        # Resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="estatisticas.xlsx"'
        
        wb.save(response)
        return response
    
    def _preench_kpis_excel(self, ws):
        """Preenche sheet de KPIs"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Headers
        ws['A1'] = "Métrica"
        ws['B1'] = "Valor"
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        # Dados
        row = 2
        metrics = [
            ("Total de Instrumentos", self.data.get('total_instrumentos', 0)),
            ("Instrumentos Vencidos", self.data.get('total_vencidos', 0)),
            ("A Vencer em 30 dias", self.data.get('vencer_30_dias', 0)),
            ("Instrumentos Vigentes", self.data.get('total_vigentes', 0)),
            ("% Vencidos", f"{self.data.get('percentage_vencidos', 0)}%"),
            ("Total de Calibrações", self.data.get('total_historicos', 0)),
            ("Aprovadas s/ Correção", self.data.get('aprovados', 0)),
            ("Aprovadas c/ Correção", self.data.get('com_correcao', 0)),
            ("Reprovadas", self.data.get('reprovados', 0)),
            ("% Aprovadas", f"{self.data.get('percentage_aprovados', 0)}%"),
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
    
    def _preench_categoria_excel(self, ws):
        """Preenche sheet de categorias"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        
        # Headers
        headers = ['Categoria', 'Total', 'Vencidos', '%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados
        row = 2
        for cat in self.data.get('por_categoria', []):
            ws.cell(row=row, column=1, value=cat.nome)
            ws.cell(row=row, column=2, value=cat.total)
            ws.cell(row=row, column=3, value=cat.vencidos)
            if cat.total > 0:
                pct = round((cat.vencidos / cat.total) * 100, 1)
            else:
                pct = 0
            ws.cell(row=row, column=4, value=f"{pct}%")
            row += 1
    
    def _preench_setor_excel(self, ws):
        """Preenche sheet de setores"""
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        
        # Headers
        headers = ['Setor', 'Total', 'Vencidos', '%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados
        row = 2
        for setor in self.data.get('por_setor', []):
            ws.cell(row=row, column=1, value=setor.nome)
            ws.cell(row=row, column=2, value=setor.total)
            ws.cell(row=row, column=3, value=setor.vencidos)
            if setor.total > 0:
                pct = round((setor.vencidos / setor.total) * 100, 1)
            else:
                pct = 0
            ws.cell(row=row, column=4, value=f"{pct}%")
            row += 1
    
    def exportar_pdf(self):
        """Exporta estatísticas para PDF"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab não está instalado. Execute: pip install reportlab")
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="estatisticas.pdf"'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph("RELATÓRIO DE ESTATÍSTICAS", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Gerado em: {self.hoje.strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # KPIs
        elements.append(Paragraph("<b>PRINCIPAIS INDICADORES</b>", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        kpi_data = [
            ['Total de Instrumentos', str(self.data.get('total_instrumentos', 0))],
            ['Vencidos', f"{self.data.get('total_vencidos', 0)} ({self.data.get('percentage_vencidos', 0)}%)"],
            ['A Vencer (30 dias)', str(self.data.get('vencer_30_dias', 0))],
            ['Vigentes', str(self.data.get('total_vigentes', 0))],
        ]
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        
        elements.append(kpi_table)
        elements.append(Spacer(1, 20))
        
        # Gerar PDF
        doc.build(elements)
        
        return response
