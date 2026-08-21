from datetime import date

from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.test import TestCase
from django.test import RequestFactory
from django.urls import reverse

from .models import (
    ComentarioRespostaAuditoria,
    ModeloAuditoria,
    PerguntaAuditoria,
    RegistroAuditoria,
    RelatorioCompartilhadoAuditoria,
)
from .views import _build_registro_report_share_token, registros_por_modelo


class ModeloDeleteProtectedTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditoria_admin",
            email="auditoria_admin@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice

            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            # Alguns ambientes de teste removem apps OTP; nesse caso seguimos sem dispositivo.
            pass

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO TESTE PROTECT",
            objeto_auditoria="Objeto de auditoria para teste",
            periodicidade="MENSAL",
        )

        RegistroAuditoria.objects.create(
            modelo=self.modelo,
            data_auditoria=date.today(),
            periodo_inicio=date.today(),
            periodo_fim=date.today(),
        )

    def test_delete_modelo_with_registros_protegidos_returns_redirect_and_message(self):
        self.client.force_login(self.user)

        response = self.client.post(reverse("auditoria:modelo_delete", args=[self.modelo.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("auditoria:modelos_list"))
        self.assertTrue(ModeloAuditoria.objects.filter(pk=self.modelo.pk).exists())

        messages = [str(m) for m in get_messages(response.wsgi_request)]
        self.assertTrue(
            any("Não foi possível remover este modelo" in m for m in messages),
            f"Mensagem esperada não encontrada. Mensagens: {messages}",
        )


class ModeloDuplicateTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditoria_admin_duplicate",
            email="auditoria_admin_duplicate@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice

            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            pass

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO BASE DUPLICACAO",
            objeto_auditoria="Objeto base",
            periodicidade="MENSAL",
        )
        self.pergunta = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta com descrição",
            descricao_detalhada="Descrição detalhada importante para instrução.",
            tipo_resposta="SIM_NAO",
            exibir_grafico=False,
            ordem=1,
            obrigatoria=True,
            ativo=True,
        )

    def test_modelo_duplicate_copies_pergunta_descricao_detalhada(self):
        self.client.force_login(self.user)

        response = self.client.post(reverse("auditoria:modelo_duplicate", args=[self.modelo.pk]))

        self.assertEqual(response.status_code, 302)
        novo_modelo = ModeloAuditoria.objects.exclude(pk=self.modelo.pk).get()
        pergunta_duplicada = PerguntaAuditoria.objects.get(modelo=novo_modelo, ordem=1)
        self.assertEqual(
            pergunta_duplicada.descricao_detalhada,
            "Descrição detalhada importante para instrução.",
        )
        self.assertFalse(pergunta_duplicada.exibir_grafico)

    def test_modelo_duplicate_uses_custom_target_name(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("auditoria:modelo_duplicate", args=[self.modelo.pk]),
            data={"novo_nome": "MODELO BASE DUPLICACAO - RENOMEADO"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ModeloAuditoria.objects.filter(nome="MODELO BASE DUPLICACAO - RENOMEADO").exists())

    def test_modelo_duplicate_rejects_existing_name(self):
        self.client.force_login(self.user)
        ModeloAuditoria.objects.create(
            nome="NOME JA EXISTENTE",
            objeto_auditoria="Objeto existente",
            periodicidade="MENSAL",
        )

        response = self.client.post(
            reverse("auditoria:modelo_duplicate", args=[self.modelo.pk]),
            data={"novo_nome": "NOME JA EXISTENTE"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(ModeloAuditoria.objects.filter(nome="NOME JA EXISTENTE").count(), 1)
        messages = [str(m) for m in get_messages(response.wsgi_request)]
        self.assertTrue(any("Já existe um modelo com este nome" in m for m in messages))


class ComentarioPerguntaSemRegistroTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditoria_user",
            email="auditoria_user@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice

            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            pass

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO COMENTARIO DATA",
            objeto_auditoria="Objeto de auditoria para comentário por data",
            periodicidade="MENSAL",
        )
        self.pergunta = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta de teste",
            ordem=1,
            tipo_resposta="SIM_NAO",
            ativo=True,
        )

    def test_add_question_comment_without_registro_persists_with_data_referencia(self):
        self.client.force_login(self.user)
        data_alvo = date(2026, 3, 10)

        response = self.client.post(
            reverse("auditoria:registros_por_modelo", args=[self.modelo.pk]),
            data={
                "action": "add_question_comment",
                "pergunta_id": str(self.pergunta.pk),
                "comentario_data": data_alvo.isoformat(),
                "comentario": "Comentário sem registro na data",
            },
        )

        self.assertEqual(response.status_code, 302)
        comentario = ComentarioRespostaAuditoria.objects.get(pergunta=self.pergunta)
        self.assertIsNone(comentario.registro)
        self.assertEqual(comentario.data_referencia, data_alvo)
        self.assertEqual(comentario.texto, "Comentário sem registro na data")

    def test_registro_create_preloads_comments_from_matching_period(self):
        self.client.force_login(self.user)
        data_alvo = date.today()
        ComentarioRespostaAuditoria.objects.create(
            registro=None,
            pergunta=self.pergunta,
            autor=self.user,
            texto="Comentário pré-registro para exibir no formulário",
            data_referencia=data_alvo,
        )

        response = self.client.get(reverse("auditoria:registro_create_modelo", args=[self.modelo.pk]))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("comentarios-dados", content)
        self.assertIn("Coment\\u00e1rio pr\\u00e9-registro para exibir no formul\\u00e1rio", content)


class PerguntaCreateTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditoria_admin_pergunta_create",
            email="auditoria_admin_pergunta_create@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice

            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            pass

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO CREATE PERGUNTA",
            objeto_auditoria="Objeto base do teste de criação",
            periodicidade="MENSAL",
        )

    def test_pergunta_create_handles_missing_checkbox_and_saves_false(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("auditoria:pergunta_create"),
            data={
                "modelo": str(self.modelo.pk),
                "subcategoria": "",
                "pergunta": "Pergunta criada via teste",
                "descricao_detalhada": "",
                "tipo_resposta": "SIM_NAO",
                "preenchimento_semanal": "UNICO",
                "opcoes_resposta": "",
                "opcoes_resposta_cores": "",
                "aplicar_no_grid": "on",
                "ordem": "1",
                "obrigatoria": "on",
                "ativo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        pergunta = PerguntaAuditoria.objects.get(modelo=self.modelo, pergunta="Pergunta criada via teste")
        self.assertFalse(pergunta.exibir_grafico)

    def test_pergunta_create_with_iso_preset_applies_official_options(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("auditoria:pergunta_create"),
            data={
                "modelo": str(self.modelo.pk),
                "subcategoria": "",
                "pergunta": "Pergunta com preset ISO",
                "descricao_detalhada": "",
                "conjunto_resposta_padrao": "ISO",
                "tipo_resposta": "SIM_NAO",
                "preenchimento_semanal": "UNICO",
                "opcoes_resposta": "",
                "opcoes_resposta_cores": "",
                "ordem": "1",
                "obrigatoria": "on",
                "ativo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        pergunta = PerguntaAuditoria.objects.get(modelo=self.modelo, pergunta="Pergunta com preset ISO")
        self.assertEqual(pergunta.tipo_resposta, "LISTA")
        self.assertEqual(
            pergunta.opcoes_resposta_list,
            ["Conforme", "Não Conforme", "Não Se Aplica", "Oportunidade de Melhoria"],
        )
        self.assertEqual(
            pergunta.opcoes_resposta_cores,
            {
                "Conforme": "#198754",
                "Não Conforme": "#ff0000",
                "Não Se Aplica": "#d9d9d9",
                "Oportunidade de Melhoria": "#fd7e14",
            },
        )
        self.assertTrue(pergunta.exibir_grafico)
        self.assertTrue(pergunta.aplicar_no_grid)


class PerguntaBulkRespostaPresetTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditoria_admin_bulk_resposta",
            email="auditoria_admin_bulk_resposta@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice

            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            pass

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO BULK RESPOSTA",
            objeto_auditoria="Objeto do teste em lote",
            periodicidade="MENSAL",
        )
        self.outro_modelo = ModeloAuditoria.objects.create(
            nome="MODELO BULK RESPOSTA EXTERNO",
            objeto_auditoria="Objeto externo ao lote",
            periodicidade="MENSAL",
        )
        self.pergunta_1 = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta 1",
            ordem=1,
            tipo_resposta="SIM_NAO",
            exibir_grafico=False,
            aplicar_no_grid=False,
            ativo=True,
        )
        self.pergunta_2 = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta 2",
            ordem=2,
            tipo_resposta="NUMERO",
            exibir_grafico=False,
            aplicar_no_grid=False,
            ativo=True,
        )
        self.pergunta_outro_modelo = PerguntaAuditoria.objects.create(
            modelo=self.outro_modelo,
            pergunta="Pergunta externa",
            ordem=1,
            tipo_resposta="DECIMAL",
            exibir_grafico=False,
            aplicar_no_grid=False,
            ativo=True,
        )

    def test_bulk_apply_iso_preset_updates_only_selected_questions_from_model(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("auditoria:perguntas_bulk_apply_resposta"),
            data={
                "modelo": str(self.modelo.pk),
                "filtro_subcategoria": "",
                "conjunto_resposta_padrao": "ISO",
                "pergunta_ids": [
                    str(self.pergunta_1.pk),
                    str(self.pergunta_2.pk),
                    str(self.pergunta_outro_modelo.pk),
                ],
            },
        )

        self.assertEqual(response.status_code, 302)

        self.pergunta_1.refresh_from_db()
        self.pergunta_2.refresh_from_db()
        self.pergunta_outro_modelo.refresh_from_db()

        for pergunta in (self.pergunta_1, self.pergunta_2):
            self.assertEqual(pergunta.tipo_resposta, "LISTA")
            self.assertEqual(
                pergunta.opcoes_resposta_list,
                ["Conforme", "Não Conforme", "Não Se Aplica", "Oportunidade de Melhoria"],
            )
            self.assertEqual(
                pergunta.opcoes_resposta_cores,
                {
                    "Conforme": "#198754",
                    "Não Conforme": "#ff0000",
                    "Não Se Aplica": "#d9d9d9",
                    "Oportunidade de Melhoria": "#fd7e14",
                },
            )
            self.assertTrue(pergunta.exibir_grafico)
            self.assertTrue(pergunta.aplicar_no_grid)

        self.assertEqual(self.pergunta_outro_modelo.tipo_resposta, "DECIMAL")
        self.assertEqual(self.pergunta_outro_modelo.opcoes_resposta, "")
        self.assertEqual(self.pergunta_outro_modelo.opcoes_resposta_cores, {})
        self.assertFalse(self.pergunta_outro_modelo.exibir_grafico)
        self.assertFalse(self.pergunta_outro_modelo.aplicar_no_grid)


class ComentarioPerguntaDeleteTests(TestCase):
    def setUp(self):
        self.rf = RequestFactory()
        user_model = get_user_model()
        self.author = user_model.objects.create_user(
            username="auditoria_comment_author",
            email="auditoria_comment_author@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        self.other = user_model.objects.create_user(
            username="auditoria_comment_other",
            email="auditoria_comment_other@example.com",
            password="senha-forte-123",
            is_staff=False,
        )

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO DELETE COMENTARIO PERGUNTA",
            objeto_auditoria="Objeto de auditoria para delete",
            periodicidade="MENSAL",
        )
        self.pergunta = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta para delete de comentário",
            ordem=1,
            tipo_resposta="SIM_NAO",
            ativo=True,
        )
        self.modelo.responsaveis.add(self.other)

        self.comentario = ComentarioRespostaAuditoria.objects.create(
            registro=None,
            pergunta=self.pergunta,
            autor=self.author,
            texto="Comentário removível",
            data_referencia=date.today(),
        )
        self.url = reverse("auditoria:registros_por_modelo", args=[self.modelo.pk])

    def test_author_can_delete_question_comment(self):
        request = self.rf.post(
            self.url,
            data={"action": "delete_question_comment", "comentario_id": str(self.comentario.id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        request.user = self.author
        response = registros_por_modelo(request, self.modelo.pk)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ComentarioRespostaAuditoria.objects.filter(id=self.comentario.id).exists())

    def test_non_author_cannot_delete_question_comment(self):
        request = self.rf.post(
            self.url,
            data={"action": "delete_question_comment", "comentario_id": str(self.comentario.id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        request.user = self.other
        response = registros_por_modelo(request, self.modelo.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(ComentarioRespostaAuditoria.objects.filter(id=self.comentario.id).exists())


class RelatorioCompartilhadoSomenteLeituraTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.owner = user_model.objects.create_user(
            username="auditoria_owner_share",
            email="auditoria_owner_share@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        self.viewer = user_model.objects.create_user(
            username="auditoria_viewer_share",
            email="auditoria_viewer_share@example.com",
            password="senha-forte-123",
            is_staff=False,
        )

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO RELATORIO COMPARTILHADO",
            objeto_auditoria="Objeto para teste de compartilhamento",
            periodicidade="MENSAL",
        )
        self.modelo.responsaveis.add(self.owner)

        self.pergunta = PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta para compartilhamento",
            ordem=1,
            tipo_resposta="SIM_NAO",
            ativo=True,
        )

        self.registro = RegistroAuditoria.objects.create(
            modelo=self.modelo,
            data_auditoria=date(2026, 3, 10),
            periodo_inicio=date(2026, 3, 1),
            periodo_fim=date(2026, 3, 31),
            avaliador=self.owner,
        )

        self.base_url = reverse("auditoria:registros_por_modelo", args=[self.modelo.pk])

    def test_viewer_can_open_shared_link_in_read_only_mode(self):
        self.client.force_login(self.viewer)
        token = _build_registro_report_share_token(
            modelo_id=self.modelo.pk,
            inicio="2026-03-01",
            fim="2026-03-31",
            subcategoria="",
        )

        response = self.client.get(f"{self.base_url}?share_token={token}")

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Visualização compartilhada: somente leitura", content)
        self.assertIn("Somente leitura", content)
        self.assertNotIn("Novo Registro", content)
        self.assertNotIn("Novo Comentário", content)

    def test_shared_link_blocks_post_actions(self):
        self.client.force_login(self.viewer)
        token = _build_registro_report_share_token(
            modelo_id=self.modelo.pk,
            inicio="2026-03-01",
            fim="2026-03-31",
            subcategoria="",
        )

        response = self.client.post(
            f"{self.base_url}?share_token={token}",
            data={"action": "add_comment", "comentario": "Tentativa indevida"},
        )

        self.assertEqual(response.status_code, 403)

    def test_invalid_shared_link_returns_403(self):
        self.client.force_login(self.viewer)

        response = self.client.get(f"{self.base_url}?share_token=token-invalido")

        self.assertEqual(response.status_code, 403)


class RelatorioCompartilhadoDirecionadoTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.remetente = user_model.objects.create_user(
            username="auditoria_sender",
            email="auditoria_sender@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        self.destinatario = user_model.objects.create_user(
            username="auditoria_target",
            email="auditoria_target@example.com",
            password="senha-forte-123",
            is_staff=False,
        )
        self.outro = user_model.objects.create_user(
            username="auditoria_other",
            email="auditoria_other@example.com",
            password="senha-forte-123",
            is_staff=False,
        )

        self.modelo = ModeloAuditoria.objects.create(
            nome="MODELO SHARE DIRECIONADO",
            objeto_auditoria="Objeto share direcionado",
            periodicidade="MENSAL",
        )
        self.modelo.responsaveis.add(self.remetente)
        PerguntaAuditoria.objects.create(
            modelo=self.modelo,
            pergunta="Pergunta base",
            ordem=1,
            tipo_resposta="SIM_NAO",
            ativo=True,
        )

        self.url = reverse("auditoria:registros_por_modelo", args=[self.modelo.pk])

    def test_sender_can_create_targeted_share(self):
        self.client.force_login(self.remetente)

        response = self.client.post(
            f"{self.url}?inicio=2026-03-01&fim=2026-03-31",
            data={"action": "share_report_targeted", "destinatario_id": str(self.destinatario.id)},
        )

        self.assertEqual(response.status_code, 302)
        share = RelatorioCompartilhadoAuditoria.objects.get(modelo=self.modelo, remetente=self.remetente)
        self.assertEqual(share.destinatario_id, self.destinatario.id)
        self.assertEqual(share.inicio.isoformat(), "2026-03-01")
        self.assertEqual(share.fim.isoformat(), "2026-03-31")

    def test_only_target_user_can_open_targeted_share(self):
        share = RelatorioCompartilhadoAuditoria.objects.create(
            modelo=self.modelo,
            remetente=self.remetente,
            destinatario=self.destinatario,
        )

        self.client.force_login(self.outro)
        response = self.client.get(f"{self.url}?share_token={share.token}")
        self.assertEqual(response.status_code, 403)

    def test_target_open_registers_receipt_proof(self):
        share = RelatorioCompartilhadoAuditoria.objects.create(
            modelo=self.modelo,
            remetente=self.remetente,
            destinatario=self.destinatario,
        )

        self.client.force_login(self.destinatario)
        response = self.client.get(f"{self.url}?share_token={share.token}")

        self.assertEqual(response.status_code, 200)
        share.refresh_from_db()
        self.assertIsNotNone(share.recebido_em)
        self.assertIsNotNone(share.primeiro_acesso_em)

    def test_home_shows_received_share_notification(self):
        RelatorioCompartilhadoAuditoria.objects.create(
            modelo=self.modelo,
            remetente=self.remetente,
            destinatario=self.destinatario,
        )

        self.client.force_login(self.destinatario)
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Relatórios Compartilhados Comigo", content)


class AuditoriaIsoExcelExportTemplateInjectionTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditor_chefe",
            first_name="Carlos",
            last_name="Silva",
            email="carlos@example.com",
            password="senha-segura-123",
            is_staff=True,
        )
        try:
            from django_otp.plugins.otp_static.models import StaticDevice
            StaticDevice.objects.create(user=self.user, name="test-device", confirmed=True)
        except Exception:
            pass

        from .models import (
            Norma,
            ItemNorma,
            AuditoriaIso,
            AgendaAuditoriaIso,
            BancoPergunta,
            RespostaEntrevistaIso,
            SolicitacaoEvidenciaIso,
            AvaliacaoFinalRequisitoIso,
        )

        self.norma = Norma.objects.create(
            codigo="ISO 13485:2016",
            descricao="Dispositivos Médicos - Sistema de Gestão da Qualidade"
        )
        self.item_pai = ItemNorma.objects.create(
            norma=self.norma,
            referencia="4",
            titulo="Sistema de Gestão da Qualidade",
            ordem=1
        )
        self.item_filho_1 = ItemNorma.objects.create(
            norma=self.norma,
            parent=self.item_pai,
            referencia="4.1.1",
            titulo="Requisitos Gerais e Documentação",
            ordem=2
        )
        self.item_filho_2 = ItemNorma.objects.create(
            norma=self.norma,
            parent=self.item_pai,
            referencia="4.1.2",
            titulo="Papel Assumido pela Organização",
            ordem=3
        )

        self.auditoria = AuditoriaIso.objects.create(
            norma=self.norma,
            data_inicio=date(2026, 6, 9),
            data_fim=date(2026, 6, 11),
            status="EM_ANDAMENTO",
            abertura_auditores="Carlos Silva (Líder)",
        )
        self.auditoria.auditores.add(self.user)
        self.auditoria.escopo_itens.add(self.item_pai, self.item_filho_1, self.item_filho_2)

        self.agenda = AgendaAuditoriaIso.objects.create(
            auditoria=self.auditoria,
            titulo="Bloco 1 - Gestão da Qualidade",
            data=date(2026, 6, 9),
        )
        self.agenda.itens_norma.add(self.item_filho_1, self.item_filho_2)

        self.pergunta = BancoPergunta.objects.create(
            texto_pergunta="Existe documentação do SGQ conforme os requisitos?",
            dica_auditor="Verificar manual da qualidade e procedimentos.",
        )
        self.pergunta.itens_norma.add(self.item_filho_1)
        self.agenda.perguntas.add(self.pergunta)

        self.resposta = RespostaEntrevistaIso.objects.create(
            auditoria=self.auditoria,
            pergunta=self.pergunta,
            texto_resposta="Procedimento PQ-001 revisado e aprovado.",
            classificacao="C"
        )
        self.solicitacao = SolicitacaoEvidenciaIso.objects.create(
            resposta=self.resposta,
            agenda=self.agenda,
            solicitacao="Apresentar cópia do Manual da Qualidade MQ-01",
            evidencia="Manual da Qualidade rev 04 apresentado em formato digital",
            conclusao="C"
        )

        # Segundo item com avaliação final gravada
        AvaliacaoFinalRequisitoIso.objects.create(
            auditoria=self.auditoria,
            item_norma=self.item_filho_2,
            classificacao="OBS",
            justificativa="Ajustar referências cruzadas no fluxo de processos.",
            atualizado_por=self.user
        )

    def test_generate_auditoria_excel_buffer_injects_metadata_and_preserves_formulas(self):
        import openpyxl
        from .services.checklist_export import generate_auditoria_excel_buffer

        buffer = generate_auditoria_excel_buffer(self.auditoria)
        self.assertIsNotNone(buffer)
        self.assertGreater(buffer.getbuffer().nbytes, 0)

        # Lê o Excel gerado
        wb = openpyxl.load_workbook(buffer, data_only=False)
        self.assertIn("Check-List", wb.sheetnames)
        self.assertIn("Resultados", wb.sheetnames)

        ws_check = wb["Check-List"]
        # Metadados no cabeçalho
        self.assertEqual(ws_check["C5"].value, "Tecnolens")
        self.assertIn("Carlos Silva", str(ws_check["C6"].value))
        self.assertEqual(ws_check["C7"].value, "09/06/2026 a 11/06/2026")
        self.assertIn("ISO 13485:2016", str(ws_check["C8"].value))
        self.assertEqual(ws_check["H6"].value, "X")  # Presencial

        # Linhas de Itens
        # 4 (Pai), 4.1.1 (Conforme), 4.1.2 (OBS)
        # Linha 13: 4
        self.assertEqual(ws_check["B13"].value, "4")
        # Linha 14: 4.1.1
        self.assertEqual(ws_check["B14"].value, "4.1.1")
        self.assertEqual(ws_check["D14"].value, "X")  # Coluna D = C (Conforme)
        self.assertEqual(ws_check["E14"].value, "")   # Coluna E = NC
        self.assertIn("Manual da Qualidade", str(ws_check["H14"].value))

        # Linha 15: 4.1.2
        self.assertEqual(ws_check["B15"].value, "4.1.2")
        self.assertIn("Ajustar referências cruzadas", str(ws_check["I15"].value))

        # Aba Evidências presente
        self.assertIn("Evidências", wb.sheetnames)
        ws_evid = wb["Evidências"]
        self.assertEqual(ws_evid["B3"].value, "EVIDÊNCIAS COM IMAGENS")

        # Aba Resultados: Fórmulas preservadas
        ws_res = wb["Resultados"]
        self.assertIn("COUNTIF", str(ws_res["D6"].value))
        self.assertIn("COUNTA", str(ws_res["D10"].value))

    def test_export_excel_endpoints(self):
        self.client.force_login(self.user)

        # Endpoint tradicional ISO
        url_iso = reverse("auditoria:iso_auditoria_export_excel", args=[self.auditoria.id])
        response_iso = self.client.get(url_iso)
        self.assertEqual(response_iso.status_code, 200)
        self.assertEqual(
            response_iso["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertIn("Checklist_Auditoria_ISO_13485_2016", response_iso["Content-Disposition"])

        # Endpoint API alias
        url_api = reverse("auditoria:api_auditoria_exportar_planilha", args=[self.auditoria.id])
        response_api = self.client.get(url_api)
        self.assertEqual(response_api.status_code, 200)
        self.assertEqual(
            response_api["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class AuditoriaIsoImagensEvidenciaTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditor_img",
            email="auditor_img@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        self.norma = Norma.objects.create(
            codigo="ISO 13485:2016",
            titulo="Dispositivos Médicos",
            ativa=True,
        )
        self.item = ItemNorma.objects.create(
            norma=self.norma,
            referencia="7.5.1",
            titulo="Controle de produção e fornecimento de serviço",
            ordem=1,
        )
        self.auditoria = AuditoriaIso.objects.create(
            norma=self.norma,
            titulo="Auditoria Imagens",
            empresa_auditada="Unidade Teste",
            criado_por=self.user,
        )
        self.auditoria.escopo_itens.add(self.item)
        self.pergunta = BancoPergunta.objects.create(
            texto_pergunta="Os processos de produção são controlados?",
            ativa=True,
        )
        self.pergunta.itens_norma.add(self.item)
        self.resposta = RespostaEntrevistaIso.objects.create(
            auditoria=self.auditoria,
            pergunta=self.pergunta,
            classificacao="C",
            respondida_por=self.user,
        )
        self.solicitacao = SolicitacaoEvidenciaIso.objects.create(
            resposta=self.resposta,
            solicitacao="Ordem de Produção OP-2026-001",
            evidencia="OP auditada na linha de montagem",
            conclusao="C",
        )

    def test_upload_imagem_via_base64_and_properties(self):
        self.client.force_login(self.user)
        url = reverse("auditoria:api_iso_solicitacao_upload_imagem", args=[self.solicitacao.id])
        
        sample_b64 = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
        response = self.client.post(
            url,
            data=json.dumps({
                "base64": sample_b64,
                "nome": "op_foto.png",
                "legenda": "Foto da OP assinada",
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["imagem"]["legenda"], "Foto da OP assinada")
        self.assertEqual(data["imagem"]["nome"], "op_foto.png")
        self.assertTrue(data["imagem"]["url"].startswith("data:image/png;base64,"))

        from auditoria.models import ImagemSolicitacaoIso
        img_obj = ImagemSolicitacaoIso.objects.get(pk=data["imagem"]["id"])
        self.assertEqual(img_obj.url_imagem, sample_b64)

    def test_update_legenda_imagem(self):
        self.client.force_login(self.user)
        from auditoria.models import ImagemSolicitacaoIso
        img_obj = ImagemSolicitacaoIso.objects.create(
            solicitacao=self.solicitacao,
            arquivo_base64="data:image/jpeg;base64,fake",
            nome_arquivo="teste.jpg",
            legenda="Legenda inicial",
        )
        url = reverse("auditoria:api_iso_solicitacao_update_legenda_imagem", args=[img_obj.id])
        response = self.client.post(
            url,
            data=json.dumps({"legenda": "Legenda atualizada pelo auditor"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        img_obj.refresh_from_db()
        self.assertEqual(img_obj.legenda, "Legenda atualizada pelo auditor")

    def test_delete_imagem(self):
        self.client.force_login(self.user)
        from auditoria.models import ImagemSolicitacaoIso
        img_obj = ImagemSolicitacaoIso.objects.create(
            solicitacao=self.solicitacao,
            arquivo_base64="data:image/jpeg;base64,fake",
            nome_arquivo="teste.jpg",
        )
        url = reverse("auditoria:api_iso_solicitacao_delete_imagem", args=[img_obj.id])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(ImagemSolicitacaoIso.objects.filter(pk=img_obj.id).exists())


class AuditoriaIsoDocxExportTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="auditor_docx",
            email="auditor_docx@example.com",
            password="senha-forte-123",
            is_staff=True,
        )
        self.norma = Norma.objects.create(
            codigo="ISO 13485:2016",
            titulo="Dispositivos Médicos",
            ativa=True,
        )
        self.item_sec4 = ItemNorma.objects.create(
            norma=self.norma,
            referencia="4",
            titulo="Sistema de Gestão da Qualidade",
            ordem=1,
        )
        self.item_41 = ItemNorma.objects.create(
            norma=self.norma,
            referencia="4.1.1",
            titulo="Requisitos gerais de documentação",
            ordem=2,
        )
        self.item_sec7 = ItemNorma.objects.create(
            norma=self.norma,
            referencia="7",
            titulo="Realização do Produto",
            ordem=3,
        )
        self.item_75 = ItemNorma.objects.create(
            norma=self.norma,
            referencia="7.5.1",
            titulo="Controle de produção",
            ordem=4,
        )
        self.auditoria = AuditoriaIso.objects.create(
            norma=self.norma,
            titulo="Auditoria DOCX",
            empresa_auditada="Unidade Teste DOCX",
            sintese="<h2>Síntese Geral</h2><p>A auditoria transcorreu em <strong>conformidade</strong>.</p><ul><li>Item auditado 1</li><li>Item auditado 2</li></ul>",
            conclusao_texto="Recomendação de aprovação com plano de ação corretiva.",
            criado_por=self.user,
        )
        self.auditoria.escopo_itens.add(self.item_41, self.item_75)

        # Resposta com Não Conformidade (NC) em 7.5.1
        self.p_75 = BancoPergunta.objects.create(texto_pergunta="Controle de produção é seguido?", ativa=True)
        self.p_75.itens_norma.add(self.item_75)
        self.r_75 = RespostaEntrevistaIso.objects.create(
            auditoria=self.auditoria,
            pergunta=self.p_75,
            classificacao="NC",
            grau_nc="MENOR",
            respondida_por=self.user,
        )
        self.sol_75 = SolicitacaoEvidenciaIso.objects.create(
            resposta=self.r_75,
            solicitacao="Ordem de Produção OP-001",
            evidencia="Falta de preenchimento do campo de lote",
            conclusao="NC",
        )

        # Resposta com Conforme em 4.1.1
        self.p_41 = BancoPergunta.objects.create(texto_pergunta="Manual da qualidade existe?", ativa=True)
        self.p_41.itens_norma.add(self.item_41)
        self.r_41 = RespostaEntrevistaIso.objects.create(
            auditoria=self.auditoria,
            pergunta=self.p_41,
            classificacao="C",
            respondida_por=self.user,
        )
        self.sol_41 = SolicitacaoEvidenciaIso.objects.create(
            resposta=self.r_41,
            solicitacao="Manual da Qualidade",
            evidencia="Manual rev. 4 auditado e em conformidade",
            conclusao="C",
        )

    def test_generate_relatorio_docx_buffer(self):
        from auditoria.services.relatorio_docx_export import generate_relatorio_docx_buffer
        buffer = generate_relatorio_docx_buffer(self.auditoria)
        self.assertIsNotNone(buffer)
        docx_bytes = buffer.getvalue()
        self.assertTrue(len(docx_bytes) > 1000)

        # Valida que o DOCX gerado pode ser lido pelo docx.Document
        doc = Document(io.BytesIO(docx_bytes))
        self.assertTrue(len(doc.paragraphs) > 0)
        self.assertTrue(len(doc.tables) > 0)

        # Verifica presença de texto nos parágrafos e tabelas
        doc_text = " ".join([p.text for p in doc.paragraphs])
        for t in doc.tables:
            for row in t.rows:
                for cell in row.cells:
                    doc_text += " " + cell.text

        self.assertIn("RELATÓRIO DE AUDITORIA INTERNA DA QUALIDADE", doc_text)
        self.assertIn("Unidade Teste DOCX", doc_text)
        self.assertIn("Síntese Geral", doc_text)
        self.assertIn("NC Menor", doc_text)

    def test_export_docx_view_endpoint(self):
        self.client.force_login(self.user)
        url = reverse("auditoria:iso_auditoria_export_docx", args=[self.auditoria.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        self.assertIn(".docx", response["Content-Disposition"])

    def test_save_fechamento_endpoint(self):
        self.client.force_login(self.user)
        url = reverse("auditoria:api_iso_fechamento_salvar", args=[self.auditoria.id])
        response = self.client.post(
            url,
            data=json.dumps({
                "sintese": "<h3>Nova Síntese</h3><p>Organograma revisado.</p>",
                "empresa_auditada": "Tecnolens Filial 1",
                "encerramento_representantes": "Carlos Gerente",
                "conclusao_texto": "Parecer de recomendação emitido.",
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])

        self.auditoria.refresh_from_db()
        self.assertEqual(self.auditoria.empresa_auditada, "Tecnolens Filial 1")
        self.assertEqual(self.auditoria.encerramento_representantes, "Carlos Gerente")
        self.assertEqual(self.auditoria.conclusao_texto, "Parecer de recomendação emitido.")
        self.assertIn("Nova Síntese", self.auditoria.sintese)

    def test_cronograma_solicitacoes_atendidas_tab(self):
        self.client.force_login(self.user)
        url = reverse("auditoria:iso_auditoria_cronograma", args=[self.auditoria.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("solicitacoes_atendidas", response.context)
        self.assertEqual(len(response.context["solicitacoes_atendidas"]), 1)
        self.assertEqual(response.context["total_solicitacoes_atendidas"], 1)
        self.assertEqual(response.context["total_atendidas_c"], 1)

        content = response.content.decode("utf-8")
        self.assertIn('id="atendidas-tab"', content)
        self.assertIn('id="atendidas"', content)
        self.assertIn("Solicitações Atendidas / Conformes", content)




