import os
import io
import re
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings


def natural_sort_key(s: str):
    """Ordenação natural para referências como 4.1, 4.1.1, 4.10, etc."""
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s or ""))]


def get_template_path() -> str:
    """Retorna o caminho do arquivo de template no servidor."""
    base_dir = getattr(settings, 'BASE_DIR', None)
    if not base_dir:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    templates_dir = os.path.join(str(base_dir), 'auditoria', 'templates_excel')
    try:
        os.makedirs(templates_dir, exist_ok=True)
    except OSError:
        pass
    return os.path.join(templates_dir, 'checklist_norma_template.xlsx')


def create_base_template_workbook() -> openpyxl.Workbook:
    """
    Gera um objeto openpyxl.Workbook base com formatações e fórmulas preservadas
    nas abas 'Check-List' e 'Resultados' diretamente em memória.
    """
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # 1. ABA 'Check-List'
    # -------------------------------------------------------------
    ws_check = wb.active
    ws_check.title = "Check-List"
    ws_check.views.sheetView[0].showGridLines = True

    # Paleta de Estilos
    navy_dark = "0B2545"
    navy_blue = "134074"
    blue_header = "1D4ED8"
    gray_light = "F1F5F9"
    gray_border = "CBD5E1"
    
    border_thin = Border(
        left=Side(style='thin', color=gray_border),
        right=Side(style='thin', color=gray_border),
        top=Side(style='thin', color=gray_border),
        bottom=Side(style='thin', color=gray_border)
    )

    # Título Principal
    ws_check.merge_cells('B2:I3')
    title_cell = ws_check['B2']
    title_cell.value = "CHECK-LIST DE AUDITORIA INTERNA DA QUALIDADE"
    title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalho / Metadados (Linhas 5 a 8)
    meta_labels = [
        ("B5", "Unidade / Empresa:", "C5:E5"),
        ("B6", "Auditor(es) Líder(es):", "C6:E6"),
        ("B7", "Data da Auditoria:", "C7:E7"),
        ("B8", "Escopo / Norma:", "C8:E8"),
    ]
    for cell_pos, label, merge_range in meta_labels:
        c_lbl = ws_check[cell_pos]
        c_lbl.value = label
        c_lbl.font = Font(name="Segoe UI", size=10, bold=True, color=navy_blue)
        c_lbl.fill = PatternFill(start_color=gray_light, end_color=gray_light, fill_type="solid")
        c_lbl.alignment = Alignment(horizontal="right", vertical="center")
        c_lbl.border = border_thin

        ws_check.merge_cells(merge_range)
        top_left_cell = ws_check[merge_range.split(':')[0]]
        top_left_cell.font = Font(name="Segoe UI", size=10, bold=False)
        top_left_cell.alignment = Alignment(horizontal="left", vertical="center")
        top_left_cell.border = border_thin

    # Tipo de Auditoria (Presencial vs Remota)
    ws_check['G5'].value = "Modalidade:"
    ws_check['G5'].font = Font(name="Segoe UI", size=10, bold=True, color=navy_blue)
    ws_check['G5'].fill = PatternFill(start_color=gray_light, end_color=gray_light, fill_type="solid")
    ws_check['G5'].border = border_thin
    ws_check['G5'].alignment = Alignment(horizontal="right", vertical="center")

    ws_check['G6'].value = "Presencial"
    ws_check['G6'].font = Font(name="Segoe UI", size=9, bold=True)
    ws_check['G6'].border = border_thin
    ws_check['G6'].alignment = Alignment(horizontal="center", vertical="center")
    ws_check['H6'].font = Font(name="Segoe UI", size=11, bold=True, color="166534")
    ws_check['H6'].alignment = Alignment(horizontal="center", vertical="center")
    ws_check['H6'].border = border_thin

    ws_check['G7'].value = "Remota"
    ws_check['G7'].font = Font(name="Segoe UI", size=9, bold=True)
    ws_check['G7'].border = border_thin
    ws_check['G7'].alignment = Alignment(horizontal="center", vertical="center")
    ws_check['H7'].font = Font(name="Segoe UI", size=11, bold=True, color="166534")
    ws_check['H7'].alignment = Alignment(horizontal="center", vertical="center")
    ws_check['H7'].border = border_thin

    # Cabeçalho da Tabela de Itens (Linha 12)
    headers_items = [
        ("B12", "Item", 12),
        ("C12", "Requisito / Questão Avaliada", 48),
        ("D12", "C", 6),
        ("E12", "NC", 6),
        ("F12", "NA", 6),
        ("G12", "OM", 6),
        ("H12", "Evidências Constatadas / Amostras", 45),
        ("I12", "Observações / Plano de Ação / OBS", 35),
    ]

    fill_table_hdr = PatternFill(start_color=navy_blue, end_color=navy_blue, fill_type="solid")
    font_table_hdr = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

    for cell_id, text, width in headers_items:
        cell = ws_check[cell_id]
        cell.value = text
        cell.font = font_table_hdr
        cell.fill = fill_table_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        col_letter = cell_id[0]
        ws_check.column_dimensions[col_letter].width = width

    ws_check.row_dimensions[12].height = 28

    # -------------------------------------------------------------
    # 2. ABA 'Evidências' (Evidências com Imagens)
    # -------------------------------------------------------------
    ws_evid = wb.create_sheet(title="Evidências")
    ws_evid.views.sheetView[0].showGridLines = True

    # Dimensões de colunas
    ws_evid.column_dimensions['A'].width = 4
    ws_evid.column_dimensions['B'].width = 18
    ws_evid.column_dimensions['C'].width = 38
    ws_evid.column_dimensions['D'].width = 38
    ws_evid.column_dimensions['E'].width = 38
    ws_evid.column_dimensions['F'].width = 38

    # Banner Superior: EVIDÊNCIAS COM IMAGENS
    ws_evid.merge_cells('B3:F3')
    evid_hdr = ws_evid['B3']
    evid_hdr.value = "EVIDÊNCIAS COM IMAGENS"
    evid_hdr.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
    evid_hdr.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    evid_hdr.alignment = Alignment(horizontal="center", vertical="center")
    
    gray_border_side = Side(style='thin', color="94A3B8")
    box_border = Border(left=gray_border_side, right=gray_border_side, top=gray_border_side, bottom=gray_border_side)
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_evid[f'{col}3'].border = box_border

    # -------------------------------------------------------------
    # 3. ABA 'Resultados' (com fórmulas e painéis automáticos)
    # -------------------------------------------------------------
    ws_res = wb.create_sheet(title="Resultados")
    ws_res.views.sheetView[0].showGridLines = True

    ws_res.merge_cells('B2:F3')
    res_title = ws_res['B2']
    res_title.value = "PAINEL DE RESULTADOS DA AUDITORIA"
    res_title.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    res_title.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    res_title.alignment = Alignment(horizontal="center", vertical="center")

    # Tabela Resumo com Fórmulas
    res_headers = [
        ("B5", "Classificação", 24),
        ("C5", "Sigla", 10),
        ("D5", "Quantidade", 14),
        ("E5", "% do Total Avaliado", 20),
    ]
    for cell_id, text, width in res_headers:
        c = ws_res[cell_id]
        c.value = text
        c.font = font_table_hdr
        c.fill = fill_table_hdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin
        ws_res.column_dimensions[cell_id[0]].width = width

    # Linhas de classificação com fórmulas COUNTIF / CONT.SE
    status_rows = [
        (6, "Conforme", "C", "COUNTIF('Check-List'!D13:D500, \"X\") + COUNTIF('Check-List'!D13:D500, \"x\")", "166534", "DCFCE7"),
        (7, "Não Conforme", "NC", "COUNTIF('Check-List'!E13:E500, \"X\") + COUNTIF('Check-List'!E13:E500, \"x\")", "991B1B", "FEE2E2"),
        (8, "Não Aplicável", "NA", "COUNTIF('Check-List'!F13:F500, \"X\") + COUNTIF('Check-List'!F13:F500, \"x\")", "374151", "F3F4F6"),
        (9, "Oportunidade de Melhoria", "OM", "COUNTIF('Check-List'!G13:G500, \"X\") + COUNTIF('Check-List'!G13:G500, \"x\")", "854D0E", "FEF9C3"),
    ]

    for row_idx, label, sigla, formula_count, text_col, bg_col in status_rows:
        c_label = ws_res[f'B{row_idx}']
        c_label.value = label
        c_label.font = Font(name="Segoe UI", size=10, bold=True)
        c_label.border = border_thin

        c_sigla = ws_res[f'C{row_idx}']
        c_sigla.value = sigla
        c_sigla.font = Font(name="Segoe UI", size=10, bold=True, color=text_col)
        c_sigla.fill = PatternFill(start_color=bg_col, end_color=bg_col, fill_type="solid")
        c_sigla.alignment = Alignment(horizontal="center", vertical="center")
        c_sigla.border = border_thin

        c_qtd = ws_res[f'D{row_idx}']
        c_qtd.value = f"={formula_count}"
        c_qtd.font = Font(name="Segoe UI", size=10, bold=True)
        c_qtd.alignment = Alignment(horizontal="center", vertical="center")
        c_qtd.border = border_thin

        c_pct = ws_res[f'E{row_idx}']
        # Percentual relativo à base (C + NC + OM)
        c_pct.value = f"=IF((D6+D7+D9)>0, D{row_idx}/(D6+D7+D9), 0)"
        c_pct.number_format = "0.0%"
        c_pct.font = Font(name="Segoe UI", size=10)
        c_pct.alignment = Alignment(horizontal="center", vertical="center")
        c_pct.border = border_thin

    # Linha Total Avaliados
    ws_res['B10'].value = "Total de Itens Auditados"
    ws_res['B10'].font = Font(name="Segoe UI", size=10, bold=True, color=navy_dark)
    ws_res['B10'].border = border_thin

    ws_res['C10'].value = "TOTAL"
    ws_res['C10'].font = Font(name="Segoe UI", size=9, bold=True, color=navy_dark)
    ws_res['C10'].alignment = Alignment(horizontal="center", vertical="center")
    ws_res['C10'].border = border_thin

    ws_res['D10'].value = "=COUNTA('Check-List'!B13:B500)"
    ws_res['D10'].font = Font(name="Segoe UI", size=10, bold=True, color=navy_dark)
    ws_res['D10'].alignment = Alignment(horizontal="center", vertical="center")
    ws_res['D10'].border = border_thin

    ws_res['E10'].value = "=SUM(E6:E9)"
    ws_res['E10'].number_format = "0.0%"
    ws_res['E10'].font = Font(name="Segoe UI", size=10, bold=True, color=navy_dark)
    ws_res['E10'].alignment = Alignment(horizontal="center", vertical="center")
    ws_res['E10'].border = border_thin

    # Card Índice de Conformidade
    ws_res.merge_cells('B12:E12')
    card_hdr = ws_res['B12']
    card_hdr.value = "ÍNDICE GERAL DE CONFORMIDADE"
    card_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    card_hdr.fill = PatternFill(start_color=navy_blue, end_color=navy_blue, fill_type="solid")
    card_hdr.alignment = Alignment(horizontal="center", vertical="center")

    ws_res.merge_cells('B13:E14')
    card_val = ws_res['B13']
    card_val.value = "=IF((D6+D7+D9)>0, D6/(D6+D7+D9), 0)"
    card_val.number_format = "0.0%"
    card_val.font = Font(name="Segoe UI", size=22, bold=True, color="166534")
    card_val.alignment = Alignment(horizontal="center", vertical="center")
    card_val.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    return wb


def process_and_add_excel_image(ws, img_obj, cell_coord, max_width=380, max_height=260) -> bool:
    """
    Processa imagem salva em base64 ou arquivo e a insere na planilha na célula especificada.
    """
    import base64
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from PIL import Image as PILImage
    except ImportError:
        return False

    try:
        pil_img = None
        if img_obj.arquivo_base64:
            raw_b64 = img_obj.arquivo_base64
            if "," in raw_b64:
                raw_b64 = raw_b64.split(",", 1)[1]
            img_bytes = base64.b64decode(raw_b64)
            pil_img = PILImage.open(io.BytesIO(img_bytes))
        elif img_obj.arquivo:
            try:
                pil_img = PILImage.open(img_obj.arquivo)
            except Exception:
                pass

        if not pil_img:
            return False

        # Converte para RGB se necessário
        if pil_img.mode in ("RGBA", "P", "LA"):
            pil_img = pil_img.convert("RGB")

        # Redimensiona proporcionalmente mantendo alta qualidade
        resample_method = getattr(PILImage, 'Resampling', PILImage).LANCZOS
        pil_img.thumbnail((max_width, max_height), resample_method)

        out_buffer = io.BytesIO()
        pil_img.save(out_buffer, format="JPEG", quality=85)
        out_buffer.seek(0)

        xl_img = OpenpyxlImage(out_buffer)
        xl_img.width, xl_img.height = pil_img.size
        ws.add_image(xl_img, cell_coord)
        return True
    except Exception:
        return False


def load_template_workbook() -> openpyxl.Workbook:
    """
    Carrega o template .xlsx se existir em disco;
    caso contrário (ou se estiver em ambiente read-only como Vercel),
    cria a estrutura base diretamente em memória.
    """
    template_path = get_template_path()
    if os.path.exists(template_path):
        try:
            return openpyxl.load_workbook(template_path, data_only=False)
        except Exception:
            pass

    wb = create_base_template_workbook()
    try:
        wb.save(template_path)
    except OSError:
        pass
    return wb


def generate_auditoria_excel_buffer(auditoria) -> io.BytesIO:
    """
    Lê o arquivo .xlsx de template e injeta os dados da auditoria nas células
    especificadas, preservando fórmulas, gráficos e estrutura da aba Resultados.
    """
    from ..models import (
        ItemNorma,
        RespostaEntrevistaIso,
        AvaliacaoFinalRequisitoIso,
    )

    wb = load_template_workbook()

    # Identifica a aba Check-List
    if "Check-List" in wb.sheetnames:
        ws = wb["Check-List"]
    else:
        ws = wb.worksheets[0]

    # -------------------------------------------------------------
    # 1. Metadados do Cabeçalho (Injeção Direta)
    # -------------------------------------------------------------
    unidade_nome = getattr(auditoria, 'empresa_auditada', '') or "Tecnolens"
    
    auditores_list = [a.get_full_name() or a.username for a in auditoria.auditores.all()]
    if not auditores_list and auditoria.abertura_auditores:
        auditores_str = auditoria.abertura_auditores
    else:
        auditores_str = ", ".join(auditores_list) if auditores_list else "Auditores Designados"

    dt_ini = auditoria.data_inicio.strftime('%d/%m/%Y') if auditoria.data_inicio else ""
    dt_fim = auditoria.data_fim.strftime('%d/%m/%Y') if auditoria.data_fim else ""
    data_str = f"{dt_ini} a {dt_fim}" if (dt_ini and dt_fim and dt_ini != dt_fim) else (dt_ini or dt_fim)

    escopo_str = f"{auditoria.norma.codigo} - {auditoria.norma.descricao}" if auditoria.norma.descricao else auditoria.norma.codigo

    ws['C5'].value = unidade_nome
    ws['C6'].value = auditores_str
    ws['C7'].value = data_str
    ws['C8'].value = escopo_str

    # Tipo de Auditoria (H6 = Presencial / H7 = Remota)
    tipo_auditoria = str(getattr(auditoria, 'tipo', '')).upper()
    if 'REMOT' in tipo_auditoria:
        ws['H6'].value = ""
        ws['H7'].value = "X"
    else:
        ws['H6'].value = "X"
        ws['H7'].value = ""

    # -------------------------------------------------------------
    # 2. Respostas da Auditoria (Loop Dinâmico a partir da Linha 13)
    # -------------------------------------------------------------
    agendas = list(auditoria.agendas.all().prefetch_related('perguntas', 'itens_norma', 'perguntas__itens_norma'))

    # Coleta todos os itens do escopo ou da norma
    itens_escopo = auditoria.escopo_itens.all()
    if not itens_escopo.exists():
        itens_escopo = ItemNorma.objects.filter(norma=auditoria.norma)
    itens_list = list(itens_escopo)

    # Identifica itens pais (que possuem subitens na hierarquia)
    parent_ids = set()
    for item in itens_list:
        prefix = item.referencia + '.'
        if any(other.referencia.startswith(prefix) for other in itens_list):
            parent_ids.add(item.id)

    respostas = RespostaEntrevistaIso.objects.filter(auditoria=auditoria).prefetch_related('solicitacoes', 'solicitacoes__imagens', 'pergunta__itens_norma')
    respostas_map = {r.pergunta_id: r for r in respostas}
    na_item_ids = set(auditoria.itens_nao_aplicaveis.values_list('id', flat=True))

    avaliacoes_finais_map = {
        av.item_norma_id: av
        for av in AvaliacaoFinalRequisitoIso.objects.filter(auditoria=auditoria)
    }

    # Mapas de vinculação de agendas/perguntas
    agenda_item_ids_map = {agenda.id: set(it.id for it in agenda.itens_norma.all()) for agenda in agendas}
    pergunta_item_ids_map = {}
    for agenda in agendas:
        for p in agenda.perguntas.all():
            if p.id not in pergunta_item_ids_map:
                pergunta_item_ids_map[p.id] = set(it.id for it in p.itens_norma.all())

    # Estilos para as células dos itens
    font_item = Font(name="Segoe UI", size=9, bold=True)
    font_text = Font(name="Segoe UI", size=9)
    font_x = Font(name="Segoe UI", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    gray_border = "CBD5E1"
    border_cell = Border(
        left=Side(style='thin', color=gray_border),
        right=Side(style='thin', color=gray_border),
        top=Side(style='thin', color=gray_border),
        bottom=Side(style='thin', color=gray_border)
    )

    current_row = 13  # Linha inicial dos itens de checklist

    for item in sorted(itens_list, key=lambda x: (x.ordem or 0, natural_sort_key(x.referencia))):
        is_parent = item.id in parent_ids

        # Perguntas e Solicitações vinculadas a este item
        perguntas_do_item = [p_id for p_id, item_ids in pergunta_item_ids_map.items() if item.id in item_ids]
        sols_do_item = []
        for p_id in perguntas_do_item:
            r = respostas_map.get(p_id)
            if r:
                for s in r.solicitacoes.all():
                    sols_do_item.append(s)

        # Determinação do Status / Classificação do Item
        av_final = avaliacoes_finais_map.get(item.id)
        if av_final and av_final.classificacao:
            status_item = av_final.classificacao
        elif is_parent:
            status_item = ""
        elif item.id in na_item_ids:
            status_item = "NA"
        elif sols_do_item:
            conclusoes = [s.conclusao for s in sols_do_item]
            if "NC" in conclusoes:
                status_item = "NC"
            elif "OM" in conclusoes:
                status_item = "OM"
            elif any(c in ["C", "OBS"] for c in conclusoes):
                status_item = "P" if all(c == "P" for c in conclusoes) else "C"
            elif all(c == "NA" for c in conclusoes):
                status_item = "NA"
            else:
                status_item = "P"
        elif perguntas_do_item:
            status_item = "C" if not is_parent else ""
        else:
            status_item = "NA" if not is_parent else ""

        # Compilação do Texto de Evidências (Col H)
        evidencias_textos = []
        for s in sols_do_item:
            sol_nome = (s.solicitacao or "").strip()
            evid_desc = (s.evidencia or "").strip()
            concl = s.get_conclusao_display()
            imgs_count = len(s.imagens.all()) if hasattr(s, 'imagens') else 0
            img_suffix = f" [📷 {imgs_count} foto(s)]" if imgs_count > 0 else ""
            
            if sol_nome and evid_desc:
                evidencias_textos.append(f"• [{concl}] {sol_nome}: {evid_desc}{img_suffix}")
            elif sol_nome:
                evidencias_textos.append(f"• [{concl}] {sol_nome}{img_suffix}")
            elif evid_desc:
                evidencias_textos.append(f"• [{concl}] {evid_desc}{img_suffix}")

        # Se não houver amostras registradas, incluir resposta/anotação geral se existir
        if not evidencias_textos:
            for p in perguntas_do_item:
                r = respostas_map.get(p.id)
                if r and r.texto_resposta and r.texto_resposta.strip():
                    evidencias_textos.append(f"• {r.texto_resposta.strip()}")

        evidencia_compilada = "\n".join(evidencias_textos)

        # Compilação de Observações / OBS / Justificativas (Col I)
        observacoes_list = []
        if av_final and av_final.justificativa:
            observacoes_list.append(f"[Revisão Final] {av_final.justificativa.strip()}")

        # Adiciona apontamentos com conclusão OBS (Observação com Correção)
        for s in sols_do_item:
            if s.conclusao == 'OBS':
                obs_detalhe = s.evidencia.strip() if (s.evidencia and s.evidencia.strip()) else (s.solicitacao.strip() if s.solicitacao else "")
                if obs_detalhe:
                    observacoes_list.append(f"[OBS com Correção] {obs_detalhe}")

        observacao_compilada = "\n".join(observacoes_list)

        # Injeção nas Células da Linha
        c_item = ws[f'B{current_row}']
        c_item.value = item.referencia
        c_item.font = font_item
        c_item.alignment = align_center
        c_item.border = border_cell

        c_questao = ws[f'C{current_row}']
        c_questao.value = item.titulo or item.descricao or ""
        c_questao.font = font_text
        c_questao.alignment = align_left_wrap
        c_questao.border = border_cell

        # Colunas D (C), E (NC), F (NA), G (OM)
        col_map = {"C": "D", "NC": "E", "NA": "F", "OM": "G"}
        for k, col_let in col_map.items():
            cell_k = ws[f'{col_let}{current_row}']
            cell_k.value = "X" if (status_item == k) else ""
            cell_k.font = font_x
            cell_k.alignment = align_center
            cell_k.border = border_cell

        c_evid = ws[f'H{current_row}']
        c_evid.value = evidencia_compilada
        c_evid.font = font_text
        c_evid.alignment = align_left_wrap
        c_evid.border = border_cell

        c_obs = ws[f'I{current_row}']
        c_obs.value = observacao_compilada
        c_obs.font = font_text
        c_obs.alignment = align_left_wrap
        c_obs.border = border_cell

        # Ajuste de altura automática da linha se houver múltiplas linhas de evidência
        num_lines = max(evidencia_compilada.count('\n') + 1, observacao_compilada.count('\n') + 1, 1)
        if num_lines > 1:
            ws.row_dimensions[current_row].height = max(20, num_lines * 16)
        else:
            ws.row_dimensions[current_row].height = 20

        current_row += 1

    # -------------------------------------------------------------
    # 2. ABA 'Evidências' (Injeção de Fotos e Imagens)
    # -------------------------------------------------------------
    if "Evidências" in wb.sheetnames:
        ws_evid = wb["Evidências"]
    elif "Evidencias" in wb.sheetnames:
        ws_evid = wb["Evidencias"]
    else:
        ws_evid = wb.create_sheet(title="Evidências", index=1)
        ws_evid.views.sheetView[0].showGridLines = True
        ws_evid.column_dimensions['A'].width = 4
        ws_evid.column_dimensions['B'].width = 18
        ws_evid.column_dimensions['C'].width = 38
        ws_evid.column_dimensions['D'].width = 38
        ws_evid.column_dimensions['E'].width = 38
        ws_evid.column_dimensions['F'].width = 38
        ws_evid.merge_cells('B3:F3')
        evid_hdr = ws_evid['B3']
        evid_hdr.value = "EVIDÊNCIAS COM IMAGENS"
        evid_hdr.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
        evid_hdr.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        evid_hdr.alignment = Alignment(horizontal="center", vertical="center")

    from ..models import ImagemSolicitacaoIso
    imagens_auditoria = list(
        ImagemSolicitacaoIso.objects.filter(
            solicitacao__resposta__auditoria=auditoria
        ).select_related(
            'solicitacao',
            'solicitacao__resposta',
            'solicitacao__resposta__pergunta'
        ).prefetch_related(
            'solicitacao__resposta__pergunta__itens_norma'
        ).order_by('criado_em')
    )

    # Índice de Requisitos na Coluna B (a partir da linha 4)
    ws_evid['B4'].value = "Item / Req."
    ws_evid['B4'].font = Font(name="Segoe UI", size=9, bold=True)
    ws_evid['B4'].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    ws_evid['B4'].alignment = Alignment(horizontal="center", vertical="center")
    ws_evid['B4'].border = border_cell

    b_row = 5
    itens_com_imagem = set()
    for img in imagens_auditoria:
        sol = img.solicitacao
        if sol and sol.resposta and sol.resposta.pergunta:
            for it in sol.resposta.pergunta.itens_norma.all():
                itens_com_imagem.add(it.referencia)

    for it in sorted(itens_list, key=lambda x: (x.ordem or 0, natural_sort_key(x.referencia)))[:35]:
        c_b = ws_evid[f'B{b_row}']
        c_b.value = it.referencia
        c_b.font = Font(name="Segoe UI", size=8.5, bold=(it.referencia in itens_com_imagem))
        c_b.alignment = Alignment(horizontal="center", vertical="center")
        c_b.border = border_cell
        if it.referencia in itens_com_imagem:
            c_b.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        b_row += 1

    if imagens_auditoria:
        evid_row = 5
        for idx, img in enumerate(imagens_auditoria, 1):
            sol = img.solicitacao
            pergunta = sol.resposta.pergunta if (sol and sol.resposta) else None
            itens_str = ", ".join([it.referencia for it in pergunta.itens_norma.all()]) if (pergunta and pergunta.itens_norma.exists()) else "Geral"
            concl_str = sol.get_conclusao_display() if sol else "Pendente"
            sol_texto = sol.solicitacao if sol else "Solicitação"

            # Cabeçalho do Card
            ws_evid.merge_cells(f'C{evid_row}:F{evid_row}')
            card_title = ws_evid[f'C{evid_row}']
            card_title.value = f"Evidência #{idx} | Requisito: {itens_str} — {sol_texto} [{concl_str}]"
            card_title.font = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
            card_title.fill = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
            card_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_evid.row_dimensions[evid_row].height = 22

            # Metadados da Imagem
            ws_evid.merge_cells(f'C{evid_row+1}:F{evid_row+1}')
            card_sub = ws_evid[f'C{evid_row+1}']
            legenda_txt = f"Legenda: {img.legenda}" if img.legenda else f"Arquivo: {img.nome_arquivo or 'foto_evidencia.jpg'}"
            data_txt = f" | Data: {img.criado_em.strftime('%d/%m/%Y %H:%M')}" if img.criado_em else ""
            card_sub.value = f"{legenda_txt}{data_txt}"
            card_sub.font = Font(name="Segoe UI", size=8.5, italic=True, color="475569")
            card_sub.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            card_sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_evid.row_dimensions[evid_row+1].height = 18

            # Inserção da Imagem
            img_coord = f'C{evid_row+2}'
            ws_evid.merge_cells(f'C{evid_row+2}:F{evid_row+2}')
            ws_evid.row_dimensions[evid_row+2].height = 160
            
            added = process_and_add_excel_image(ws_evid, img, img_coord, max_width=420, max_height=200)
            if not added:
                c_fallback = ws_evid[img_coord]
                c_fallback.value = "[Imagem anexada no sistema - visualização disponível no portal CalibraWeb]"
                c_fallback.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
                c_fallback.alignment = Alignment(horizontal="center", vertical="center")

            ws_evid.row_dimensions[evid_row+3].height = 10
            evid_row += 4
    else:
        ws_evid.merge_cells('C5:F7')
        msg_cell = ws_evid['C5']
        msg_cell.value = "Nenhuma foto de evidência foi anexada às solicitações desta auditoria até o momento.\nPara incluir registros fotográficos neste relatório, utilize o botão de anexo na tela de entrevista ou na matriz."
        msg_cell.font = Font(name="Segoe UI", size=9.5, italic=True, color="64748B")
        msg_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        msg_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
