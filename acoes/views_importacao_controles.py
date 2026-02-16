# -*- coding: utf-8 -*-
from datetime import datetime
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from acoes.forms import ImportacaoControleRegistrosForm
import unicodedata

from acoes.models import AcaoCorretiva, Solucao, PlanoAcao, LinhaAcao, KPIOpcao, TipoSolucao
from rh.models import Colaborador


TEMPLATE_HEADERS = [
    "Data de abertura",
    "Ano",
    "Unidade",
    "Nº do Registro",
    "Tipo de Solução",
    "Origem do Problema",
    "Descrição da NC e/ou Melhoria",
    "Causa Raiz",
    "Responsável",
    "Observação",
    "Link do registro",
    "Data de Fechamento Programada",
    "Data Fechamento",
    "Status",
]


HEADER_MAP = {
    "datadeabertura": "data_abertura",
    "ano": "ano",
    "unidade": "unidade",
    "ndoregistro": "numero_registro",
    "tipodesoluo": "tipo_solucao",
    "origemdoproblema": "origem",
    "descriodanceoumelhoria": "descricao",
    "causaraiz": "causa_raiz",
    "responsvel": "responsavel_matricula",
    "observao": "observacoes",
    "linkdoregistro": "link_registro",
    "datadefechamentoprogramada": "data_vencimento",
    "datafechamento": "data_conclusao",
    "status": "status",
}

SOLUCAO_TIPO_MAP = {
    "plano de acao": "plano_acao",
    "planoacao": "plano_acao",
    "plano_acao": "plano_acao",
    "a3": "a3",
    "8d": "8d",
    "rnc": "rnc",
    "gestao de mudanca": "gestao_mudanca",
    "gestaodemudanca": "gestao_mudanca",
    "gestao_mudanca": "gestao_mudanca",
    "revisao gerencial": "revisao_gerencial",
    "revisaogerencial": "revisao_gerencial",
    "revisao_gerencial": "revisao_gerencial",
}


SOLUCAO_STATUS_MAP = {
    "planejamento": "planejamento",
    "analise": "analise",
    "implementacao": "implementacao",
    "validacao": "validacao",
    "encerrada": "encerrada",
}


ACAO_STATUS_MAP = {
    "aberta": "aberta",
    "em progresso": "em_progresso",
    "em_progresso": "em_progresso",
    "concluida": "concluida",
    "cancelada": "cancelada",
}


ACAO_TIPO_MAP = {
    "corretiva": "corretiva",
    "preventiva": "preventiva",
}


ACAO_TIPO_SOLUCAO_MAP = {
    "corretiva": "corretiva",
    "preventiva": "preventiva",
    "melhoria": "melhoria",
}


ACAO_PRIORIDADE_MAP = {
    "baixa": "baixa",
    "media": "media",
    "alta": "alta",
    "critica": "critica",
}


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def build_tipo_solucao_lookup():
    lookup = {}
    for tipo in TipoSolucao.objects.filter(ativo=True):
        key = normalize_text(tipo.nome)
        if key:
            lookup[key] = tipo.nome
    return lookup


def map_tipo_solucao(value, lookup):
    if not value:
        return ""
    key = normalize_text(value)
    if not key:
        return ""

    map_padrao = {
        "plano de acao": "Plano de Ação",
        "planoacao": "Plano de Ação",
        "plano acao": "Plano de Ação",
        "a3": "A3",
        "8d": "8D",
        "rnc": "RNC",
        "gestao de mudanca": "Gestão de Mudança",
        "gestao mudanca": "Gestão de Mudança",
        "revisao gerencial": "Revisão Gerencial",
    }

    return lookup.get(key) or map_padrao.get(key) or str(value).strip()


def parse_bool(value, default=None):
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "sim", "s", "yes"}:
        return True
    if text in {"0", "false", "nao", "n", "no"}:
        return False
    return default


def parse_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def resolve_choice(value, mapping, default=None):
    if value is None or value == "":
        return default
    key = str(value).strip().lower()
    return mapping.get(key, default)


def get_colaborador_by_matricula(matricula):
    matricula = str(matricula).strip()
    if not matricula:
        return None
    return Colaborador.objects.filter(matricula=matricula).first()


def build_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Controles"
    ws.append(TEMPLATE_HEADERS)
    ws.append([
        "13/02/2026",
        2026,
        "Tecnolens",
        "PA-TEC-001/2026",
        "Melhoria",
        "Processo",
        "Implementar novo sistema de gestao",
        "Necessidade de otimizacao",
        "202",
        "Observacoes gerenciais",
        "https://tecnolens.sharepoint.com/...",
        "31/12/2026",
        "",
        "aberta",
    ])
    return wb


@login_required
@require_http_methods(["GET"])
def download_template_controle_registros(request):
    wb = build_template_workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=template_controle_registros.xlsx"
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET", "POST"])
def importar_controle_registros(request):
    if request.method == "GET":
        form = ImportacaoControleRegistrosForm()
        return render(request, "acoes/importar_controle_registros.html", {"form": form})

    form = ImportacaoControleRegistrosForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Envie uma planilha valida para importacao.")
        return redirect("acoes:importar_controle_registros")

    arquivo = request.FILES.get("arquivo_excel")
    if not arquivo:
        messages.error(request, "Nenhum arquivo foi enviado.")
        return redirect("acoes:importar_controle_registros")

    try:
        wb = load_workbook(filename=arquivo, data_only=True)
        ws = wb.active
    except Exception as exc:
        messages.error(request, f"Erro ao ler a planilha: {exc}")
        return redirect("acoes:importar_controle_registros")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "Planilha vazia.")
        return redirect("acoes:importar_controle_registros")

    header_row = rows[0]
    header_indexes = {}
    for idx, header in enumerate(header_row):
        normalized = normalize_header(header)
        if normalized in HEADER_MAP:
            header_indexes[HEADER_MAP[normalized]] = idx

    required_fields = [
        "numero_registro",
        "descricao",
        "data_vencimento",
    ]
    missing_required = [field for field in required_fields if field not in header_indexes]
    if missing_required:
        messages.error(
            request,
            "Colunas obrigatorias ausentes: " + ", ".join(missing_required),
        )
        return redirect("acoes:importar_controle_registros")

    created = 0
    updated = 0
    errors = []

    tipo_solucao_lookup = build_tipo_solucao_lookup()

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        def get_cell(field):
            col_idx = header_indexes.get(field)
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx]

        numero_registro = str(get_cell("numero_registro") or "").strip()
        if not numero_registro:
            errors.append(f"Linha {row_idx}: Numero Registro vazio")
            continue

        descricao = str(get_cell("descricao") or "").strip()
        data_vencimento = parse_date(get_cell("data_vencimento"))
        if not descricao or not data_vencimento:
            errors.append(
                f"Linha {row_idx}: Descricao e Data de Vencimento sao obrigatorios"
            )
            continue

        acao_defaults = {
            "descricao": descricao,
            "data_vencimento": data_vencimento,
            "data_abertura": parse_date(get_cell("data_abertura")),
            "unidade": get_cell("unidade") or "",
            "origem": get_cell("origem") or "",
            "causa_raiz": get_cell("causa_raiz") or "",
            "observacoes": get_cell("observacoes") or "",
            "link_registro": get_cell("link_registro") or "",
            "ano": get_cell("ano") or None,
            "tipo_solucao": map_tipo_solucao(get_cell("tipo_solucao"), tipo_solucao_lookup),
            "data_conclusao": parse_date(get_cell("data_conclusao")),
            "status": get_cell("status") or "aberta",
        }

        responsavel = get_colaborador_by_matricula(get_cell("responsavel_matricula"))
        acao_defaults["responsavel"] = responsavel

        acao, created_flag = AcaoCorretiva.objects.get_or_create(
            numero_registro=numero_registro,
            defaults=acao_defaults,
        )
        if created_flag:
            created += 1
        else:
            for field, value in acao_defaults.items():
                if value is not None and value != "":
                    setattr(acao, field, value)
            acao.save()
            updated += 1

    if errors:
        messages.warning(
            request,
            f"Importacao concluida com avisos. Criadas: {created}, Atualizadas: {updated}, Erros: {len(errors)}",
        )
        for erro in errors[:8]:
            messages.info(request, erro)
        if len(errors) > 8:
            messages.info(request, f"... e mais {len(errors) - 8} erros")
    else:
        messages.success(request, f"Importacao concluida! Criadas: {created}, Atualizadas: {updated}")

    return redirect("acoes:listar_solucoes")


# ============================================================================
# Segunda Importacao: PlanoAcao + LinhaAcao
# ============================================================================

PLANO_ACAO_TEMPLATE_HEADERS = [
    "N° de Registro",
    "Nº Ação",
    "Input",
    "Problema",
    "Lab",
    "KPI",
    "Descrição",
    "Classificação",
    "Status",
    "Prioridade",
    "Responsável",
    "1° Deadline",
    "Deadline",
    "COMENTÁRIOS",
    "Ação Eficaz",
]


PLANO_ACAO_HEADER_MAP = {
    "nderegistro": "numero_registro",
    "ndoregistro": "numero_registro",
    "numeroregistro": "numero_registro",
    "nacao": "numero_acao",
    "numeroacao": "numero_acao",
    "input": "input_origem",
    "inputorigem": "input_origem",
    "problema": "problema",
    "lab": "lab",
    "laboratorio": "lab",
    "kpi": "kpi",
    "descricao": "descricao",
    "classificacao": "classificacao",
    "status": "status",
    "prioridade": "prioridade",
    "responsavel": "responsavel",
    "responsavelmatricula": "responsavel",
    "responsaveismultiplosmatriculas": "responsaveis_multiplos",
    "responsavelexterno": "responsavel_externo",
    "responsavelexterno1": "responsavel_externo_1",
    "responsavelexterno2": "responsavel_externo_2",
    "responsavelexterno3": "responsavel_externo_3",
    "1deadline": "data_primeira_deadline",
    "dataprimeiradeadline": "data_primeira_deadline",
    "deadline": "data_deadline",
    "datadeadline": "data_deadline",
    "comentarios": "comentarios",
    "acaoeficaz": "acao_eficaz",
}


LINHA_ACAO_STATUS_MAP = {
    "planejada": "planejada",
    "em andamento": "em_andamento",
    "em_andamento": "em_andamento",
    "completa": "completa",
    "cancelada": "cancelada",
}


LINHA_ACAO_CLASSIFICACAO_MAP = {
    "corretiva": "corretiva",
    "preventiva": "preventiva",
    "melhoria": "melhoria",
}


ACOES_ASSOCIADAS_TEMPLATE_HEADERS = [
    "Numero Acao",
    "Input Origem",
    "KPI",
    "Problema",
    "Descricao",
    "Classificacao",
    "Status",
    "Prioridade",
    "Responsavel Matricula",
    "Responsavel Externo",
    "Data Primeira Deadline",
    "Data Deadline",
    "Comentarios",
    "Acao Eficaz",
    "Data Conclusao",
]


ACOES_ASSOCIADAS_HEADER_MAP = {
    "numeroacao": "numero_acao",
    "inputorigem": "input_origem",
    "kpi": "kpi",
    "problema": "problema",
    "descricao": "descricao",
    "classificacao": "classificacao",
    "status": "status",
    "prioridade": "prioridade",
    "responsavelmatricula": "responsavel_matricula",
    "responsaveismultiplosmatriculas": "responsaveis_multiplos",
    "responsavelexterno": "responsavel_externo",
    "responsavelexterno1": "responsavel_externo_1",
    "responsavelexterno2": "responsavel_externo_2",
    "responsavelexterno3": "responsavel_externo_3",
    "dataprimeiradeadline": "data_primeira_deadline",
    "datadeadline": "data_deadline",
    "comentarios": "comentarios",
    "acaoeficaz": "acao_eficaz",
    "dataconclusao": "data_conclusao",
}


ACOES_ASSOCIADAS_STATUS_MAP = {
    "planejada": "planejada",
    "em curso": "em_curso",
    "em_curso": "em_curso",
    "em andamento": "em_curso",
    "em_andamento": "em_curso",
    "retardo": "retardo",
    "atrasada": "retardo",
    "completa": "completa",
    "completa concluido": "completa",
    "completa concluida": "completa",
    "concluido": "completa",
    "concluida": "completa",
    "cancelada": "cancelada",
}


ACOES_ASSOCIADAS_EFICAZ_MAP = {
    "sim": "eficaz",
    "eficaz": "eficaz",
    "true": "eficaz",
    "1": "eficaz",
    "nao": "nao_eficaz",
    "nao eficaz": "nao_eficaz",
    "nao_eficaz": "nao_eficaz",
    "false": "nao_eficaz",
    "0": "nao_eficaz",
    "parcialmente eficaz": "parcialmente_eficaz",
    "parcialmene eficaz": "parcialmente_eficaz",
    "parcialmente_eficaz": "parcialmente_eficaz",
}


def parse_int(value):
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def build_plano_acao_template_workbook():
    """Cria workbook com aba 'Acoes' para importacao em massa."""
    wb = Workbook()
    ws_acoes = wb.active
    ws_acoes.title = "Acoes"
    ws_acoes.append(PLANO_ACAO_TEMPLATE_HEADERS)
    ws_acoes.append([
        "PA-TEC-001/2026",
        1,
        "Processo",
        "Falha no fluxo",
        "Lab Metrologia",
        "Tempo de ciclo",
        "Implantar novo processo",
        "corretiva",
        "planejada",
        "sim",
        "202",
        "2026-03-10",
        "2026-04-10",
        "Acao de exemplo",
        "EFICAZ",
    ])

    # Data validation para coluna "Ação Eficaz" (coluna 15 = O)
    dv_eficaz = DataValidation(
        type="list",
        formula1='"EFICAZ,NÃO EFICAZ,PARCIALMENTE EFICAZ"',
        allow_blank=True,
    )
    dv_eficaz.error = "Selecione uma opção válida"
    dv_eficaz.errorTitle = "Valor inválido"
    dv_eficaz.prompt = "Selecione a eficácia da ação"
    dv_eficaz.promptTitle = "Ação Eficaz"
    ws_acoes.add_data_validation(dv_eficaz)
    dv_eficaz.add(f"O2:O1048576")

    return wb


@login_required
@require_http_methods(["GET"])
def download_template_plano_acao(request):
    wb = build_plano_acao_template_workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=template_plano_acao.xlsx"
    wb.save(response)
    return response


def build_acoes_associadas_template_workbook():
    wb = Workbook()
    ws_acoes = wb.active
    ws_acoes.title = "Acoes"
    ws_acoes.append(ACOES_ASSOCIADAS_TEMPLATE_HEADERS)
    ws_acoes.append([
        1,
        "Processo",
        "Tempo de ciclo",
        "Falha no fluxo",
        "Implantar novo processo",
        "corretiva",
        "planejada",
        "sim",
        "202",
        "Fornecedor Alfa",
        "2026-03-10",
        "2026-04-10",
        "Acao de exemplo",
        "Parcialmente Eficaz",
        "",
    ])
    return wb


@login_required
@require_http_methods(["GET"])
def download_template_acoes_associadas(request, acao_id):
    wb = build_acoes_associadas_template_workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=template_acoes_associadas.xlsx"
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET", "POST"])
def importar_acoes_associadas(request, acao_id):
    acao = AcaoCorretiva.objects.filter(id=acao_id).first()
    if not acao:
        messages.error(request, "Acao corretiva nao encontrada.")
        return redirect("acoes:listar_acoes")

    if request.method == "GET":
        form = ImportacaoControleRegistrosForm()
        return render(
            request,
            "acoes/importar_acoes_associadas.html",
            {"form": form, "acao": acao},
        )

    form = ImportacaoControleRegistrosForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Envie uma planilha valida para importacao.")
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    arquivo = request.FILES.get("arquivo_excel")
    if not arquivo:
        messages.error(request, "Nenhum arquivo foi enviado.")
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    try:
        wb = load_workbook(filename=arquivo, data_only=True)
    except Exception as exc:
        messages.error(request, f"Erro ao ler a planilha: {exc}")
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    ws = None
    for sheet_name in wb.sheetnames:
        if normalize_header(sheet_name) == normalize_header("Acoes"):
            ws = wb[sheet_name]
            break

    if not ws:
        messages.error(request, "Planilha nao contem aba 'Acoes'.")
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "Aba 'Acoes' vazia.")
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    header_row = rows[0]
    header_indexes = {}
    for idx, header in enumerate(header_row):
        normalized = normalize_header(header)
        if normalized in ACOES_ASSOCIADAS_HEADER_MAP:
            header_indexes[ACOES_ASSOCIADAS_HEADER_MAP[normalized]] = idx

    required_fields = ["numero_acao", "descricao"]
    missing_required = [field for field in required_fields if field not in header_indexes]
    if missing_required:
        messages.error(
            request,
            "Colunas obrigatorias ausentes na aba 'Acoes': " + ", ".join(missing_required),
        )
        return redirect("acoes:importar_acoes_associadas", acao_id=acao.id)

    solucao, _ = Solucao.objects.get_or_create(
        acao_corretiva=acao,
        tipo="plano_acao",
        defaults={
            "titulo": f"Plano de Acao - {acao.numero_registro or acao.titulo}",
            "descricao": acao.descricao or "",
            "responsavel": acao.responsavel,
            "status": "planejamento",
        },
    )
    plano_acao, _ = PlanoAcao.objects.get_or_create(solucao=solucao)

    criadas = 0
    atualizadas = 0
    errors = []
    linhas_map = {}

    def get_cell_from_row(row_values, field):
        col_idx = header_indexes.get(field)
        if col_idx is None or col_idx >= len(row_values):
            return None
        return row_values[col_idx]

    def merge_if_empty(target, field, value):
        if value is None or value == "":
            return
        if target.get(field) in (None, ""):
            target[field] = value

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        numero_acao = parse_int(get_cell_from_row(row, "numero_acao"))
        descricao = str(get_cell_from_row(row, "descricao") or "").strip()
        if not numero_acao or not descricao:
            errors.append(f"Linha {row_idx}: Numero Acao e Descricao sao obrigatorios")
            continue

        status_key = normalize_text(get_cell_from_row(row, "status"))
        status = ACOES_ASSOCIADAS_STATUS_MAP.get(status_key, "planejada")
        acao_eficaz_raw = normalize_text(get_cell_from_row(row, "acao_eficaz"))
        acao_eficaz = ACOES_ASSOCIADAS_EFICAZ_MAP.get(acao_eficaz_raw)

        base_defaults = {
            "input_origem": get_cell_from_row(row, "input_origem") or "",
            "kpi": get_cell_from_row(row, "kpi") or "",
            "problema": get_cell_from_row(row, "problema") or "",
            "descricao": descricao,
            "classificacao": resolve_choice(
                get_cell_from_row(row, "classificacao"), LINHA_ACAO_CLASSIFICACAO_MAP, "corretiva"
            ),
            "status": status,
            "prioridade": parse_bool(get_cell_from_row(row, "prioridade"), False),
            "data_primeira_deadline": parse_date(get_cell_from_row(row, "data_primeira_deadline")),
            "data_deadline": parse_date(get_cell_from_row(row, "data_deadline")),
            "comentarios": get_cell_from_row(row, "comentarios") or "",
            "acao_eficaz": acao_eficaz,
            "data_conclusao": parse_date(get_cell_from_row(row, "data_conclusao")),
        }

        if numero_acao not in linhas_map:
            linhas_map[numero_acao] = {
                "defaults": base_defaults,
                "responsaveis_ids": set(),
                "responsavel_principal": None,
                "externos": [],
            }
        else:
            for field, value in base_defaults.items():
                merge_if_empty(linhas_map[numero_acao]["defaults"], field, value)

        matricula_raw = str(get_cell_from_row(row, "responsavel_matricula") or "").strip()
        responsaveis_raw = str(get_cell_from_row(row, "responsaveis_multiplos") or "").strip()
        externo_raw = str(get_cell_from_row(row, "responsavel_externo") or "").strip()

        if matricula_raw:
            responsavel = get_colaborador_by_matricula(matricula_raw)
            if not responsavel:
                errors.append(f"Linha {row_idx}: Responsavel matricula nao encontrado")
            else:
                linhas_map[numero_acao]["responsaveis_ids"].add(responsavel.id)
                if not linhas_map[numero_acao]["responsavel_principal"]:
                    linhas_map[numero_acao]["responsavel_principal"] = responsavel

            if responsaveis_raw:
                matriculas = [
                    item.strip() for item in re.split(r"[;,]", responsaveis_raw) if item.strip()
                ]
                for matricula in matriculas:
                    colaborador = get_colaborador_by_matricula(matricula)
                    if not colaborador:
                        errors.append(
                            f"Linha {row_idx}: Responsavel multiplo matricula '{matricula}' nao encontrado"
                        )
                        continue
                    linhas_map[numero_acao]["responsaveis_ids"].add(colaborador.id)
        else:
            if responsaveis_raw:
                linhas_map[numero_acao]["externos"].append(responsaveis_raw)

        if externo_raw:
            linhas_map[numero_acao]["externos"].append(externo_raw)

        for field in ("responsavel_externo_1", "responsavel_externo_2", "responsavel_externo_3"):
            value = str(get_cell_from_row(row, field) or "").strip()
            if value:
                linhas_map[numero_acao]["externos"].append(value)

    for numero_acao, payload in linhas_map.items():
        linha_defaults = payload["defaults"]
        if payload["responsavel_principal"]:
            linha_defaults["responsavel_acao"] = payload["responsavel_principal"]
        if payload["externos"]:
            externos_unique = []
            for item in payload["externos"]:
                if item not in externos_unique:
                    externos_unique.append(item)
            linha_defaults["responsaveis_externos"] = "; ".join(externos_unique)

        linha, created_flag = LinhaAcao.objects.get_or_create(
            plano_acao=plano_acao,
            numero_acao=numero_acao,
            defaults=linha_defaults,
        )
        if created_flag:
            criadas += 1
        else:
            for field, value in linha_defaults.items():
                if value is not None and value != "":
                    setattr(linha, field, value)
            linha.save()
            atualizadas += 1

        if payload["responsaveis_ids"]:
            linha.responsaveis_multiplos.set(list(payload["responsaveis_ids"]))

    if errors:
        messages.warning(
            request,
            f"Importacao concluida com avisos. Criadas: {criadas}, Atualizadas: {atualizadas}, Erros: {len(errors)}",
        )
        for erro in errors[:8]:
            messages.info(request, erro)
        if len(errors) > 8:
            messages.info(request, f"... e mais {len(errors) - 8} erros")
    else:
        messages.success(
            request, f"Importacao concluida! Criadas: {criadas}, Atualizadas: {atualizadas}"
        )

    return redirect("acoes:detalhe_acao", acao_id=acao.id)


@login_required
@require_http_methods(["GET"])
def exportar_acoes_associadas(request, acao_id):
    acao = AcaoCorretiva.objects.filter(id=acao_id).first()
    if not acao:
        messages.error(request, "Acao corretiva nao encontrada.")
        return redirect("acoes:listar_acoes")

    solucao_plano = Solucao.objects.filter(acao_corretiva=acao, tipo="plano_acao").first()
    plano_acao = solucao_plano.plano_acao if solucao_plano else None
    acoes = LinhaAcao.objects.none()
    if plano_acao:
        acoes = (
            LinhaAcao.objects.filter(plano_acao=plano_acao)
            .select_related("responsavel_acao")
            .prefetch_related("responsaveis_multiplos")
            .order_by("numero_acao")
        )

    def get_matricula(colaborador):
        if not colaborador:
            return ""
        return str(getattr(colaborador, "matricula", "") or "").strip()

    acao_eficaz_label = {
        "eficaz": "Eficaz",
        "nao_eficaz": "Não Eficaz",
        "parcialmente_eficaz": "Parcialmente Eficaz",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Acoes"
    ws.append(ACOES_ASSOCIADAS_TEMPLATE_HEADERS)

    def append_row(linha, responsavel_matricula, responsavel_externo):
        ws.append([
            linha.numero_acao,
            linha.input_origem or "",
            linha.kpi or "",
            linha.problema or "",
            linha.descricao or "",
            linha.classificacao or "",
            linha.status or "",
            "sim" if linha.prioridade else "nao",
            responsavel_matricula,
            responsavel_externo,
            linha.data_primeira_deadline.isoformat() if linha.data_primeira_deadline else "",
            linha.data_deadline.isoformat() if linha.data_deadline else "",
            linha.comentarios or "",
            acao_eficaz_label.get(linha.acao_eficaz, ""),
            linha.data_conclusao.isoformat() if linha.data_conclusao else "",
        ])

    for linha in acoes:
        internos = []
        if linha.responsavel_acao:
            internos.append(linha.responsavel_acao)
        internos.extend(list(linha.responsaveis_multiplos.all()))

        externos = []
        if linha.responsaveis_externos:
            externos = [
                item.strip()
                for item in re.split(r"[;|]", linha.responsaveis_externos)
                if item.strip()
            ]

        if internos:
            for resp in internos:
                append_row(linha, get_matricula(resp), "")

        if externos:
            for ext in externos:
                append_row(linha, "", ext)

        if not internos and not externos:
            append_row(linha, "", "")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = f"attachment; filename=acoes_associadas_{acao.numero_registro or acao.id}.xlsx"
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET", "POST"])
def importar_plano_acao(request):
    """
    Importa LinhaAcao em massa para Solucoes existentes.
    Aba 'Acoes' contem as linhas de acao com Numero do Registro.
    """
    if request.method == "GET":
        form = ImportacaoControleRegistrosForm()
        return render(request, "acoes/importar_plano_acao.html", {"form": form})

    form = ImportacaoControleRegistrosForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Envie uma planilha valida para importacao.")
        return redirect("acoes:importar_plano_acao")

    arquivo = request.FILES.get("arquivo_excel")
    if not arquivo:
        messages.error(request, "Nenhum arquivo foi enviado.")
        return redirect("acoes:importar_plano_acao")

    try:
        wb = load_workbook(filename=arquivo, data_only=True)
    except Exception as exc:
        messages.error(request, f"Erro ao ler a planilha: {exc}")
        return redirect("acoes:importar_plano_acao")

    # Procura pela aba "Acoes"
    ws = None
    for sheet_name in wb.sheetnames:
        if normalize_header(sheet_name) == normalize_header("Acoes"):
            ws = wb[sheet_name]
            break
    
    if not ws:
        messages.error(request, "Planilha nao contem aba 'Acoes'.")
        return redirect("acoes:importar_plano_acao")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "Aba 'Acoes' vazia.")
        return redirect("acoes:importar_plano_acao")

    header_row = rows[0]
    header_indexes = {}
    for idx, header in enumerate(header_row):
        normalized = normalize_header(header)
        if normalized in PLANO_ACAO_HEADER_MAP:
            header_indexes[PLANO_ACAO_HEADER_MAP[normalized]] = idx

    required_fields = [
        "numero_registro",
        "numero_acao",
        "descricao",
    ]
    missing_required = [field for field in required_fields if field not in header_indexes]
    if missing_required:
        messages.error(
            request,
            "Colunas obrigatorias ausentes na aba 'Acoes': " + ", ".join(missing_required),
        )
        return redirect("acoes:importar_plano_acao")

    criadas = 0
    atualizadas = 0
    errors = []
    planos_map = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        def get_cell(field):
            col_idx = header_indexes.get(field)
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx]

        numero_registro = str(get_cell("numero_registro") or "").strip()
        if not numero_registro:
            errors.append(f"Linha {row_idx} (Acoes): N° do Registro obrigatorio")
            continue

        acao = AcaoCorretiva.objects.filter(numero_registro__iexact=numero_registro).first()
        if not acao:
            errors.append(
                f"Linha {row_idx} (Acoes): Registro '{numero_registro}' nao encontrado"
            )
            continue

        if acao.id not in planos_map:
            solucao, _ = Solucao.objects.get_or_create(
                acao_corretiva=acao,
                tipo="plano_acao",
                defaults={
                    "titulo": f"Plano de Acao - {acao.numero_registro or acao.titulo}",
                    "descricao": acao.descricao or "",
                    "responsavel": acao.responsavel,
                    "status": "planejamento",
                },
            )
            plano_acao, _ = PlanoAcao.objects.get_or_create(solucao=solucao)
            planos_map[acao.id] = {
                "acao": acao,
                "plano_acao": plano_acao,
                "linhas_map": {},
            }

        plano_payload = planos_map[acao.id]
        linhas_map = plano_payload["linhas_map"]

        # Atualiza campo Lab no PlanoAcao se fornecido
        lab_value = str(get_cell("lab") or "").strip()
        if lab_value:
            plano_acao_obj = plano_payload["plano_acao"]
            if not plano_acao_obj.laboratorio or plano_acao_obj.laboratorio != lab_value:
                plano_acao_obj.laboratorio = lab_value
                plano_acao_obj.save(update_fields=["laboratorio"])

        numero_acao = parse_int(get_cell("numero_acao"))
        descricao = str(get_cell("descricao") or "").strip()
        if not numero_acao or not descricao:
            errors.append(f"Linha {row_idx} (Acoes): Numero Acao e Descricao sao obrigatorios")
            continue

        status_key = normalize_text(get_cell("status"))
        status = ACOES_ASSOCIADAS_STATUS_MAP.get(status_key, "planejada")
        acao_eficaz_raw = normalize_text(get_cell("acao_eficaz"))
        acao_eficaz = ACOES_ASSOCIADAS_EFICAZ_MAP.get(acao_eficaz_raw)

        base_defaults = {
            "input_origem": get_cell("input_origem") or "",
            "kpi": get_cell("kpi") or "",
            "problema": get_cell("problema") or "",
            "descricao": descricao,
            "classificacao": resolve_choice(
                get_cell("classificacao"), LINHA_ACAO_CLASSIFICACAO_MAP, "corretiva"
            ),
            "status": status,
            "prioridade": parse_bool(get_cell("prioridade"), False),
            "data_primeira_deadline": parse_date(get_cell("data_primeira_deadline")),
            "data_deadline": parse_date(get_cell("data_deadline")),
            "comentarios": get_cell("comentarios") or "",
            "acao_eficaz": acao_eficaz,
        }

        if numero_acao not in linhas_map:
            linhas_map[numero_acao] = {
                "defaults": base_defaults,
                "responsaveis_ids": set(),
                "responsavel_principal": None,
                "externos": [],
            }
        else:
            for field, value in base_defaults.items():
                merge_if_empty(linhas_map[numero_acao]["defaults"], field, value)

        # Coluna unificada "Responsavel": tenta matricula, senao trata como externo
        responsavel_raw = str(get_cell("responsavel") or "").strip()
        responsaveis_raw = str(get_cell("responsaveis_multiplos") or "").strip()
        externo_raw = str(get_cell("responsavel_externo") or "").strip()

        if responsavel_raw:
            responsavel = get_colaborador_by_matricula(responsavel_raw)
            if responsavel:
                linhas_map[numero_acao]["responsaveis_ids"].add(responsavel.id)
                if not linhas_map[numero_acao]["responsavel_principal"]:
                    linhas_map[numero_acao]["responsavel_principal"] = responsavel
            else:
                # Nao encontrou como matricula, trata como responsavel externo
                linhas_map[numero_acao]["externos"].append(responsavel_raw)

        if responsaveis_raw:
            matriculas = [
                item.strip() for item in re.split(r"[;,]", responsaveis_raw) if item.strip()
            ]
            for matricula in matriculas:
                colaborador = get_colaborador_by_matricula(matricula)
                if colaborador:
                    linhas_map[numero_acao]["responsaveis_ids"].add(colaborador.id)
                else:
                    linhas_map[numero_acao]["externos"].append(matricula)

        if externo_raw:
            linhas_map[numero_acao]["externos"].append(externo_raw)

        for field in ("responsavel_externo_1", "responsavel_externo_2", "responsavel_externo_3"):
            value = str(get_cell(field) or "").strip()
            if value:
                linhas_map[numero_acao]["externos"].append(value)

    for payload in planos_map.values():
        plano_acao = payload["plano_acao"]
        linhas_map = payload["linhas_map"]
        for numero_acao, linha_payload in linhas_map.items():
            linha_defaults = linha_payload["defaults"]
            if linha_payload["responsavel_principal"]:
                linha_defaults["responsavel_acao"] = linha_payload["responsavel_principal"]
            if linha_payload["externos"]:
                externos_unique = []
                for item in linha_payload["externos"]:
                    if item not in externos_unique:
                        externos_unique.append(item)
                linha_defaults["responsaveis_externos"] = "; ".join(externos_unique)

            linha, created_flag = LinhaAcao.objects.get_or_create(
                plano_acao=plano_acao,
                numero_acao=numero_acao,
                defaults=linha_defaults,
            )
            if created_flag:
                criadas += 1
            else:
                for field, value in linha_defaults.items():
                    if value is not None and value != "":
                        setattr(linha, field, value)
                linha.save()
                atualizadas += 1

            if linha_payload["responsaveis_ids"]:
                linha.responsaveis_multiplos.set(list(linha_payload["responsaveis_ids"]))

    if errors:
        messages.warning(
            request,
            f"Importacao concluida com avisos. Criadas: {criadas}, Atualizadas: {atualizadas}, Erros: {len(errors)}",
        )
        for erro in errors[:8]:
            messages.info(request, erro)
        if len(errors) > 8:
            messages.info(request, f"... e mais {len(errors) - 8} erros")
    else:
        messages.success(request, f"Importacao concluida! Criadas: {criadas}, Atualizadas: {atualizadas}")

    return redirect("acoes:listar_solucoes")


@login_required
@require_http_methods(["POST"])
def deletar_acoes_associadas(request, acao_id):
    """
    Deleta múltiplas LinhaAcao associadas a uma AcaoCorretiva.
    Recebe lista de IDs via POST['ids'].
    """
    print(f"\n{'='*60}")
    print(f"DELETAR ACOES ASSOCIADAS - View chamada")
    print(f"{'='*60}")
    print(f"acao_id: {acao_id}")
    print(f"request.method: {request.method}")
    print(f"request.POST: {dict(request.POST)}")
    
    try:
        acao = AcaoCorretiva.objects.get(id=acao_id)
        print(f"✓ AcaoCorretiva encontrada: {acao}")
    except AcaoCorretiva.DoesNotExist:
        print(f"✗ AcaoCorretiva não encontrada com ID {acao_id}")
        messages.error(request, "Ação Corretiva não encontrada.")
        return redirect("acoes:listar_acoes")

    ids = request.POST.getlist('ids')
    print(f"IDs recebidos: {ids} (total: {len(ids)})")
    
    if not ids:
        print(f"✗ Nenhum ID foi enviado")
        messages.warning(request, "Nenhuma ação foi selecionada para deletar.")
        return redirect("acoes:detalhe_acao", acao_id=acao_id)

    try:
        # PlanoAcao está relacionado com Solucao, não diretamente com AcaoCorretiva
        # Precisamos fazer: AcaoCorretiva -> Solucao -> PlanoAcao
        plano_acao = PlanoAcao.objects.filter(solucao__acao_corretiva=acao).first()
        if not plano_acao:
            print(f"✗ PlanoAcao não encontrado para a ação {acao_id}")
            messages.error(request, "Plano de Ação não encontrado para esta Ação Corretiva.")
            return redirect("acoes:detalhe_acao", acao_id=acao_id)
        
        print(f"✓ PlanoAcao encontrado: {plano_acao.id}")

        # Filter only the LinhaAcao that belong to this PlanoAcao
        linhas_to_delete = LinhaAcao.objects.filter(id__in=ids, plano_acao=plano_acao)
        count = linhas_to_delete.count()
        print(f"LinhaAcao encontradas para deletar: {count}")
        
        if count > 0:
            for linha in linhas_to_delete:
                print(f"  - Deletando LinhaAcao #{linha.id}: {linha.numero_acao}")

        if count == 0:
            print(f"✗ Nenhuma LinhaAcao válida encontrada")
            messages.warning(request, "Nenhuma ação válida encontrada para deletar.")
        else:
            linhas_to_delete.delete()
            print(f"✓ {count} LinhaAcao(ões) deletada(s) com sucesso")
            messages.success(request, f"{count} ação(ões) deletada(s) com sucesso.")

    except Exception as exc:
        print(f"✗ ERRO durante deleção: {exc}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Erro ao deletar ações: {exc}")

    print(f"Redirecionando para detalhe_acao...")
    print(f"{'='*60}\n")
    return redirect("acoes:detalhe_acao", acao_id=acao_id)
